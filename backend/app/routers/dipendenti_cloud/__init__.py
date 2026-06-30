"""
Dipendenti in Cloud - Router Module
Sistema HR completo per gestione personale
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Body
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
import re
import os
import io
import zipfile
import hashlib
import base64
import tempfile
from datetime import datetime, timezone, timedelta

from backend.app.database import Database

# Router principale
router = APIRouter(prefix="/dipendenti-cloud", tags=["Dipendenti Cloud"])

# ============ HELPERS ============

def get_db():
    """Get database instance"""
    return Database.get_db()

def generate_id():
    return str(uuid.uuid4())

def now_iso():
    return datetime.now(timezone.utc).isoformat()

def serialize_doc(doc):
    if doc and '_id' in doc:
        del doc['_id']
    return doc

# ============ MODELS ============

class DipendenteCloud(BaseModel):
    nome: str
    cognome: str
    matricola: Optional[str] = None
    codice_fiscale: Optional[str] = None
    data_nascita: Optional[str] = None
    email: Optional[str] = None
    telefono: Optional[str] = None
    indirizzo: Optional[str] = None
    ruolo: Optional[str] = None
    luogo_lavoro: Optional[str] = None
    contratto: str = "Indeterminato"
    data_assunzione: Optional[str] = None
    data_fine_contratto: Optional[str] = None
    iban: Optional[str] = None
    stato: str = "attivo"

class PresenzaCloud(BaseModel):
    dipendente_id: str
    data: str
    entrata: Optional[str] = None
    uscita: Optional[str] = None
    stato: str = "presente"
    giustificativo: Optional[str] = None
    ore_lavorate: float = 0
    note: Optional[str] = None

class FerieCloud(BaseModel):
    dipendente_id: str
    tipo: str  # Ferie, Permesso, Malattia, ROL
    data_inizio: str
    data_fine: str
    giorni: int = 1
    stato: str = "in_attesa"
    nota: Optional[str] = None

class TurnoCloud(BaseModel):
    nome: str
    orario_inizio: str
    orario_fine: str
    colore: str = "#3b82f6"

class BustaPagaCloud(BaseModel):
    dipendente_id: str
    mese: int
    anno: int
    lordo: float
    netto: float
    inps: float = 0
    irpef: float = 0
    trattenute: float = 0
    stato: str = "DA_PAGARE"
    data_pagamento: Optional[str] = None

class MissioneCloud(BaseModel):
    dipendente_id: str
    destinazione: str
    data_inizio: str
    data_fine: str
    scopo: str
    rimborso: float = 0
    stato: str = "in_attesa"

class DocumentoCloud(BaseModel):
    dipendente_id: str
    titolo: str
    tipo: str
    scadenza: Optional[str] = None
    file_url: Optional[str] = None

# ============ DIPENDENTI ============

@router.get("/dipendenti")
async def get_dipendenti():
    """Legge dalla collezione 'dipendenti' esistente nel database Gestionale"""
    dipendenti = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    # Normalizza i campi per compatibilità con il frontend
    result = []
    for d in dipendenti:
        result.append({
            "id": d.get("id") or str(d.get("_id", "")),
            "nome": d.get("nome", ""),
            "cognome": d.get("cognome", ""),
            "codice_fiscale": d.get("codice_fiscale", ""),
            "stato": d.get("stato", "attivo"),
            "ruolo": d.get("ruolo", ""),
            "iban": d.get("iban", ""),
            "email": d.get("email", ""),
            "telefono": d.get("telefono", ""),
            "contratto": d.get("contratto", "Indeterminato"),
            "data_assunzione": d.get("data_assunzione", ""),
            "luogo_lavoro": d.get("luogo_lavoro", ""),
            "importo_stipendio": d.get("importo_stipendio", 0),
            "created_at": d.get("created_at", "")
        })
    return result

@router.get("/dipendenti/{dipendente_id}")
async def get_dipendente(dipendente_id: str):
    dip = await get_db().dipendenti.find_one({"id": dipendente_id}, {"_id": 0})
    if not dip:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    return dip

@router.post("/dipendenti")
async def create_dipendente(dip: DipendenteCloud):
    dip_dict = dip.model_dump()
    dip_dict["id"] = generate_id()
    dip_dict["created_at"] = now_iso()
    await get_db().dipendenti.insert_one(dip_dict)
    return serialize_doc(dip_dict)

@router.put("/dipendenti/{dipendente_id}")
async def update_dipendente(dipendente_id: str, dip: DipendenteCloud):
    result = await get_db().dipendenti.update_one(
        {"id": dipendente_id},
        {"$set": dip.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    return {"message": "Dipendente aggiornato"}

@router.delete("/dipendenti/{dipendente_id}")
async def delete_dipendente(dipendente_id: str):
    result = await get_db().dipendenti.delete_one({"id": dipendente_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    return {"message": "Dipendente eliminato"}

@router.post("/dipendenti/{dipendente_id}/cessa")
async def cessa_dipendente(dipendente_id: str, data: dict = Body(default={})):
    """Cessa il rapporto: aggiorna lo stato e innesca l'iter completo di chiusura
    (termina contratti, rifiuta assenze future, annulla partite, risolve alert)
    tramite l'evento DIPENDENTE_CESSATO già agganciato all'handler dedicato."""
    db = get_db()
    dip = await db.dipendenti.find_one({"id": dipendente_id}, {"_id": 0})
    if not dip:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    data_cessazione = (data.get("data_cessazione") or now_iso()[:10])
    nome = dip.get("nome_completo") or f"{dip.get('cognome','')} {dip.get('nome','')}".strip()
    await db.dipendenti.update_one({"id": dipendente_id}, {"$set": {
        "stato": "cessato", "attivo": False,
        "data_dimissione": data_cessazione, "cessato_il": now_iso(),
        "motivo_cessazione": data.get("motivo") or "cessazione_manuale",
    }})
    try:
        from backend.app.services.event_bus import propagate_event, EventTypes
        risultati = await propagate_event(EventTypes.DIPENDENTE_CESSATO, {
            "dipendente_id": dipendente_id, "nome_completo": nome,
            "data_cessazione": data_cessazione,
        }, db, source_module="gestione", user="admin")
    except Exception as e:
        risultati = [{"error": str(e)}]
    return {"ok": True, "stato": "cessato", "data_cessazione": data_cessazione, "automazioni": risultati}

@router.get("/ordine-dipendenti")
async def get_ordine_dipendenti():
    doc = await get_db().dipendenti_ordine.find_one({"id": "ordine"}, {"_id": 0})
    return {"ordine": (doc or {}).get("lista", [])}

@router.post("/ordine-dipendenti")
async def set_ordine_dipendenti(data: dict):
    lista = data.get("ordine", [])
    await get_db().dipendenti_ordine.update_one(
        {"id": "ordine"}, {"$set": {"lista": lista}}, upsert=True)
    return {"ok": True}

# ============ PAGHE MENSILI (importo busta + bonifico + acconti) ============

@router.get("/paghe")
async def get_paghe(anno: int, mese: int):
    return await get_db().paghe_mensili.find(
        {"anno": int(anno), "mese": int(mese)}, {"_id": 0}).to_list(500)

@router.post("/paghe")
async def upsert_pagha(data: dict):
    dip = data.get("dipendente_id"); anno = data.get("anno"); mese = data.get("mese")
    if not dip or not anno or not mese:
        raise HTTPException(status_code=400, detail="dipendente_id, anno, mese obbligatori")
    # Normalizza gli acconti: massimo 3, solo importo+data
    acconti = []
    for a in (data.get("acconti") or [])[:3]:
        acconti.append({"importo": a.get("importo"), "data": a.get("data")})
    doc = {
        "dipendente_id": dip, "anno": int(anno), "mese": int(mese),
        "importo_busta": data.get("importo_busta"),
        "bonifico_ricevuto": bool(data.get("bonifico_ricevuto", False)),
        "bonifico_importo": data.get("bonifico_importo"),
        "bonifico_data": data.get("bonifico_data"),
        "acconti": acconti,
        "updated_at": now_iso(),
    }
    await get_db().paghe_mensili.update_one(
        {"dipendente_id": dip, "anno": int(anno), "mese": int(mese)},
        {"$set": doc}, upsert=True)
    await _ricalcola_stato_paga(get_db(), dip, int(anno), int(mese))
    return {"ok": True, "pagha": doc}


async def _ricalcola_stato_paga(db, dip, anno, mese):
    """MOTORE UNICO buste↔bonifici. Aggancia i pagamenti bancari già arrivati
    (pagamenti_esiti) come bonifico del mese e ricalcola lo stato:
    in_attesa_pagamento (busta senza pagamento) / parziale / pagato / vuoto.
    Chiamato da OGNI ingresso (busta da LUL/email, prima nota, CSV, modifica manuale),
    così il popolamento di un dato aggiorna automaticamente gli altri."""
    anno, mese = int(anno), int(mese)
    p = await db.paghe_mensili.find_one({"dipendente_id": dip, "anno": anno, "mese": mese})
    if not p:
        return None
    tot_esiti, n_esiti = 0.0, 0
    async for e in db.pagamenti_esiti.find({"dipendente_id": dip, "mese": mese, "anno": anno}, {"_id": 0, "importo": 1}):
        tot_esiti += e.get("importo") or 0
        n_esiti += 1
    bonifico = round(tot_esiti, 2) if n_esiti else float(p.get("bonifico_importo") or 0)
    busta = float(p.get("importo_busta") or 0)
    acc = sum(float(a.get("importo") or 0) for a in (p.get("acconti") or []))
    erogato = bonifico + acc
    if busta <= 0 and erogato <= 0:
        stato = "vuoto"
    elif erogato <= 0:
        stato = "in_attesa_pagamento"
    elif erogato + 0.5 >= busta:
        stato = "pagato"
    else:
        stato = "parziale"
    upd = {"stato_pagamento": stato, "saldo": round(busta - erogato, 2), "updated_at": now_iso()}
    if n_esiti:
        upd["bonifico_importo"] = bonifico
        upd["bonifico_ricevuto"] = bonifico > 0
    await db.paghe_mensili.update_one({"dipendente_id": dip, "anno": anno, "mese": mese}, {"$set": upd})
    return stato

@router.delete("/paghe")
async def delete_pagha(dipendente_id: str, anno: int, mese: int):
    res = await get_db().paghe_mensili.delete_one(
        {"dipendente_id": dipendente_id, "anno": int(anno), "mese": int(mese)})
    return {"ok": True, "eliminati": res.deleted_count}


@router.post("/paghe/importa-excel-salari")
async def importa_excel_salari(file: UploadFile = File(...)):
    """Importa l'Excel 'prima nota salari' (colonne: DIPENDENTE, MESE, ANNO,
    STIPENDIO NETTO, IMPORTO EROGATO). Il netto fissa il valore atteso della busta,
    l'erogato il valore atteso del bonifico. I flag di riconciliazione partono a False
    e diventano True quando arriva il PDF (busta o ricevuta bonifico) con importo che
    combacia. I dipendenti non presenti in anagrafica vengono solo segnalati."""
    import openpyxl
    nome_file = (file.filename or "").lower()
    if not nome_file.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Serve un file Excel (.xlsx)")

    _MESI_IT = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
                "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12}

    def _num(x):
        if x in (None, ""):
            return None
        if isinstance(x, (int, float)):
            return float(x)
        try:
            return float(str(x).strip().replace(".", "").replace(",", "."))
        except Exception:
            return None

    data = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Excel non leggibile: {e}")
    ws = wb.active

    dips = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    anag = {}
    for d in dips:
        cg = (d.get("cognome") or "").upper().strip()
        nm = (d.get("nome") or "").upper().strip()
        if cg or nm:
            anag[f"{cg} {nm}".strip()] = d   # Cognome Nome
            anag[f"{nm} {cg}".strip()] = d   # Nome Cognome (ordine invertito)

    try:
        await get_db().paghe_mensili.create_index(
            [("dipendente_id", 1), ("anno", 1), ("mese", 1)], unique=True, name="uniq_dip_anno_mese")
    except Exception:
        pass

    anno_corrente = datetime.now(timezone.utc).year
    importati, mesi_set = 0, set()
    non_trovati, scartati = {}, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        nome = str(row[0]).strip()
        mese = _MESI_IT.get(str(row[1]).strip().lower()) if len(row) > 1 and row[1] else None
        try:
            anno = int(row[2]) if len(row) > 2 and row[2] else None
        except Exception:
            anno = None
        netto = _num(row[3]) if len(row) > 3 else None
        erogato = _num(row[4]) if len(row) > 4 else None

        dip = anag.get(nome.upper())
        if not dip:
            non_trovati[nome] = non_trovati.get(nome, 0) + 1
            continue
        if not mese or not anno or anno not in ANNI_AMMESSI:
            scartati.append({"nome": nome, "motivo": f"periodo non valido ({row[1]} {row[2]})"})
            continue

        set_doc = {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                   "fonte_excel": True, "updated_at": now_iso()}
        if netto is not None:
            set_doc["importo_busta"] = netto
            set_doc["netto_atteso"] = netto
        if erogato is not None:
            set_doc["bonifico_importo"] = erogato
            set_doc["erogato_atteso"] = erogato
        await get_db().paghe_mensili.update_one(
            {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
            {"$set": set_doc,
             "$setOnInsert": {"busta_riconciliata": False, "bonifico_riconciliato": False}},
            upsert=True)
        importati += 1
        mesi_set.add((anno, mese))

    mesi = sorted([{"anno": y, "mese": m} for (y, m) in mesi_set], key=lambda x: (x["anno"], x["mese"]))
    return {"importati": importati,
            "mesi": mesi,
            "dipendenti_non_in_anagrafica": [{"nome": k, "righe": v} for k, v in sorted(non_trovati.items())],
            "scartati": scartati}


# --- Import automatico Libro Unico (PDF) → divide per dipendente e memorizza i netti ---

_CF_RE = re.compile(r'\b([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])\b')
_MESI = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
         "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12}
# Regola atomica: si importano SOLO questi anni. Tutto il resto è bloccato.
ANNI_AMMESSI = {2023, 2024, 2025, 2026}

def _lul_netto(text):
    m = re.findall(r'([\d]{1,3}(?:\.\d{3})*,\d{2})\s*€', text)
    return m[-1] if m else None

def _lul_acconto(text):
    """Rileva acconti/anticipi erogati durante il mese e trattenuti nel cedolino
    (righe con 'acconto', 'anticipo', 'rec. acconto', escluso il TFR). Serve a sapere
    quanto è già stato dato, così il bonifico del solo saldo chiude comunque la busta.
    Ritorna l'importo totale o None."""
    tot = 0.0
    for line in text.split("\n"):
        low = line.lower()
        if ("acconto" in low or "anticipo" in low) and "tfr" not in low and "trattamento fine" not in low:
            nums = re.findall(r'([\d]{1,3}(?:\.\d{3})*,\d{2})', line)
            if nums:
                tot += _to_float(nums[-1]) or 0
    return round(tot, 2) if tot > 0 else None

def _lul_periodo(text):
    m = re.search(r'(Gennaio|Febbraio|Marzo|Aprile|Maggio|Giugno|Luglio|Agosto|Settembre|Ottobre|Novembre|Dicembre)\s+(\d{4})', text, re.I)
    if m:
        return _MESI[m.group(1).lower()], int(m.group(2))
    return None, None

_LUL_NUM = re.compile(r'-?\d{1,3}(?:\.\d{3})*,\d{2,6}|-?\d+,\d{2,6}')


def _lul_dati_busta(text: str) -> dict:
    """Estrae dal testo della busta i dati chiave (per codice voce o descrizione).
    Robusto sul prefisso (C/F/Z…). L'ultimo numero della riga voce = importo competenza."""
    voci = []
    voci_obj = []
    for line in text.split("\n"):
        m = re.match(r'^\s*([A-Z]\d{4,5})\b\s*(.*)$', line)
        if m:
            resto, valori = m.group(2), _LUL_NUM.findall(m.group(2))
            voci.append((m.group(1), resto, valori))
            voci_obj.append({"codice": m.group(1),
                             "descrizione": _LUL_NUM.split(resto)[0].strip(' .-'),
                             "valori": valori})

    def find(codici=None, testo=None):
        for codice, resto, valori in voci:
            if (codici and codice in codici) or (testo and testo.lower() in resto.lower()):
                return valori[-1] if valori else None
        return None

    dati = {
        "rateo_13ma": find(codici={"C50000", "Z50000"}, testo="13ma Mensilit"),
        "rateo_14ma": find(codici={"C50022", "Z50022"}, testo="14ma Mensilit"),
        "indennita_l207_24": find(codici={"F02703"}),
        "indennita_l207_24_cng_ann": find(codici={"F09088"}),
        "tratt_integrativo_l21": find(codici={"F09081"}),
        "tratt_integrativo_l21_rata": find(codici={"F09083"}),
        "tratt_integrativo_l21_cng": find(codici={"F09084"}),
        # tutte le voci del cedolino (codici+descrizione+importi) per il motore di ricerca
        "voci": voci_obj or None,
    }
    # Rimborso da 730 (residuo + importo del mese)
    for codice, resto, valori in voci:
        if "730" in resto:
            dati["rimborso_730"] = valori[-1] if valori else None
            if len(valori) >= 2:
                dati["rimborso_730_residuo"] = valori[0]
            break
    # Ore lavorate + giorni retribuiti (riquadro 'Lavorato', best effort)
    lav = re.search(r'(?:Lavorato|Ore\s*lavorat\w*)\D{0,15}?(\d{1,3},\d{2})\s+(\d{1,2})\b', text, re.IGNORECASE)
    if lav:
        dati["ore_lavorate"] = lav.group(1)
        dati["giorni_retribuiti"] = lav.group(2)
    # Giorni effettivamente lavorati: righe del foglio presenze con ore (LU 19 6,40 ...)
    gg = set()
    for line in text.split("\n"):
        pm = re.search(r'\b(LU|MA|ME|GI|VE|SA|DO)\s+(\d{1,2})\s+\d{1,2},\d{2}\b', line)
        if pm:
            gg.add(pm.group(2))
    if gg:
        dati["giorni_lavorati"] = len(gg)
    return {k: v for k, v in dati.items() if v is not None}


def _parse_lul(pdf_path):
    """Raggruppa le pagine per codice fiscale (gestisce 1, 2 o 3 pagine a dipendente).
    Tiene anche traccia degli indici di pagina di ciascun dipendente, così l'import
    può ritagliare il PDF reale del suo cedolino."""
    import pdfplumber
    ced = {}
    with pdfplumber.open(pdf_path) as pdf:
        cur = None
        for idx, page in enumerate(pdf.pages):
            t = page.extract_text() or ""
            cfs = _CF_RE.findall(t)
            mese, anno = _lul_periodo(t)
            if cfs:
                cur = cfs[0]
                d = ced.setdefault(cur, {"nome": None, "netto": None, "mese": None, "anno": None, "pagine": []})
                if mese:
                    d["mese"], d["anno"] = mese, anno
                for line in t.split("\n"):
                    mm = re.search(r'\b0[0-9]{6}\b\s+([A-ZÀ-Ù\' ]{4,}?)\s+[A-Z]{6}\d{2}[A-Z]', line)
                    if mm:
                        d["nome"] = mm.group(1).strip()
                        break
            if cur:
                ced[cur].setdefault("pagine", []).append(idx)
                n = _lul_netto(t)
                if n:
                    ced[cur]["netto"] = n
                acc = _lul_acconto(t)
                if acc:
                    ced[cur]["acconto"] = round((ced[cur].get("acconto") or 0) + acc, 2)
                if not ced[cur].get("mese") and mese:
                    ced[cur]["mese"], ced[cur]["anno"] = mese, anno
                # Dati chiave della busta (rateo 13/14, indennità L.207/24, tratt. integ. L.21, giorni)
                for k, v in _lul_dati_busta(t).items():
                    ced[cur][k] = v
    return ced

def _to_float(s):
    return float(s.replace(".", "").replace(",", ".")) if s else None


def _ritaglia_pdf(pdf_path, pagine):
    """Estrae le pagine indicate dal PDF originale e le restituisce come bytes:
    è il cedolino reale del singolo dipendente dentro il Libro Unico."""
    import fitz
    src = fitz.open(pdf_path)
    out = fitz.open()
    for i in sorted(set(pagine)):
        if 0 <= i < src.page_count:
            out.insert_pdf(src, from_page=i, to_page=i)
    data = out.tobytes()
    out.close(); src.close()
    return data

def _estrai_testo(pdf_path):
    import pdfplumber
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            parts.append(page.extract_text() or "")
    return "\n".join(parts)

def _classifica_doc(text):
    """Distingue: bonifico (ricevuta bancaria), presenze (LUL ore/timbrature),
    cedolino (busta paga con netto). Default: cedolino (per il Libro Unico multi-dipendente)."""
    T = (text or "").upper()
    if "RICEVUTA PER ORDINANTE" in T or "A VOSTRO DEBITO A FAVORE DI" in T or ("BONIFICO" in T and "IBAN BENEFICIARIO" in T):
        return "bonifico"
    ha_netto = "NETTO DEL MESE" in T or "NETTOSDELSMESE" in T
    if ha_netto:
        return "cedolino"
    if ("PERIODO DI RIFERIMENTO" in T or "TIMBRATURE" in T or "ORE ORDINARIE" in T):
        return "presenze"
    return "cedolino"

def _competenza_da_causale(causale):
    """Estrae mese/anno SOLO se dichiarati esplicitamente nella causale."""
    c = (causale or "").lower()
    m = re.search(r'(gennaio|febbraio|marzo|aprile|maggio|giugno|luglio|agosto|settembre|ottobre|novembre|dicembre)\s*(\d{4})?', c)
    if m:
        return _MESI[m.group(1)], (int(m.group(2)) if m.group(2) else None), True
    m = re.search(r'\b(0?[1-9]|1[0-2])[-/](\d{4})\b', c)
    if m:
        return int(m.group(1)), int(m.group(2)), True
    return None, None, False

def _parse_bonifico(text):
    imp = None
    m = re.search(r'EUR\s+([\d.]+,\d{2})', text) or re.search(r'IMPORTO\s+([\d.]+,\d{2})', text)
    if m:
        imp = _to_float(m.group(1))
    data = None
    md = re.search(r'DATA\s+(\d{2})/(\d{2})/(\d{4})', text)
    if md:
        data = f"{md.group(3)}-{md.group(2)}-{md.group(1)}"
    caus = None
    mc = re.search(r'CAUSALE\s*\n\s*([^\n]+)', text)
    if mc:
        caus = mc.group(1).strip()
    cro = None
    mr = re.search(r'(MB0B\w+)', text)
    if mr:
        cro = mr.group(1).strip()
    mese_c, anno_c, esplicita = _competenza_da_causale(caus)
    is_tfr = bool(re.search(r'\btfr\b|trattamento fine rapporto|anticipo\s+t\.?f\.?r', (caus or "").lower()))
    return {"importo": imp, "data": data, "causale": caus, "cro": cro,
            "mese_causale": mese_c, "anno_causale": anno_c, "esplicita": esplicita, "is_tfr": is_tfr}


async def _importa_documenti(pdf_items, errori_iniziali=None, forza=False):
    """Pipeline condivisa: riceve una lista di (origine, pdf_bytes) già espansi (da upload
    file o da posta elettronica), li classifica e li importa in paghe_mensili / prestiti.
    L'anti-duplicazione per hash evita di re-importare gli stessi documenti."""
    dips = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    by_cf = {(d.get("codice_fiscale") or "").upper(): d for d in dips if d.get("codice_fiscale")}
    by_nome = {}
    for d in dips:
        cg = (d.get("cognome") or "").upper().strip()
        nm = (d.get("nome") or "").upper().strip()
        if cg or nm:
            by_nome[f"{cg} {nm}".strip()] = d
            by_nome[f"{nm} {cg}".strip()] = d

    # Vincolo: una sola busta per (dipendente, anno, mese) — i duplicati diventano impossibili
    try:
        await get_db().paghe_mensili.create_index(
            [("dipendente_id", 1), ("anno", 1), ("mese", 1)], unique=True, name="uniq_dip_anno_mese")
    except Exception:
        pass
    # Registro documenti importati (anti-duplicazione): impronta del file + chiave logica
    try:
        await get_db().documenti_importati.create_index([("hash", 1)], unique=True, name="uniq_hash")
        await get_db().documenti_importati.create_index([("chiave", 1)], name="idx_chiave")
    except Exception:
        pass

    async def _registra_doc(h, tipo, chiave, origine):
        try:
            await get_db().documenti_importati.update_one(
                {"hash": h},
                {"$set": {"hash": h, "tipo": tipo, "chiave": chiave, "file": origine,
                          "imported_at": now_iso()}}, upsert=True)
        except Exception:
            pass

    async def _imputa_competenza(dip_id, b):
        """Determina (mese, anno, fonte) di competenza del bonifico secondo le regole:
        1) mese esplicito in causale; 2) match per importo con la busta (acconto=busta o
        somma cumulativa=busta) nella finestra mese precedente→mese stesso; 3) ripiego sul
        mese precedente. Sfondamento d'anno (gen→dic anno prima) solo dal 2024 (2023 blindato)."""
        if b["esplicita"] and b["mese_causale"]:
            anno = b["anno_causale"] or (int(b["data"][:4]) if b.get("data") else None)
            return b["mese_causale"], anno, "causale"
        data = b.get("data")
        if not data:
            return None, None, "data assente"
        y, mo = int(data[:4]), int(data[5:7])
        pm, py = (mo - 1, y) if mo > 1 else (12, y - 1)
        finestra = []
        if not (mo == 1 and py < 2023):   # 2023 blindato: gennaio 2023 non sfonda a dic 2022
            finestra.append((py, pm))     # mese precedente (priorità)
        finestra.append((y, mo))          # mese stesso
        for (a, m) in finestra:
            rec = await get_db().paghe_mensili.find_one(
                {"dipendente_id": dip_id, "anno": a, "mese": m})
            if not rec:
                continue
            busta = rec.get("importo_busta") or rec.get("netto_atteso")
            if busta:
                if abs(busta - b["importo"]) <= 1:
                    return m, a, "importo (= busta)"
                # acconto già dato nel mese (rilevato nel cedolino o bonifico precedente):
                # acconto + questo bonifico = busta  ->  saldo che chiude la busta
                gia = (rec.get("bonifico_importo") or 0) + (rec.get("acconto_cedolino") or 0)
                if abs((gia + b["importo"]) - busta) <= 1:
                    return m, a, "importo (acconto+saldo = busta)"
                # il bonifico copre esattamente il saldo residuo dopo l'acconto
                residuo = rec.get("saldo_residuo")
                if residuo and abs(residuo - b["importo"]) <= 1:
                    return m, a, "importo (= saldo dopo acconto)"
        a, m = finestra[0]
        return m, a, "mese precedente (dedotta)"

    async def _processa_pdf(pdfbytes, origine):
        with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp:
            tmp.write(pdfbytes)
            path = tmp.name
        ass, dac, bon, pres, dup, tfr, prestiti = [], [], [], [], [], [], []
        # Anti-duplicazione 1: stesso file già importato (impronta del contenuto)
        h = hashlib.sha256(pdfbytes).hexdigest()
        try:
            if not forza and await get_db().documenti_importati.find_one({"hash": h}):
                dup.append({"file": origine, "motivo": "documento già importato (stesso file)"})
                try: os.unlink(path)
                except Exception: pass
                return ass, dac, bon, pres, dup, tfr, prestiti
        except Exception:
            pass
        try:
            text = _estrai_testo(path)
            tipo = _classifica_doc(text)

            # ---- BONIFICO (ricevuta bancaria) ----
            if tipo == "bonifico":
                b = _parse_bonifico(text)
                # match dipendente: "COGNOME NOME" presente nel testo; fallback cognome nella causale
                T = text.upper()
                dip = None
                for cand in dips:
                    cg = (cand.get("cognome") or "").upper().strip()
                    nm = (cand.get("nome") or "").upper().strip()
                    if cg and nm and f"{cg} {nm}" in T:
                        dip = cand; break
                if not dip:
                    cau = (b.get("causale") or "").upper()
                    for cand in dips:
                        cg = (cand.get("cognome") or "").upper().strip()
                        if cg and cg in cau:
                            dip = cand; break
                manca = []
                if not dip: manca.append("dipendente non riconosciuto")
                if not b.get("importo"): manca.append("importo")
                if manca:
                    dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                "motivo": "bonifico: " + ", ".join(manca)})
                else:
                    caus_low = (b.get("causale") or "").lower()
                    if "prestito" in caus_low:
                        # PRESTITO: non imputare a buste paga; mastrino prestiti con saldo progressivo
                        data = b.get("data")
                        if not data:
                            dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                        "motivo": "prestito: data assente"})
                        elif int(data[:4]) not in ANNI_AMMESSI:
                            dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                        "motivo": f"anno {data[:4]} non ammesso — bloccato (solo 2023-2026)"})
                        else:
                            pa, pm = int(data[:4]), int(data[5:7])
                            cro = b.get("cro")
                            gia = await get_db().documenti_importati.find_one({"chiave": f"cro:{cro}"}) if cro else None
                            if gia:
                                dup.append({"file": origine, "motivo": f"prestito già importato (CRO {cro})"})
                            else:
                                await get_db().prestiti_dipendenti.insert_one({
                                    "id": str(uuid.uuid4()), "dipendente_id": dip["id"],
                                    "importo": b["importo"], "data": data, "mese": pm, "anno": pa,
                                    "causale": b.get("causale"), "cro": cro, "pdf": origine,
                                    "created_at": now_iso()})
                                saldo = await _ricalcola_saldo_prestiti(dip["id"])
                                await _registra_doc(h, "prestito",
                                    f"cro:{cro}" if cro else f"pre:{dip['id']}:{pa}:{pm}:{b['importo']}", origine)
                                prestiti.append({"dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                                                 "importo": b["importo"], "mese": pm, "anno": pa,
                                                 "data": data, "saldo": saldo})
                        return ass, dac, bon, pres, dup, tfr, prestiti
                    mese, anno, fonte = await _imputa_competenza(dip["id"], b)
                    if not mese or not anno:
                        dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                    "motivo": "bonifico: competenza non determinabile"})
                    elif anno not in ANNI_AMMESSI:
                        dac.append({"nome": (b.get("causale") or "?")[:30], "origine": origine,
                                    "motivo": f"anno {anno} non ammesso — bloccato (solo 2023-2026)"})
                    else:
                        cro = b.get("cro")
                        gia = await get_db().documenti_importati.find_one({"chiave": f"cro:{cro}"}) if cro else None
                        if gia:
                            dup.append({"file": origine, "motivo": f"bonifico già importato (CRO {cro})"})
                        elif b.get("is_tfr"):
                            # Anticipo TFR: fuori dal saldo stipendi
                            await get_db().paghe_mensili.update_one(
                                {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                                {"$set": {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                                          "tfr_anticipo_importo": b["importo"], "tfr_anticipo_data": b.get("data"),
                                          "tfr_anticipo_pdf": origine, "updated_at": now_iso()},
                                 "$setOnInsert": {"busta_riconciliata": False, "bonifico_riconciliato": False}},
                                upsert=True)
                            await _registra_doc(h, "tfr", f"cro:{cro}" if cro else f"tfr:{dip['id']}:{anno}:{mese}", origine)
                            tfr.append({"dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                                        "importo": b["importo"], "mese": mese, "anno": anno, "data": b.get("data")})
                        else:
                            esist = await get_db().paghe_mensili.find_one(
                                {"dipendente_id": dip["id"], "anno": anno, "mese": mese}, {"erogato_atteso": 1})
                            atteso = (esist or {}).get("erogato_atteso")
                            discrep = atteso if (atteso is not None and abs(atteso - b["importo"]) > 1) else None
                            set_doc = {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                                       "bonifico_importo": b["importo"], "bonifico_data": b.get("data"),
                                       "bonifico_ricevuto": True, "bonifico_causale": b.get("causale"),
                                       "bonifico_cro": cro, "bonifico_pdf": origine,
                                       "bonifico_riconciliato": True, "updated_at": now_iso()}
                            await get_db().paghe_mensili.update_one(
                                {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                                {"$set": set_doc, "$setOnInsert": {"busta_riconciliata": False}}, upsert=True)
                            await _registra_doc(h, "bonifico", f"cro:{cro}" if cro else f"bon:{dip['id']}:{anno}:{mese}", origine)
                            bon.append({"dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                                        "importo": b["importo"], "mese": mese, "anno": anno,
                                        "causale": b.get("causale"), "data": b.get("data"),
                                        "riconciliato": True, "discrepanza": discrep, "fonte": fonte})
                return ass, dac, bon, pres, dup, tfr, prestiti

            # ---- FOGLIO PRESENZE (ore/timbrature, non è una busta) ----
            if tipo == "presenze":
                cf = (_CF_RE.findall(text) or [None])[0]
                mese, anno = _lul_periodo(text)
                if anno and anno not in ANNI_AMMESSI:
                    dac.append({"nome": cf or "?", "origine": origine,
                                "motivo": f"presenze anno {anno} non ammesso — bloccato (solo 2023-2026)"})
                    return ass, dac, bon, pres, dup, tfr, prestiti
                dip = by_cf.get((cf or "").upper())
                await _registra_doc(h, "presenze", f"pres:{cf}:{anno}:{mese}", origine)
                pres.append({"dipendente": (f"{dip.get('cognome')} {dip.get('nome')}".strip() if dip else (cf or "?")),
                             "mese": mese, "anno": anno, "origine": origine})
                return ass, dac, bon, pres, dup, tfr, prestiti

            # ---- CEDOLINO / LIBRO UNICO multi-dipendente (netti) ----
            ced = _parse_lul(path)
            for cf, info in ced.items():
                dip = by_cf.get(cf)
                metodo = "codice fiscale"
                if not dip:
                    dip = by_nome.get((info.get("nome") or "").upper())
                    metodo = "nome (CF non combacia)"
                netto = _to_float(info.get("netto"))
                mese, anno = info.get("mese"), info.get("anno")
                if not dip or not mese:
                    dac.append({"nome": info.get("nome"), "cf": cf, "netto": netto, "origine": origine,
                                "motivo": "dipendente non trovato" if not dip else "periodo non rilevato"})
                    continue
                if not netto or netto <= 0:
                    dac.append({"nome": info.get("nome"), "cf": cf, "netto": netto, "origine": origine,
                                "motivo": "netto non rilevato (non salvato)"})
                    continue
                if anno not in ANNI_AMMESSI:
                    dac.append({"nome": info.get("nome"), "cf": cf, "netto": netto, "origine": origine,
                                "motivo": f"anno {anno} non ammesso — bloccato (solo 2023-2026)"})
                    continue
                esistente = await get_db().paghe_mensili.find_one(
                    {"dipendente_id": dip["id"], "anno": anno, "mese": mese}, {"netto_atteso": 1})
                atteso = (esistente or {}).get("netto_atteso")
                discrep = atteso if (atteso is not None and abs(atteso - netto) > 1) else None
                acconto = info.get("acconto")
                set_doc = {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                           "importo_busta": netto, "busta_da_lul": True,
                           "busta_riconciliata": True, "updated_at": now_iso()}
                if acconto and acconto > 0:
                    set_doc["acconto_cedolino"] = acconto
                    set_doc["saldo_residuo"] = round(netto - acconto, 2)
                await get_db().paghe_mensili.update_one(
                    {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                    {"$set": set_doc}, upsert=True)
                # Motore unico: busta arrivata → aggancia il pagamento o la mette in attesa
                await _ricalcola_stato_paga(get_db(), dip["id"], anno, mese)
                # Cedolino (fonte del portale): salvo il PDF REALE ritagliato dal Libro
                # Unico + il netto, così il dipendente scarica la sua busta vera.
                ced_set = {"dipendente_id": dip["id"], "anno": anno, "mese": mese,
                           "netto": netto,
                           "dipendente_nome": f"{dip.get('cognome','')} {dip.get('nome','')}".strip(),
                           "updated_at": now_iso()}
                if acconto and acconto > 0:
                    ced_set["acconto_cedolino"] = acconto
                    ced_set["saldo_residuo"] = round(netto - acconto, 2)
                # Dati chiave estratti dalla busta (salvati nel cedolino)
                for k in ("rateo_13ma", "rateo_14ma", "indennita_l207_24",
                          "indennita_l207_24_cng_ann", "tratt_integrativo_l21",
                          "tratt_integrativo_l21_rata", "tratt_integrativo_l21_cng",
                          "rimborso_730", "rimborso_730_residuo",
                          "ore_lavorate", "giorni_retribuiti", "giorni_lavorati", "voci"):
                    if info.get(k) is not None:
                        ced_set[k] = info[k]
                try:
                    if info.get("pagine"):
                        ced_set["pdf_data"] = base64.b64encode(_ritaglia_pdf(path, info["pagine"])).decode()
                        ced_set["filename"] = f"busta_{anno}_{str(mese).zfill(2)}.pdf"
                except Exception:
                    pass
                await get_db().cedolini.update_one(
                    {"dipendente_id": dip["id"], "anno": anno, "mese": mese},
                    {"$set": ced_set,
                     "$setOnInsert": {"id": generate_id(), "created_at": now_iso(), "stato": "importato"}},
                    upsert=True)
                ass.append({"dipendente_id": dip["id"],
                            "dipendente": f"{dip.get('cognome')} {dip.get('nome')}".strip(),
                            "netto": netto, "metodo": metodo, "mese": mese, "anno": anno,
                            "riconciliata": True, "discrepanza": discrep,
                            "acconto": acconto, "saldo_residuo": (round(netto - acconto, 2) if acconto else None)})
            if ass:
                await _registra_doc(h, "cedolino", f"file:{origine}", origine)
        finally:
            try:
                os.unlink(path)
            except Exception:
                pass
        return ass, dac, bon, pres, dup, tfr, prestiti

    associati, da_controllare, errori, bonifici, presenze, duplicati, tfr_list, prestiti_list = [], [], [], [], [], [], [], []
    errori = list(errori_iniziali or [])
    file_pdf = 0
    for (nome, data) in pdf_items:
        try:
            a, d, b, p, du, tf, pr = await _processa_pdf(data, nome)
            associati += a; da_controllare += d; bonifici += b; presenze += p; duplicati += du; tfr_list += tf; prestiti_list += pr; file_pdf += 1
        except Exception as e:
            errori.append(f"{nome}: {e}")

    if file_pdf == 0:
        raise HTTPException(status_code=400, detail="Nessun PDF elaborabile. " + ("; ".join(errori) if errori else ""))

    # Dedup: se lo stesso dipendente/mese è arrivato da più file, tieni una riga sola
    visti = {}
    for a in associati:
        visti[(a["dipendente_id"], a["anno"], a["mese"])] = a
    associati = list(visti.values())

    mesi_set = sorted({(a["anno"], a["mese"]) for a in associati})
    mesi = [{"anno": y, "mese": m, "n": sum(1 for a in associati if a["anno"] == y and a["mese"] == m)}
            for (y, m) in mesi_set]
    associati.sort(key=lambda x: (x["anno"], x["mese"], x["dipendente"]))
    bonifici.sort(key=lambda x: (x["anno"], x["mese"], x["dipendente"]))
    return {"associati": associati, "da_controllare": da_controllare,
            "totale_associati": len(associati), "file_pdf": file_pdf,
            "mesi": mesi, "errori": errori,
            "bonifici": bonifici, "presenze": presenze, "duplicati": duplicati, "tfr": tfr_list, "prestiti": prestiti_list}


async def _ricalcola_saldo_prestiti(dip_id):
    """Riporto continuo: azzera i campi prestito da tutti i mesi del dipendente, poi somma
    i movimenti in ordine cronologico riscrivendo erogato del mese e saldo cumulativo.
    Ritorna il saldo totale corrente."""
    await get_db().paghe_mensili.update_many(
        {"dipendente_id": dip_id},
        {"$unset": {"prestito_importo": "", "prestito_saldo": ""}})
    movs = await get_db().prestiti_dipendenti.find({"dipendente_id": dip_id}).to_list(2000)
    erog = {}
    for mv in movs:
        k = (mv["anno"], mv["mese"])
        erog[k] = erog.get(k, 0) + (mv.get("importo") or 0)
    saldo = 0
    for (a, m) in sorted(erog.keys()):
        saldo += erog[(a, m)]
        await get_db().paghe_mensili.update_one(
            {"dipendente_id": dip_id, "anno": a, "mese": m},
            {"$set": {"dipendente_id": dip_id, "anno": a, "mese": m,
                      "prestito_importo": erog[(a, m)], "prestito_saldo": saldo,
                      "updated_at": now_iso()},
             "$setOnInsert": {"busta_riconciliata": False, "bonifico_riconciliato": False}},
            upsert=True)
    return saldo


@router.get("/_unif_diag")
async def diagnostica_unificazione():
    """SOLA LETTURA. Fotografa cedolini vs paghe_mensili per pianificare l'unificazione:
    conteggi, sovrapposizioni per (dipendente_id, anno, mese), confronto netto vs importo_busta,
    record presenti solo in paghe_mensili, e dump completo di paghe_mensili per backup."""
    db = get_db()
    ced = await db.cedolini.find({}, {"_id": 0}).to_list(5000)
    pm = await db.paghe_mensili.find({}, {"_id": 0}).to_list(5000)
    ced_idx = {}
    for c in ced:
        ced_idx.setdefault((c.get("dipendente_id"), c.get("anno"), c.get("mese")), c)
    solo_in_pm, con_match, mismatch_netto = [], 0, []
    for p in pm:
        k = (p.get("dipendente_id"), p.get("anno"), p.get("mese"))
        c = ced_idx.get(k)
        if not c:
            solo_in_pm.append({"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"), "mese": p.get("mese")})
        else:
            con_match += 1
            nb = p.get("importo_busta") or p.get("netto_atteso")
            nc = c.get("netto")
            if nb is not None and nc is not None and abs(float(nb) - float(nc)) > 1:
                mismatch_netto.append({"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"),
                                       "mese": p.get("mese"), "paghe_mensili": nb, "cedolini": nc})
    # campi accessori presenti in paghe_mensili (riconciliazione)
    campi = set()
    for p in pm:
        campi.update(p.keys())
    pm_con_riconciliazione = [p for p in pm if any(p.get(k) for k in
        ("bonifico_importo", "acconti", "prestito_importo", "tfr_anticipo_importo",
         "busta_riconciliata", "bonifico_riconciliato"))]
    return {
        "cedolini_totali": len(ced),
        "paghe_mensili_totali": len(pm),
        "paghe_mensili_con_match_in_cedolini": con_match,
        "paghe_mensili_solo_loro": solo_in_pm,
        "mismatch_netto": mismatch_netto,
        "campi_presenti_in_paghe_mensili": sorted(campi),
        "paghe_mensili_con_dati_riconciliazione": len(pm_con_riconciliazione),
        "backup_paghe_mensili": pm,
    }


_RICON_FIELDS = ["bonifico_importo", "bonifico_data", "bonifico_ricevuto", "bonifico_causale",
                 "bonifico_cro", "bonifico_pdf", "bonifico_riconciliato", "busta_riconciliata",
                 "busta_da_lul", "acconti", "acconto_cedolino", "saldo_residuo", "netto_atteso",
                 "erogato_atteso", "fonte_excel", "tfr_anticipo_importo", "tfr_anticipo_data",
                 "tfr_anticipo_pdf", "prestito_importo", "prestito_saldo"]


@router.post("/_unif_esegui")
async def esegui_unificazione(dry_run: bool = True, limit: int = 25):
    """Unifica paghe_mensili dentro cedolini, A PICCOLI BATCH leggeri: processa solo i record
    non ancora migrati (flag _migrato sul documento paghe_mensili), così ogni chiamata è veloce.
    Chiamare ripetutamente finché completato=True. NON cancella paghe_mensili."""
    db = get_db()
    pendenti = await db.paghe_mensili.find({"_migrato": {"$ne": True}}).to_list(limit)
    if not pendenti:
        return {"dry_run": dry_run, "fatti_ora": 0, "restanti_da_fare": 0, "completato": True}
    arricchiti = creati = saltati = 0
    for p in pendenti:
        k = {"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"), "mese": p.get("mese")}
        ricon = {f: p[f] for f in _RICON_FIELDS if f in p and p[f] is not None}
        c = await db.cedolini.find_one(k, {"_id": 0, "id": 1})
        if dry_run:
            if c: arricchiti += 1
            elif (p.get("importo_busta") or p.get("netto_atteso") or 0) > 0: creati += 1
            else: saltati += 1
            continue
        if c:
            upd = dict(ricon); upd["unif_arricchito"] = True
            await db.cedolini.update_one({"id": c["id"]}, {"$set": upd})
            arricchiti += 1
        else:
            netto = p.get("importo_busta") or p.get("netto_atteso")
            if not netto or float(netto) <= 0:
                saltati += 1
            else:
                dip = await db.dipendenti.find_one({"id": p.get("dipendente_id")},
                                                   {"_id": 0, "nome": 1, "cognome": 1, "nome_completo": 1})
                nome = (dip or {}).get("nome_completo") or (f"{(dip or {}).get('cognome','')} {(dip or {}).get('nome','')}".strip() if dip else "")
                nuovo = {"id": str(uuid.uuid4()), "dipendente_id": p.get("dipendente_id"),
                         "dipendente_nome": nome, "anno": p.get("anno"), "mese": p.get("mese"),
                         "netto": float(netto), "stato": "importato",
                         "origine_unificazione": True, "unif_arricchito": True, "created_at": now_iso()}
                nuovo.update(ricon)
                await db.cedolini.insert_one(nuovo)
                creati += 1
        await db.paghe_mensili.update_one(
            {"dipendente_id": p.get("dipendente_id"), "anno": p.get("anno"), "mese": p.get("mese")},
            {"$set": {"_migrato": True}})
    restanti = await db.paghe_mensili.count_documents({"_migrato": {"$ne": True}})
    return {"dry_run": dry_run, "arricchiti_ora": arricchiti, "creati_ora": creati,
            "saltati_ora": saltati, "restanti_da_fare": restanti, "completato": restanti == 0}


@router.get("/prestiti")
async def lista_prestiti(dipendente_id: Optional[str] = None):
    """Mastrino prestiti: movimenti con saldo progressivo. Filtrabile per dipendente."""
    q = {"dipendente_id": dipendente_id} if dipendente_id else {}
    movs = await get_db().prestiti_dipendenti.find(q, {"_id": 0}).to_list(2000)
    movs.sort(key=lambda x: (x.get("anno", 0), x.get("mese", 0), x.get("data") or ""))
    # saldo progressivo per dipendente
    saldi = {}
    for mv in movs:
        d = mv["dipendente_id"]
        saldi[d] = saldi.get(d, 0) + (mv.get("importo") or 0)
        mv["saldo"] = saldi[d]
    return movs


@router.delete("/prestiti/{prestito_id}")
async def elimina_prestito(prestito_id: str):
    """Elimina un movimento di prestito e ricalcola il saldo progressivo del dipendente."""
    mv = await get_db().prestiti_dipendenti.find_one({"id": prestito_id})
    if not mv:
        raise HTTPException(status_code=404, detail="Prestito non trovato")
    await get_db().prestiti_dipendenti.delete_one({"id": prestito_id})
    # libera anche l'anti-dup così un eventuale re-import è possibile
    if mv.get("cro"):
        await get_db().documenti_importati.delete_many({"chiave": f"cro:{mv['cro']}"})
    saldo = await _ricalcola_saldo_prestiti(mv["dipendente_id"])
    return {"ok": True, "saldo_aggiornato": saldo}


def _espandi_in_pdf(nome, data):
    """Espande un allegato/file in lista di (origine, pdf_bytes): PDF diretto, oppure
    PDF contenuti in uno ZIP. Ritorna (items, errori)."""
    items, errori = [], []
    low = (nome or "").lower()
    if low.endswith(".pdf"):
        items.append((nome, data))
    elif low.endswith(".zip"):
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                interni = [n for n in z.namelist() if n.lower().endswith(".pdf") and "__MACOSX" not in n]
                if not interni:
                    errori.append(f"{nome}: ZIP senza PDF")
                for zi in interni:
                    items.append((f"{nome} › {zi}", z.read(zi)))
        except zipfile.BadZipFile:
            errori.append(f"{nome}: ZIP non valido")
        except Exception as e:
            errori.append(f"{nome}: {e}")
    else:
        errori.append(f"{nome}: tipo non supportato (servono PDF o ZIP)")
    return items, errori


@router.post("/paghe/importa-lul")
async def importa_libro_unico(files: List[UploadFile] = File(...), forza: bool = False):
    """Importa uno o più PDF (anche dentro ZIP) caricati dall'utente: buste paga,
    fogli presenze, bonifici (acconti, saldi, TFR, prestiti). Vedi _importa_documenti."""
    pdf_items, errori = [], []
    for uf in files:
        nome = uf.filename or ""
        try:
            data = await uf.read()
        except Exception:
            errori.append(f"{nome}: lettura fallita")
            continue
        its, err = _espandi_in_pdf(nome, data)
        pdf_items += its
        errori += err
    if not pdf_items:
        raise HTTPException(status_code=400,
            detail="Nessun PDF valido trovato. " + ("; ".join(errori) if errori else ""))
    return await _importa_documenti(pdf_items, errori, forza=forza)


@router.post("/paghe/importa-email")
async def importa_da_email(cartella: Optional[str] = None, solo_non_letti: bool = False):
    """Scarica gli allegati PDF dalla casella di posta (INBOX + tutte le cartelle) e li
    importa con la stessa pipeline. Credenziali dalle variabili ambiente Render:
    IMAP_HOST, IMAP_PORT (default 993), IMAP_USER, IMAP_PASSWORD.
    L'anti-duplicazione per hash evita di re-importare email già lette in passato."""
    import imaplib, email
    host = os.getenv("IMAP_HOST") or os.getenv("IMAP_SERVER")
    user = os.getenv("IMAP_USER") or os.getenv("IMAP_EMAIL")
    pwd = os.getenv("IMAP_PASSWORD") or os.getenv("IMAP_PASS")
    port = int(os.getenv("IMAP_PORT") or 993)
    mancano = [n for n, v in [("IMAP_HOST", host), ("IMAP_USER", user), ("IMAP_PASSWORD", pwd)] if not v]
    if mancano:
        raise HTTPException(status_code=400,
            detail="Variabili ambiente IMAP mancanti su Render: " + ", ".join(mancano) +
                   ". Servono IMAP_HOST, IMAP_USER, IMAP_PASSWORD (IMAP_PORT opzionale, default 993).")
    try:
        M = imaplib.IMAP4_SSL(host, port)
        M.login(user, pwd)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Connessione/login IMAP fallito ({host}:{port}): {e}")

    pdf_items, errori, cartelle_lette = [], [], []
    try:
        # Elenco cartelle: una specifica se richiesta, altrimenti tutte
        if cartella:
            target = [cartella]
        else:
            target = []
            typ, data = M.list()
            if typ == "OK":
                for raw in data:
                    line = raw.decode(errors="ignore") if isinstance(raw, bytes) else str(raw)
                    # l'ultimo token tra virgolette è il nome cartella
                    nome_c = line.split(' "')[-1].strip().strip('"') if '"' in line else line.split()[-1]
                    if nome_c and "\\Noselect" not in line:
                        target.append(nome_c)
            if "INBOX" not in target:
                target.insert(0, "INBOX")
        for box in target:
            try:
                typ, _ = M.select(f'"{box}"', readonly=True)
                if typ != "OK":
                    continue
                crit = "(UNSEEN)" if solo_non_letti else "ALL"
                typ, msgnums = M.search(None, crit)
                if typ != "OK":
                    continue
                ids = msgnums[0].split()
                cartelle_lette.append({"cartella": box, "messaggi": len(ids)})
                for num in ids:
                    typ, msgdata = M.fetch(num, "(RFC822)")
                    if typ != "OK" or not msgdata or not msgdata[0]:
                        continue
                    msg = email.message_from_bytes(msgdata[0][1])
                    for part in msg.walk():
                        if part.get_content_maintype() == "multipart":
                            continue
                        fn = part.get_filename()
                        if not fn:
                            continue
                        try:
                            payload = part.get_payload(decode=True)
                        except Exception:
                            continue
                        if not payload:
                            continue
                        its, err = _espandi_in_pdf(fn, payload)
                        pdf_items += [(f"[{box}] {o}", d) for (o, d) in its]
                        errori += err
            except Exception as e:
                errori.append(f"cartella {box}: {e}")
    finally:
        try:
            M.logout()
        except Exception:
            pass

    if not pdf_items:
        return {"associati": [], "da_controllare": [], "totale_associati": 0, "file_pdf": 0,
                "mesi": [], "errori": errori, "bonifici": [], "presenze": [], "duplicati": [],
                "tfr": [], "prestiti": [], "cartelle_lette": cartelle_lette,
                "messaggio": "Nessun allegato PDF trovato nella casella."}
    res = await _importa_documenti(pdf_items, errori)
    res["cartelle_lette"] = cartelle_lette
    return res


# ============ PRESENZE ============

@router.get("/presenze")
async def get_presenze(anno: Optional[int] = None, mese: Optional[int] = None, dipendente_id: Optional[str] = None):
    """
    Recupera presenze dalla collezione 'presenze' (dati storici dal Libro Unico).
    Le presenze sono raggruppate per mese con un array 'giorni'.
    """
    db = get_db()
    
    # Prima leggi da presenze_cloud (inserimenti manuali)
    query_cloud = {}
    if dipendente_id:
        query_cloud["dipendente_id"] = dipendente_id
    if anno and mese:
        query_cloud["data"] = {"$regex": f"^{anno}-{str(mese).zfill(2)}"}
    
    presenze_cloud = await db.presenze_cloud.find(query_cloud, {"_id": 0}).to_list(5000)
    
    # Poi leggi da presenze (dati storici dal LUL - struttura diversa)
    query_lul = {}
    if anno:
        query_lul["anno"] = anno
    if mese:
        query_lul["mese"] = mese
    
    presenze_lul = await db.presenze.find(query_lul, {"_id": 0}).to_list(500)
    
    # Converti presenze LUL in formato giornaliero
    result = list(presenze_cloud)
    cloud_keys = {(p.get("dipendente_id"), p.get("data")) for p in presenze_cloud}
    
    for p_lul in presenze_lul:
        cf = p_lul.get("codice_fiscale", "")
        anno_p = p_lul.get("anno", 2026)
        mese_p = p_lul.get("mese", 1)
        giorni = p_lul.get("giorni", [])
        
        # Trova l'ID dipendente dal codice fiscale
        dip = await db.dipendenti.find_one({"codice_fiscale": cf})
        dip_id = dip.get("id", cf) if dip else cf
        
        for g in giorni:
            giorno_num = g.get("giorno", 1)
            data_str = f"{anno_p}-{str(mese_p).zfill(2)}-{str(giorno_num).zfill(2)}"
            
            key = (dip_id, data_str)
            if key in cloud_keys:
                continue  # Già presente nei dati manuali
            
            # Determina lo stato dal giustificativo
            giust = g.get("giustificativo", "")
            ore = g.get("ore_ordinarie", 0)
            
            if giust:
                stato = giust  # AI, FE, MA, RL, etc.
            elif ore > 0:
                stato = "presente"
            else:
                stato = "assente"
            
            result.append({
                "id": f"{cf}_{data_str}",
                "dipendente_id": dip_id,
                "data": data_str,
                "entrata": None,
                "uscita": None,
                "stato": stato,
                "giustificativo": giust,
                "ore_lavorate": ore,
                "note": ""
            })
    
    return result

@router.post("/presenze")
async def create_presenza(presenza: PresenzaCloud):
    pres_dict = presenza.model_dump()
    pres_dict["id"] = generate_id()
    pres_dict["created_at"] = now_iso()
    
    # Calculate hours worked
    if pres_dict.get("entrata") and pres_dict.get("uscita"):
        try:
            ent = datetime.strptime(pres_dict["entrata"], "%H:%M")
            usc = datetime.strptime(pres_dict["uscita"], "%H:%M")
            pres_dict["ore_lavorate"] = round((usc - ent).seconds / 3600, 2)
        except:
            pass
    
    await get_db().presenze_cloud.insert_one(pres_dict)
    return serialize_doc(pres_dict)

@router.put("/presenze/{presenza_id}")
async def update_presenza(presenza_id: str, presenza: PresenzaCloud):
    pres_dict = presenza.model_dump()
    
    if pres_dict.get("entrata") and pres_dict.get("uscita"):
        try:
            ent = datetime.strptime(pres_dict["entrata"], "%H:%M")
            usc = datetime.strptime(pres_dict["uscita"], "%H:%M")
            pres_dict["ore_lavorate"] = round((usc - ent).seconds / 3600, 2)
        except:
            pass
    
    result = await get_db().presenze_cloud.update_one(
        {"id": presenza_id},
        {"$set": pres_dict}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Presenza non trovata")
    return {"message": "Presenza aggiornata"}

@router.delete("/presenze/{presenza_id}")
async def delete_presenza(presenza_id: str):
    result = await get_db().presenze_cloud.delete_one({"id": presenza_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Presenza non trovata")
    return {"message": "Presenza eliminata"}

@router.post("/presenze/batch")
async def create_presenze_batch(presenze: List[PresenzaCloud]):
    created = []
    for p in presenze:
        pres_dict = p.model_dump()
        pres_dict["id"] = generate_id()
        pres_dict["created_at"] = now_iso()
        
        existing = await get_db().presenze_cloud.find_one({
            "dipendente_id": pres_dict["dipendente_id"],
            "data": pres_dict["data"]
        })
        
        if existing:
            await get_db().presenze_cloud.update_one(
                {"id": existing["id"]},
                {"$set": pres_dict}
            )
        else:
            await get_db().presenze_cloud.insert_one(pres_dict)
        created.append(pres_dict)
    
    return {"message": f"Inserite/aggiornate {len(created)} presenze"}

# ============ FERIE E PERMESSI ============

@router.get("/ferie")
async def get_ferie(dipendente_id: Optional[str] = None, stato: Optional[str] = None):
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    if stato:
        query["stato"] = stato
    ferie = await get_db().ferie_cloud.find(query, {"_id": 0}).to_list(1000)
    return ferie

@router.post("/ferie")
async def create_ferie(ferie: FerieCloud):
    ferie_dict = ferie.model_dump()
    ferie_dict["id"] = generate_id()
    ferie_dict["created_at"] = now_iso()
    await get_db().ferie_cloud.insert_one(ferie_dict)
    return serialize_doc(ferie_dict)

@router.post("/ferie-giorno")
async def set_ferie_giorno(data: dict):
    """Assegna/aggiorna/rimuove un'assenza di un singolo giorno dal calendario.
    tipo=None rimuove. Usato dalla vista calendario di Ferie & Permessi."""
    dip = data.get("dipendente_id")
    giorno = data.get("data")
    tipo = data.get("tipo")
    if not dip or not giorno:
        raise HTTPException(status_code=400, detail="dipendente_id e data obbligatori")
    existing = await get_db().ferie_cloud.find_one({
        "dipendente_id": dip, "data_inizio": giorno, "data_fine": giorno
    })
    if tipo:
        if existing:
            await get_db().ferie_cloud.update_one({"id": existing["id"]}, {"$set": {"tipo": tipo}})
        else:
            await get_db().ferie_cloud.insert_one({
                "id": generate_id(), "dipendente_id": dip, "tipo": tipo,
                "data_inizio": giorno, "data_fine": giorno, "giorni": 1,
                "stato": "approvata", "created_at": now_iso()
            })
    elif existing:
        await get_db().ferie_cloud.delete_one({"id": existing["id"]})
    return {"ok": True}

@router.put("/ferie/{ferie_id}/approva")
async def approva_ferie(ferie_id: str):
    result = await get_db().ferie_cloud.update_one(
        {"id": ferie_id},
        {"$set": {"stato": "approvata"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    return {"message": "Richiesta approvata"}

@router.put("/ferie/{ferie_id}/rifiuta")
async def rifiuta_ferie(ferie_id: str):
    result = await get_db().ferie_cloud.update_one(
        {"id": ferie_id},
        {"$set": {"stato": "rifiutata"}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    return {"message": "Richiesta rifiutata"}

@router.delete("/ferie/{ferie_id}")
async def delete_ferie(ferie_id: str):
    result = await get_db().ferie_cloud.delete_one({"id": ferie_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Richiesta non trovata")
    return {"message": "Richiesta eliminata"}

# ============ TURNI ============

@router.get("/impostazioni-turni")
async def get_impostazioni_turni():
    """Impostazioni della generazione turni (documento unico)."""
    doc = await get_db().impostazioni_cloud.find_one({"chiave": "turni"}, {"_id": 0})
    return (doc or {}).get("valore", {"bar_chiuso_domenica_pomeriggio": True})

@router.post("/impostazioni-turni")
async def set_impostazioni_turni(valore: dict):
    await get_db().impostazioni_cloud.update_one(
        {"chiave": "turni"},
        {"$set": {"chiave": "turni", "valore": valore, "updated_at": now_iso()}},
        upsert=True)
    return {"ok": True, "valore": valore}

@router.get("/turni")
async def get_turni():
    turni = await get_db().turni_cloud.find({}, {"_id": 0}).to_list(100)
    return turni

@router.post("/turni")
async def create_turno(turno: TurnoCloud):
    turno_dict = turno.model_dump()
    turno_dict["id"] = generate_id()
    await get_db().turni_cloud.insert_one(turno_dict)
    return serialize_doc(turno_dict)

@router.put("/turni/{turno_id}")
async def update_turno(turno_id: str, turno: TurnoCloud):
    result = await get_db().turni_cloud.update_one(
        {"id": turno_id},
        {"$set": turno.model_dump()}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Turno non trovato")
    return {"message": "Turno aggiornato"}

@router.delete("/turni/{turno_id}")
async def delete_turno(turno_id: str):
    result = await get_db().turni_cloud.delete_one({"id": turno_id})
    await get_db().assegnazioni_turni_cloud.delete_many({"turno_id": turno_id})
    return {"message": "Turno eliminato"}

@router.get("/assegnazioni-turni")
async def get_assegnazioni(settimana: Optional[str] = None):
    query = {"settimana": settimana} if settimana else {}
    assegnazioni = await get_db().assegnazioni_turni_cloud.find(query, {"_id": 0}).to_list(2000)
    return assegnazioni

@router.post("/assegnazioni-turni/migra")
async def migra_settimana_assegnazioni(data: dict):
    """Una-tantum: assegna una settimana ai record che non ce l'hanno."""
    settimana = data.get("settimana")
    if not settimana:
        raise HTTPException(status_code=400, detail="settimana obbligatoria")
    res = await get_db().assegnazioni_turni_cloud.update_many(
        {"$or": [{"settimana": {"$exists": False}}, {"settimana": None}]},
        {"$set": {"settimana": settimana}}
    )
    return {"migrati": res.modified_count}

@router.post("/assegnazioni-turni")
async def create_or_update_assegnazione(data: dict):
    dipendente_id = data.get("dipendente_id")
    giorno = data.get("giorno")
    turno_id = data.get("turno_id")
    settimana = data.get("settimana")
    
    if not dipendente_id or not giorno:
        raise HTTPException(status_code=400, detail="dipendente_id e giorno sono obbligatori")
    
    motivo = data.get("motivo")  # es. "onomastico" → reso visibile nei turni
    match = {"dipendente_id": dipendente_id, "giorno": giorno}
    if settimana:
        match["settimana"] = settimana
    existing = await get_db().assegnazioni_turni_cloud.find_one(match)

    if turno_id:
        if existing:
            upd = {"$set": {"turno_id": turno_id}}
            if motivo:
                upd["$set"]["motivo"] = motivo
            else:
                upd["$unset"] = {"motivo": ""}
            await get_db().assegnazioni_turni_cloud.update_one({"id": existing["id"]}, upd)
        else:
            ass = {
                "id": generate_id(),
                "dipendente_id": dipendente_id,
                "giorno": giorno,
                "turno_id": turno_id,
                "settimana": settimana,
            }
            if motivo:
                ass["motivo"] = motivo
            await get_db().assegnazioni_turni_cloud.insert_one(ass)
    else:
        if existing:
            await get_db().assegnazioni_turni_cloud.delete_one({"id": existing["id"]})
    
    return {"message": "Assegnazione salvata"}

# ============ CONFIG TURNI PER DIPENDENTE ============
# Per ogni dipendente: turno abituale (turno_id) + giorno di riposo fisso
# settimanale (riposo_giorno, nome italiano). Usati da "Genera settimana".
@router.get("/turni-config")
async def get_turni_config():
    return await get_db().turni_config.find({}, {"_id": 0}).to_list(1000)


@router.post("/turni-config")
async def save_turni_config(data: dict = Body(...)):
    """Body: {voci: [{dipendente_id, turno_id, riposo_giorno}]}."""
    db = get_db()
    for v in (data.get("voci") or []):
        if not v.get("dipendente_id"):
            continue
        await db.turni_config.update_one(
            {"dipendente_id": v["dipendente_id"]},
            {"$set": {"dipendente_id": v["dipendente_id"],
                      "turno_id": v.get("turno_id") or None,
                      "riposo_giorno": v.get("riposo_giorno") or None,
                      "lunga_giorni": v.get("lunga_giorni") or [],
                      "rotazione": v.get("rotazione") or None,
                      "sala": bool(v.get("sala")),
                      "updated_at": now_iso()}}, upsert=True)
    return {"ok": True, "salvati": len(data.get("voci") or [])}

# ============ ONOMASTICI (riposo per onomastico nei turni) ============
# Date standard italiane (mese, giorno) per nome proprio. Prefillate e
# MODIFICABILI in gestione. I nomi non presenti sono "stranieri" → esclusi.
ONOMASTICI_DEFAULT = {
    "angela": (1, 27), "angelo": (10, 2), "anna": (7, 26), "antonella": (6, 13),
    "antonietta": (6, 13), "antonio": (6, 13), "carmela": (7, 16), "carmine": (7, 16),
    "caterina": (11, 25), "ciro": (1, 31), "domenico": (8, 8), "elena": (8, 18),
    "emanuele": (3, 26), "fabio": (5, 11), "francesca": (3, 9), "francesco": (10, 4),
    "gaetano": (8, 7), "gennaro": (9, 19), "giorgio": (4, 23), "giovanna": (5, 30),
    "giovanni": (6, 24), "giulia": (5, 22), "giuliano": (1, 9), "giuseppa": (3, 19),
    "giuseppe": (3, 19), "ignazio": (7, 31), "liliana": (7, 27), "lucia": (12, 13),
    "luigi": (6, 21), "luigia": (6, 21), "marcella": (1, 31), "marco": (4, 25),
    "margherita": (2, 22), "maria": (9, 12), "mariano": (8, 19), "marina": (7, 17),
    "mario": (1, 19), "michele": (9, 29), "ottavio": (11, 20), "pasquale": (5, 17),
    "paolo": (6, 29), "pietro": (6, 29), "raffaele": (9, 29), "rosa": (8, 23),
    "salvatore": (8, 6), "simone": (10, 28), "stefano": (12, 26), "teresa": (10, 15),
    "valerio": (1, 29), "vincenzo": (1, 22), "vincenza": (1, 22),
}
NOMI_GIORNO_IT = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato", "Domenica"]
# Dipendenti che NON seguono i turni → niente riposo onomastico (decisione titolare).
NON_TURNI = [("vincenzo", "ceraldi"), ("valerio", "ceraldi"),
             ("antonietta", "ceraldi"), ("marina", "liuzza")]


def _non_turni(d: dict) -> bool:
    f = f"{d.get('nome','')} {d.get('cognome','')} {d.get('nome_completo','')}".lower()
    return any(a in f and b in f for a, b in NON_TURNI)


def _nome_proprio(dip: dict) -> str:
    n = (dip.get("nome") or "").strip()
    if not n and dip.get("nome_completo"):
        n = dip["nome_completo"].split()[0]
    return n.split()[0].lower() if n else ""


@router.get("/onomastici")
async def get_onomastici():
    """Onomastico per ogni dipendente attivo: data (prefillata dal nome o salvata),
    attivo e flag 'straniero' (nome senza onomastico italiano)."""
    db = get_db()
    dips = await db.dipendenti.find(
        {"merged_into": {"$exists": False}}, {"_id": 0}).to_list(1000)
    salvati = {o["dipendente_id"]: o async for o in db.onomastici.find({}, {"_id": 0})}
    out = []
    for d in dips:
        if d.get("attivo") is False or (d.get("stato") or "attivo") in ("cessato", "dimesso", "archiviato"):
            continue
        nome = _nome_proprio(d)
        default = ONOMASTICI_DEFAULT.get(nome)
        straniero = default is None
        s = salvati.get(d.get("id"))
        if s:
            mese, giorno, attivo = s.get("mese"), s.get("giorno"), s.get("attivo", True)
        else:
            mese, giorno = (default if default else (None, None))
            attivo = (not straniero) and (not _non_turni(d))
        out.append({
            "dipendente_id": d.get("id"),
            "nome": d.get("nome_completo") or f"{d.get('cognome','')} {d.get('nome','')}".strip(),
            "mese": mese, "giorno": giorno, "attivo": bool(attivo), "straniero": straniero,
            "non_turni": _non_turni(d),
        })
    out.sort(key=lambda x: (x["nome"] or "").lower())
    return out


@router.post("/onomastici")
async def save_onomastici(data: dict = Body(...)):
    """Salva le date/attivo onomastico. Body: {voci: [{dipendente_id, mese, giorno, attivo}]}."""
    db = get_db()
    for v in (data.get("voci") or []):
        if not v.get("dipendente_id"):
            continue
        await db.onomastici.update_one(
            {"dipendente_id": v["dipendente_id"]},
            {"$set": {"dipendente_id": v["dipendente_id"],
                      "mese": v.get("mese"), "giorno": v.get("giorno"),
                      "attivo": bool(v.get("attivo", True)),
                      "updated_at": now_iso()}}, upsert=True)
    return {"ok": True, "salvati": len(data.get("voci") or [])}


@router.get("/onomastici/settimana")
async def onomastici_settimana(settimana: str):
    """Onomastici (idonei al riposo) che cadono nella settimana indicata (lunedì
    ISO). Esclude stranieri, esclusi (attivo=False) e la domenica (bar chiuso)."""
    try:
        lun = datetime.strptime(settimana, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="settimana deve essere YYYY-MM-DD (lunedì)")
    voci = await get_onomastici()
    giorni_sett = [(lun + timedelta(days=i)) for i in range(7)]
    out = []
    for v in voci:
        if not v["attivo"] or v["straniero"] or not v["mese"] or not v["giorno"]:
            continue
        for i, gd in enumerate(giorni_sett):
            if gd.month == v["mese"] and gd.day == v["giorno"] and i < 6:  # esclude domenica
                out.append({
                    "dipendente_id": v["dipendente_id"], "nome": v["nome"],
                    "data": gd.strftime("%Y-%m-%d"), "giorno_nome": NOMI_GIORNO_IT[i],
                    "data_label": gd.strftime("%d/%m"),
                })
    return out

# ============ MOTORE DI INTERROGAZIONE CEDOLINI ============

@router.get("/cedolini/cerca-voce")
async def cerca_voce(codice: Optional[str] = None, testo: Optional[str] = None,
                     anno: Optional[int] = None, dipendente_id: Optional[str] = None):
    """Cerca una voce in TUTTI i cedolini salvati (campo voci). Per codice (es. F09081)
    o per testo della descrizione (es. '730', '13ma'). Filtrabile per anno/dipendente."""
    if not codice and not testo:
        raise HTTPException(status_code=400, detail="Indica 'codice' (es. F09081) o 'testo' (es. 730) da cercare")
    q: dict = {}
    if anno:
        q["anno"] = anno
    if dipendente_id:
        q["dipendente_id"] = dipendente_id
    cod = (codice or "").upper().strip()
    txt = (testo or "").lower().strip()
    out = []
    async for c in get_db().cedolini.find(q, {"_id": 0, "dipendente_id": 1, "dipendente_nome": 1, "anno": 1, "mese": 1, "voci": 1}):
        for v in (c.get("voci") or []):
            if (cod and v.get("codice") == cod) or (txt and txt in (v.get("descrizione") or "").lower()):
                out.append({"dipendente_id": c.get("dipendente_id"), "dipendente": c.get("dipendente_nome"),
                            "anno": c.get("anno"), "mese": c.get("mese"),
                            "codice": v.get("codice"), "descrizione": v.get("descrizione"),
                            "importo": (v.get("valori") or [None])[-1], "valori": v.get("valori")})
    out.sort(key=lambda x: (x.get("anno") or 0, x.get("mese") or 0))
    return {"risultati": out, "totale": len(out)}


@router.post("/cedolini/riscansiona")
async def riscansiona_cedolini(anno: Optional[int] = None, dipendente_id: Optional[str] = None):
    """Ri-estrae tutte le voci dai cedolini storici (2023→oggi) che hanno il PDF salvato,
    così il motore di ricerca trova ogni codice anche sulle buste già importate."""
    import io
    import pdfplumber
    db = get_db()
    q: dict = {"pdf_data": {"$exists": True}}
    if anno:
        q["anno"] = anno
    if dipendente_id:
        q["dipendente_id"] = dipendente_id
    aggiornati, errori = 0, 0
    async for c in db.cedolini.find(q, {"_id": 0, "id": 1, "pdf_data": 1}):
        try:
            raw = base64.b64decode(c["pdf_data"])
            text = ""
            with pdfplumber.open(io.BytesIO(raw)) as pdf:
                for p in pdf.pages:
                    text += (p.extract_text() or "") + "\n"
            dati = _lul_dati_busta(text)
            if dati:
                await db.cedolini.update_one({"id": c["id"]}, {"$set": dati})
                aggiornati += 1
        except Exception:
            errori += 1
    return {"aggiornati": aggiornati, "errori": errori,
            "nota": "I cedolini senza PDF salvato non possono essere riscansionati: vanno re-importati dal Libro Unico."}


# ============ IMPORT PRIMA NOTA SALARI (Excel) ============

@router.post("/paghe/importa-prima-nota")
async def importa_prima_nota(file: UploadFile = File(...)):
    """Importa la 'Prima Nota Salari' (Excel: Dipendente, Mese, Anno, Stipendio Netto,
    Importo Erogato). Per ogni dipendente/mese/anno SOMMA gli Importi Erogati (più bonifici
    nello stesso mese) e li scrive in paghe_mensili.bonifico_importo. Riempie l'importo
    busta se mancante. Confronta col dato già in app e segnala differenze e nomi non trovati."""
    import io
    import openpyxl
    raw = await file.read()
    if raw[:2] != b"PK":
        raise HTTPException(400, "Il file deve essere un .xlsx")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel non valido: {e}")
    ws = wb["Salari"] if "Salari" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "Foglio vuoto")
    header = [(str(c).strip().lower() if c is not None else "") for c in rows[0]]

    def col(*names):
        for i, h in enumerate(header):
            if h in names:
                return i
        return None
    ci_dip, ci_mese, ci_anno = col("dipendente"), col("mese"), col("anno")
    ci_netto = col("stipendio netto", "netto", "importo busta")
    ci_erog = col("importo erogato", "erogato", "bonifico")
    if None in (ci_dip, ci_mese, ci_anno, ci_erog):
        raise HTTPException(400, "Colonne attese: Dipendente, Mese, Anno, Stipendio Netto, Importo Erogato")

    MESI = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
            "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12}

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()

    def fnum(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    agg = {}
    for r in rows[1:]:
        if ci_dip >= len(r) or not r[ci_dip]:
            continue
        mese = MESI.get(norm(r[ci_mese]))
        try:
            anno = int(r[ci_anno])
        except (TypeError, ValueError):
            anno = None
        if not mese or not anno:
            continue
        k = (norm(r[ci_dip]), mese, anno)
        a = agg.setdefault(k, {"nome": str(r[ci_dip]).strip(), "netto": 0.0, "erogato": 0.0})
        a["erogato"] += fnum(r[ci_erog])
        if ci_netto is not None and ci_netto < len(r):
            a["netto"] += fnum(r[ci_netto])

    db = get_db()
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1}).to_list(1000)
    by_nome = {}
    for d in dips:
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for v in {norm(d.get("nome_completo")), f"{c} {n}".strip(), f"{n} {c}".strip()}:
            if v:
                by_nome[v] = d

    aggiornati, non_trovati, discrepanze = [], [], []
    for (nome_n, mese, anno), a in agg.items():
        erog, netto = round(a["erogato"], 2), round(a["netto"], 2)
        if erog <= 0 and netto <= 0:
            continue
        d = by_nome.get(nome_n)
        if not d:
            non_trovati.append({"nome": a["nome"], "mese": mese, "anno": anno, "bonifico": erog})
            continue
        existing = await db.paghe_mensili.find_one(
            {"dipendente_id": d["id"], "anno": anno, "mese": mese}, {"importo_busta": 1}) or {}
        set_doc = {"dipendente_id": d["id"], "anno": anno, "mese": mese,
                   "bonifico_importo": erog, "bonifico_ricevuto": erog > 0,
                   "bonifico_da_prima_nota": True, "updated_at": now_iso()}
        busta_app = existing.get("importo_busta")
        if (busta_app in (None, 0, "")) and netto > 0:
            set_doc["importo_busta"] = netto
        elif busta_app and netto > 0 and abs(float(busta_app) - netto) > 1:
            discrepanze.append({"dipendente": a["nome"], "mese": mese, "anno": anno,
                                "busta_app": round(float(busta_app), 2), "busta_excel": netto})
        await db.paghe_mensili.update_one(
            {"dipendente_id": d["id"], "anno": anno, "mese": mese}, {"$set": set_doc}, upsert=True)
        await _ricalcola_stato_paga(db, d["id"], anno, mese)
        aggiornati.append({"dipendente": a["nome"], "mese": mese, "anno": anno, "bonifico": erog})

    nomi_non_trovati = sorted({x["nome"] for x in non_trovati})
    return {"aggiornati": len(aggiornati),
            "righe_aggregate": len(agg),
            "non_trovati": len(non_trovati),
            "nomi_non_trovati": nomi_non_trovati,
            "discrepanze": sorted(discrepanze, key=lambda x: (x["anno"], x["mese"]))}


_MESI_IT = {"gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5, "giugno": 6,
            "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12}


@router.post("/dipendenti/importa-anagrafica")
async def importa_anagrafica(file: UploadFile = File(...)):
    """Importa/aggiorna l'anagrafica da Excel (Cognome, Nome, CF, Data di nascita,
    Mansione, Telefono, Email, Indirizzo). Match per codice fiscale: aggiorna se esiste,
    altrimenti crea."""
    import io
    import openpyxl
    raw = await file.read()
    if raw[:2] != b"PK":
        raise HTTPException(400, "Il file deve essere un .xlsx")
    wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [(str(c).strip().lower() if c is not None else "") for c in rows[0]]

    def col(*names):
        for i, h in enumerate(header):
            if any(n in h for n in names):
                return i
        return None
    ci = {"cognome": col("cognome"), "nome": col("nome"), "cf": col("cf", "codice fiscale"),
          "nascita": col("nascita"), "mansione": col("mansione"), "tel": col("telefono", "cell"),
          "email": col("email", "mail"), "indirizzo": col("indirizzo")}
    db = get_db()
    creati, aggiornati = 0, 0

    def val(r, k):
        i = ci.get(k)
        if i is None or i >= len(r) or r[i] is None:
            return None
        return str(r[i]).strip()
    for r in rows[1:]:
        cf = (val(r, "cf") or "").upper().replace(" ", "")
        nome, cognome = val(r, "nome"), val(r, "cognome")
        if not (cf or (nome and cognome)):
            continue
        campi = {"nome": nome, "cognome": cognome, "codice_fiscale": cf or None,
                 "data_nascita": (val(r, "nascita") or "")[:10] or None,
                 "mansione": val(r, "mansione"), "telefono": val(r, "tel"),
                 "email": val(r, "email"), "indirizzo": val(r, "indirizzo")}
        campi = {k: v for k, v in campi.items() if v}
        campi["nome_completo"] = f"{cognome or ''} {nome or ''}".strip()
        esistente = await db.dipendenti.find_one({"codice_fiscale": cf}) if cf else None
        if esistente:
            await db.dipendenti.update_one({"id": esistente["id"]}, {"$set": campi})
            aggiornati += 1
        else:
            campi.update({"id": generate_id(), "attivo": True, "stato": "attivo", "created_at": now_iso()})
            await db.dipendenti.insert_one(campi)
            creati += 1
    return {"creati": creati, "aggiornati": aggiornati}


@router.post("/riduzioni-orario")
async def save_riduzioni_orario(data: dict = Body(...)):
    """Salva la riduzione oraria per dipendente: ore/giorno ridotte, paga oraria,
    data inizio e data fine (scadenza sorvegliata). Body: {voci:[{dipendente_id,...}]}."""
    db = get_db()
    n = 0
    for v in (data.get("voci") or []):
        did = v.get("dipendente_id")
        if not did:
            continue

        def num(x):
            try:
                return float(str(x).replace(",", ".")) if x not in (None, "") else None
            except (TypeError, ValueError):
                return None
        rid = {"attiva": bool(v.get("attiva")),
               "ore_giorno": num(v.get("ore_giorno")),
               "paga_oraria": num(v.get("paga_oraria")),
               "data_inizio": v.get("data_inizio") or None,
               "data_fine": v.get("data_fine") or None,
               "note": (v.get("note") or "").strip(),
               "updated_at": now_iso()}
        await db.dipendenti.update_one({"id": did}, {"$set": {"riduzione_orario": rid}})
        n += 1
    return {"salvati": n}


@router.get("/riduzioni-orario/scadenze")
async def riduzioni_in_scadenza(giorni: int = 30):
    """Riduzioni attive con scadenza entro N giorni (o già scadute) — vigilanza contratto."""
    db = get_db()
    oggi = datetime.now(timezone.utc).date()
    out = []
    async for d in db.dipendenti.find({"riduzione_orario.attiva": True}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "riduzione_orario": 1}):
        rid = d.get("riduzione_orario") or {}
        df = rid.get("data_fine")
        if not df:
            continue
        try:
            scad = datetime.strptime(df[:10], "%Y-%m-%d").date()
        except ValueError:
            continue
        gg = (scad - oggi).days
        if gg <= giorni:
            out.append({"dipendente_id": d.get("id"),
                        "nome": f"{d.get('cognome','')} {d.get('nome','')}".strip(),
                        "data_fine": df, "giorni_alla_scadenza": gg,
                        "scaduta": gg < 0, "ore_giorno": rid.get("ore_giorno")})
    out.sort(key=lambda x: x["giorni_alla_scadenza"])
    return out


@router.post("/paghe/importa-pagamenti")
async def importa_pagamenti(file: UploadFile = File(...)):
    """Importa i bonifici/pagamenti dal CSV banca. Riconosce due formati dall'intestazione:
    1) ESITI bonifici (Esecuzione;Ordinante;Beneficiario;Importo;Div;Causale;CRO);
    2) ANDAMENTO conto (Ragione Sociale;Data contabile;Data valuta;Banca;Rapporto;Importo;
       Divisa;Descrizione;Categoria;Hashtag): tiene solo le USCITE (importo negativo),
       scarta commissioni bancarie, estrae il nominativo dal 'FAVORE <Nome>' nella descrizione.
    In entrambi i casi aggancia solo chi è in anagrafica (fornitori esclusi automaticamente).
    Mese di competenza dalla causale (es. '9-2025', 'luglio') o, in mancanza, dalla data del
    movimento. Idempotente (dedup per CRO o hash riga). Aggiorna il bonifico del mese = somma
    dei pagamenti di quel mese e ricalcola lo stato paga (alimenta la prima nota)."""
    import io
    import csv as _csv
    raw = await file.read()
    text = raw.decode("utf-8", errors="ignore")
    reader = _csv.reader(io.StringIO(text), delimiter=";")
    righe = list(reader)
    if not righe:
        raise HTTPException(400, "CSV vuoto")
    db = get_db()
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1}).to_list(1000)

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()
    by_nome, by_cogn = {}, {}
    for d in dips:
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for v in {norm(d.get("nome_completo")), f"{c} {n}".strip(), f"{n} {c}".strip()}:
            if v and len(v) > 5:
                by_nome[v] = d
        if len(c) >= 4:
            by_cogn.setdefault(c, []).append(d)

    def trova_dip(beneficiario):
        b = norm(beneficiario)
        for nome_n, d in by_nome.items():
            if nome_n in b or b in nome_n:
                return d
        for cogn, lst in by_cogn.items():
            if cogn in b and len(lst) == 1:
                return lst[0]
        return None

    def to_float(s):
        try:
            return float(str(s).replace(".", "").replace(",", "."))
        except (TypeError, ValueError):
            return None

    def mese_anno(causale, data_dt):
        c = norm(causale)
        m = re.search(r'\b(\d{1,2})[-/](20\d{2})\b', c)
        if m:
            return int(m.group(1)), int(m.group(2))
        for nome, n in _MESI_IT.items():
            if nome in c:
                y = re.search(r'(20\d{2})', c)
                return n, int(y.group(1)) if y else data_dt.year
        return data_dt.month, data_dt.year

    def favore(s):
        m = re.search(r'favore\s+(.+?)(?:\s+-|\s+notprovide|$)', norm(s))
        return (m.group(1) if m else norm(s))[:50]

    # Rileva il formato dall'intestazione: ESITI bonifici o ESTRATTO CONTO (entrate/uscite)
    hdr = [norm(c) for c in (righe[0] if righe else [])]

    def col(*names):
        return next((i for i, h in enumerate(hdr) if any(n in h for n in names)), None)
    i_ben = col("beneficiario")
    if i_ben is not None:
        formato = "esiti"
        i_data = col("esecuzione", "data") if col("esecuzione", "data") is not None else 0
        i_imp = col("importo") if col("importo") is not None else 3
        i_caus = col("causale", "descrizione")
        i_cro = col("cro")
        i_cat = None
    else:
        formato = "andamento"
        i_data = col("data contabile", "data valuta", "data")
        i_imp = col("importo")
        i_caus = col("descrizione")
        i_cat = col("categoria")
        i_cro = None
        i_ben = i_caus

    importati, non_trovati, affected = 0, [], set()
    for r in righe[1:]:
        if i_imp is None or i_imp >= len(r) or i_data is None or i_data >= len(r):
            continue
        importo = to_float(r[i_imp])
        if importo is None:
            continue
        try:
            data_dt = datetime.strptime(str(r[i_data]).strip()[:10], "%d/%m/%Y")
        except (ValueError, TypeError):
            continue
        causale = (r[i_caus] if i_caus is not None and i_caus < len(r) else "") or ""
        if formato == "andamento":
            if importo >= 0:  # solo uscite = pagamenti
                continue
            cat = (r[i_cat] if i_cat is not None and i_cat < len(r) else "") or ""
            if "commission" in norm(cat) or norm(causale).startswith("comm"):
                continue  # niente commissioni bancarie
            importo = -importo
            beneficiario = favore(causale)
        else:
            if importo <= 0:
                continue
            beneficiario = (r[i_ben] if i_ben is not None and i_ben < len(r) else "") or ""
        if importo < 5:
            continue
        d = trova_dip(beneficiario if formato == "esiti" else causale)
        if not d:
            non_trovati.append(beneficiario or favore(causale))
            continue
        mese, anno = mese_anno(causale, data_dt)
        cro = (r[i_cro].strip() if i_cro is not None and i_cro < len(r) and r[i_cro] else "")
        key = cro or hashlib.sha1(f"{d['id']}|{r[i_data]}|{importo}|{causale}".encode()).hexdigest()
        await db.pagamenti_esiti.update_one(
            {"key": key},
            {"$set": {"key": key, "cro": cro, "dipendente_id": d["id"], "data": data_dt.strftime("%Y-%m-%d"),
                      "importo": importo, "causale": causale, "beneficiario": beneficiario,
                      "mese": mese, "anno": anno}}, upsert=True)
        affected.add((d["id"], mese, anno))
        importati += 1
    # ricalcola il bonifico del mese = somma dei pagamenti di quel mese
    for dip_id, mese, anno in affected:
        tot = 0.0
        async for p in db.pagamenti_esiti.find({"dipendente_id": dip_id, "mese": mese, "anno": anno}, {"_id": 0, "importo": 1}):
            tot += p.get("importo") or 0
        await db.paghe_mensili.update_one(
            {"dipendente_id": dip_id, "anno": anno, "mese": mese},
            {"$set": {"dipendente_id": dip_id, "anno": anno, "mese": mese,
                      "bonifico_importo": round(tot, 2), "bonifico_ricevuto": tot > 0,
                      "bonifico_da_esiti": True, "updated_at": now_iso()}}, upsert=True)
        await _ricalcola_stato_paga(db, dip_id, anno, mese)
    return {"importati": importati, "mesi_aggiornati": len(affected),
            "non_trovati": sorted(set(non_trovati))}


@router.get("/paghe/in-attesa")
async def paghe_in_attesa():
    """Buste in attesa di pagamento (o parziali): elenco per il pannello/avvisi."""
    db = get_db()
    dip_map = {d["id"]: f"{d.get('cognome','')} {d.get('nome','')}".strip()
               async for d in db.dipendenti.find({}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1})}
    out = []
    async for p in db.paghe_mensili.find(
            {"stato_pagamento": {"$in": ["in_attesa_pagamento", "parziale"]}}, {"_id": 0}):
        saldo = p.get("saldo")
        if saldo is None:
            saldo = round(float(p.get("importo_busta") or 0) - float(p.get("bonifico_importo") or 0), 2)
        if not saldo or saldo <= 0.5:
            continue
        out.append({"dipendente_id": p.get("dipendente_id"),
                    "dipendente": dip_map.get(p.get("dipendente_id"), p.get("dipendente_id")),
                    "anno": p.get("anno"), "mese": p.get("mese"),
                    "stato": p.get("stato_pagamento"),
                    "busta": round(float(p.get("importo_busta") or 0), 2),
                    "saldo": round(saldo, 2)})
    out.sort(key=lambda x: (x["anno"] or 0, x["mese"] or 0))
    return {"righe": out, "totale": len(out), "importo": round(sum(x["saldo"] for x in out), 2)}


@router.get("/paghe/prima-nota")
async def prima_nota(dipendente_id: str):
    """Prima nota salari di un dipendente: tutti i mesi con busta, erogato (bonifici+acconti)
    e saldo progressivo (cumulato busta − cumulato erogato; >0 = ancora da pagare)."""
    db = get_db()
    paghe = await db.paghe_mensili.find({"dipendente_id": dipendente_id}, {"_id": 0}).to_list(2000)
    paghe.sort(key=lambda p: (p.get("anno") or 0, p.get("mese") or 0))
    out, saldo = [], 0.0
    for p in paghe:
        busta = float(p.get("importo_busta") or 0)
        acc = sum(float(a.get("importo") or 0) for a in (p.get("acconti") or []))
        bon = float(p.get("bonifico_importo") or 0)
        erogato = bon + acc
        if busta == 0 and erogato == 0:
            continue
        saldo += busta - erogato
        out.append({"anno": p.get("anno"), "mese": p.get("mese"), "busta": round(busta, 2),
                    "bonifico": round(bon, 2), "acconti": round(acc, 2),
                    "erogato": round(erogato, 2), "saldo_progressivo": round(saldo, 2)})
    return {"righe": out, "saldo_finale": round(saldo, 2)}


@router.get("/paghe/associazioni-bonifici")
async def associazioni_bonifici(anno: Optional[int] = None, mese: Optional[int] = None,
                                stato: Optional[str] = None):
    """Vista UNICA cedolino↔bonifico. Per ogni busta del periodo mostra l'importo busta,
    i bonifici REALMENTE pagati (collezione pagamenti_esiti: data, importo, causale, riferimento/CRO),
    gli acconti, il saldo e lo stato di associazione:
      - pagato            = erogato (bonifici+acconti) ≥ busta
      - parziale          = erogato > 0 ma < busta
      - da_pagare         = busta presente, nessun pagamento
      - bonifico_senza_busta = pagamento presente ma nessuna busta
    Inoltre indica la 'fonte' del bonifico (banca/prima_nota/manuale), la 'qualita' del match
    (esatto/per_importo/aggregato/da_verificare) e se esiste il PDF del cedolino.
    Sorgente dati = sistema vivo paghe_mensili + pagamenti_esiti (nessun sistema parallelo)."""
    db = get_db()
    q = {}
    if anno:
        q["anno"] = int(anno)
    if mese:
        q["mese"] = int(mese)

    dip_map = {}
    async for d in db.dipendenti.find({}, {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "codice_fiscale": 1}):
        dip_map[d["id"]] = d

    righe = []
    tot = {"buste": 0.0, "bonifici": 0.0, "acconti": 0.0, "saldo": 0.0,
           "pagati": 0, "parziali": 0, "da_pagare": 0, "senza_busta": 0,
           "associati": 0, "da_verificare": 0}

    async for p in db.paghe_mensili.find(q, {"_id": 0}):
        busta = float(p.get("importo_busta") or 0)
        bon = float(p.get("bonifico_importo") or 0)
        acc_list = p.get("acconti") or []
        acc = sum(float(a.get("importo") or 0) for a in acc_list)
        if busta <= 0 and bon <= 0 and acc <= 0:
            continue

        dip_id = p.get("dipendente_id")
        dip = dip_map.get(dip_id) or {}
        nome = f"{dip.get('cognome', '')} {dip.get('nome', '')}".strip() or dip_id

        # Bonifici reali pagati (esiti banca) per questo dipendente/mese/anno
        esiti = []
        async for e in db.pagamenti_esiti.find(
                {"dipendente_id": dip_id, "mese": p.get("mese"), "anno": p.get("anno")},
                {"_id": 0}).sort("data", 1):
            esiti.append({
                "data": e.get("data"),
                "importo": round(float(e.get("importo") or 0), 2),
                "causale": e.get("causale") or "",
                "beneficiario": e.get("beneficiario") or "",
                "riferimento": e.get("cro") or e.get("key") or "",
            })

        erogato = bon + acc
        if busta <= 0 and erogato > 0:
            st = "bonifico_senza_busta"
            tot["senza_busta"] += 1
        elif erogato <= 0:
            st = "da_pagare"
            tot["da_pagare"] += 1
        elif erogato + 0.5 >= busta:
            st = "pagato"
            tot["pagati"] += 1
        else:
            st = "parziale"
            tot["parziali"] += 1

        # Fonte del bonifico
        if esiti:
            fonte = "banca"
        elif p.get("bonifico_da_prima_nota"):
            fonte = "prima_nota"
        elif bon > 0:
            fonte = "manuale"
        else:
            fonte = None

        # Qualità dell'associazione (quanto è affidabile il legame busta↔bonifico)
        if bon <= 0:
            qualita = None
        elif esiti:
            if len(esiti) == 1 and busta > 0 and abs(esiti[0]["importo"] - busta) <= 0.5:
                qualita = "esatto"          # un solo bonifico che combacia con la busta
            elif busta > 0 and abs(bon - busta) <= 0.5:
                qualita = "per_importo"     # somma bonifici = busta
            elif len(esiti) > 1:
                qualita = "aggregato"       # più bonifici nello stesso mese
            else:
                qualita = "per_importo"
        else:
            qualita = "da_verificare"       # importo inserito a mano / da prima nota, senza prova banca

        associato = bool(p.get("bonifico_riconciliato")) or qualita in ("esatto", "per_importo")
        if st in ("pagato", "parziale", "bonifico_senza_busta"):
            if associato:
                tot["associati"] += 1
            else:
                tot["da_verificare"] += 1

        # Esiste il PDF del cedolino?
        ced = await db.cedolini.find_one(
            {"dipendente_id": dip_id, "mese": p.get("mese"), "anno": p.get("anno")},
            {"_id": 0, "id": 1, "pdf_data": 1})
        if not ced and dip.get("cognome"):
            ced = await db.cedolini.find_one(
                {"nome_dipendente": {"$regex": dip.get("cognome"), "$options": "i"},
                 "mese": p.get("mese"), "anno": p.get("anno")},
                {"_id": 0, "id": 1, "pdf_data": 1})
        has_pdf = bool(ced and ced.get("pdf_data"))
        cedolino_id = ced.get("id") if ced else None

        if stato and st != stato:
            continue

        tot["buste"] += busta
        tot["bonifici"] += bon
        tot["acconti"] += acc
        tot["saldo"] += (busta - erogato)

        righe.append({
            "dipendente_id": dip_id,
            "dipendente": nome,
            "anno": p.get("anno"),
            "mese": p.get("mese"),
            "busta": round(busta, 2),
            "bonifico": round(bon, 2),
            "acconti": round(acc, 2),
            "erogato": round(erogato, 2),
            "saldo": round(busta - erogato, 2),
            "stato": st,
            "fonte": fonte,
            "qualita": qualita,
            "associato": associato,
            "riconciliato": bool(p.get("bonifico_riconciliato")),
            "bonifico_data": p.get("bonifico_data"),
            "bonifici": esiti,
            "n_bonifici": len(esiti),
            "cedolino_pdf": has_pdf,
            "cedolino_id": cedolino_id,
        })

    righe.sort(key=lambda r: ((r["anno"] or 0), (r["mese"] or 0), r["dipendente"]), reverse=True)
    for k in ("buste", "bonifici", "acconti", "saldo"):
        tot[k] = round(tot[k], 2)
    return {"righe": righe, "totali": tot, "count": len(righe)}


@router.post("/paghe/conferma-associazione")
async def conferma_associazione(data: dict):
    """Conferma/annulla manualmente l'associazione bonifico↔cedolino di una busta.
    Imposta bonifico_riconciliato e traccia data/nota. Non crea record nuovi:
    agisce sul record paghe_mensili esistente (sistema unico)."""
    dip = data.get("dipendente_id")
    anno = data.get("anno")
    mese = data.get("mese")
    if not dip or not anno or not mese:
        raise HTTPException(status_code=400, detail="dipendente_id, anno, mese obbligatori")
    val = bool(data.get("riconciliato", True))
    set_doc = {"bonifico_riconciliato": val, "updated_at": now_iso()}
    if val:
        set_doc["associazione_confermata_at"] = now_iso()
    if data.get("nota") is not None:
        set_doc["associazione_nota"] = str(data.get("nota"))
    res = await get_db().paghe_mensili.update_one(
        {"dipendente_id": dip, "anno": int(anno), "mese": int(mese)}, {"$set": set_doc})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Busta non trovata per quel dipendente/mese")
    return {"ok": True, "riconciliato": val}


# ============ BUSTE PAGA ============

@router.get("/buste-paga")
async def get_buste_paga(anno: Optional[int] = None, mese: Optional[int] = None, dipendente_id: Optional[str] = None):
    """
    Recupera cedolini dalla collezione 'cedolini' (dati storici dal 2014).
    Se dipendente_id è fornito, cerca anche per nome del dipendente in dipendenti_cloud.
    """
    query = {}
    if anno:
        query["anno"] = anno
    if mese:
        query["mese"] = mese
    
    # Se abbiamo un dipendente_id, dobbiamo trovare il nome per cercare nei cedolini
    if dipendente_id:
        dip = await get_db().dipendenti_cloud.find_one({"id": dipendente_id})
        if dip:
            nome_completo = f"{dip.get('nome', '')} {dip.get('cognome', '')}".strip().upper()
            query["$or"] = [
                {"dipendente_id": dipendente_id},
                {"nome_dipendente": {"$regex": dip.get('cognome', ''), "$options": "i"}}
            ]
    
    # Leggi dalla collezione cedolini (dati storici)
    cedolini = await get_db().cedolini.find(query, {"_id": 0}).sort([("anno", -1), ("mese", -1)]).to_list(1000)
    
    # Normalizza i campi per compatibilità con il frontend
    result = []
    for c in cedolini:
        result.append({
            "id": c.get("id", str(c.get("_id", ""))),
            "dipendente_id": c.get("dipendente_id", ""),
            "dipendente_nome": c.get("nome_dipendente") or c.get("dipendente_nome") or "",
            "mese": c.get("mese"),
            "anno": c.get("anno"),
            "lordo": c.get("lordo", 0),
            "netto": c.get("netto", 0),
            "inps": c.get("inps_dipendente", 0),
            "irpef": c.get("irpef", 0),
            "trattenute": c.get("trattenute", 0),
            "stato": c.get("stato_pagamento") or c.get("stato") or "DA_PAGARE",
            "created_at": c.get("created_at", "")
        })
    
    return result

@router.post("/buste-paga")
async def create_busta_paga(busta: BustaPagaCloud):
    busta_dict = busta.model_dump()
    busta_dict["id"] = generate_id()
    busta_dict["created_at"] = now_iso()
    await get_db().buste_paga_cloud.insert_one(busta_dict)
    return serialize_doc(busta_dict)

@router.post("/buste-paga/genera")
async def genera_buste_paga(data: dict):
    """Genera buste paga per tutti i dipendenti attivi per un mese specifico"""
    mese = data.get("mese")
    anno = data.get("anno")
    lordo_default = data.get("lordo", 1500)
    
    if not mese or not anno:
        raise HTTPException(status_code=400, detail="mese e anno sono obbligatori")
    
    dipendenti = await get_db().dipendenti_cloud.find({"stato": "attivo"}, {"_id": 0}).to_list(1000)
    created = 0
    
    for dip in dipendenti:
        existing = await get_db().buste_paga_cloud.find_one({
            "dipendente_id": dip["id"],
            "mese": mese,
            "anno": anno
        })
        
        if not existing:
            inps = round(lordo_default * 0.0919, 2)
            irpef = round((lordo_default - inps) * 0.23, 2)
            netto = round(lordo_default - inps - irpef, 2)
            
            busta = {
                "id": generate_id(),
                "dipendente_id": dip["id"],
                "mese": mese,
                "anno": anno,
                "lordo": lordo_default,
                "inps": inps,
                "irpef": irpef,
                "trattenute": 0,
                "netto": netto,
                "stato": "DA_PAGARE",
                "created_at": now_iso()
            }
            await get_db().buste_paga_cloud.insert_one(busta)
            created += 1
    
    return {"message": f"Generate {created} buste paga"}

@router.put("/buste-paga/{busta_id}/paga")
async def paga_busta(busta_id: str):
    result = await get_db().buste_paga_cloud.update_one(
        {"id": busta_id},
        {"$set": {"stato": "PAGATO", "data_pagamento": now_iso()}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Busta paga non trovata")
    return {"message": "Busta paga marcata come pagata"}

# ============ MISSIONI ============

@router.get("/missioni")
async def get_missioni(dipendente_id: Optional[str] = None, stato: Optional[str] = None):
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    if stato:
        query["stato"] = stato
    missioni = await get_db().missioni_cloud.find(query, {"_id": 0}).to_list(1000)
    return missioni

@router.post("/missioni")
async def create_missione(missione: MissioneCloud):
    miss_dict = missione.model_dump()
    miss_dict["id"] = generate_id()
    miss_dict["created_at"] = now_iso()
    await get_db().missioni_cloud.insert_one(miss_dict)
    return serialize_doc(miss_dict)

@router.put("/missioni/{missione_id}/approva")
async def approva_missione(missione_id: str):
    db = get_db()
    miss = await db.missioni_cloud.find_one({"id": missione_id}, {"_id": 0})
    if not miss:
        raise HTTPException(status_code=404, detail="Missione non trovata")
    await db.missioni_cloud.update_one(
        {"id": missione_id}, {"$set": {"stato": "approvata", "approvata_il": now_iso()}})

    automazioni = []
    rimborso = float(miss.get("rimborso") or 0)
    dip_id = miss.get("dipendente_id")
    dip = await db.dipendenti.find_one({"id": dip_id}, {"_id": 0, "nome_completo": 1, "nome": 1, "cognome": 1}) if dip_id else None
    nome = (dip or {}).get("nome_completo") or (f"{(dip or {}).get('cognome','')} {(dip or {}).get('nome','')}".strip() if dip else "")

    # Rimborso missione → partita aperta (tracciamento finanziario)
    if rimborso > 0 and dip_id:
        try:
            from backend.app.services.partite_aperte_engine import crea_partita, TipoPartita
            await crea_partita(
                tipo=TipoPartita.ALTRO, documento_id=missione_id,
                documento_collection="missioni_cloud", controparte_id=dip_id,
                controparte_nome=nome, importo=rimborso, db=db, data_documento=now_iso()[:10],
                extra={"categoria": "rimborso_missione", "destinazione": miss.get("destinazione")})
            automazioni.append("partita_rimborso")
        except Exception:
            pass
    # Notifica al dipendente
    if dip_id:
        try:
            from backend.app.services.notifiche import crea_notifica
            await crea_notifica(db, dip_id, "missione", "Missione approvata",
                                f"La missione a {miss.get('destinazione','')} è stata approvata"
                                + (f" · rimborso € {rimborso:.2f}" if rimborso > 0 else "") + ".",
                                extra={"missione_id": missione_id})
            automazioni.append("notifica_dipendente")
        except Exception:
            pass
    return {"message": "Missione approvata", "automazioni": automazioni}

@router.delete("/missioni/{missione_id}")
async def delete_missione(missione_id: str):
    result = await get_db().missioni_cloud.delete_one({"id": missione_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Missione non trovata")
    return {"message": "Missione eliminata"}

# ============ DOCUMENTI ============

@router.get("/documenti")
async def get_documenti(dipendente_id: Optional[str] = None):
    query = {}
    if dipendente_id:
        query["dipendente_id"] = dipendente_id
    documenti = await get_db().documenti_cloud.find(query, {"_id": 0}).to_list(1000)
    return documenti

@router.post("/documenti")
async def create_documento(doc: DocumentoCloud):
    doc_dict = doc.model_dump()
    doc_dict["id"] = generate_id()
    doc_dict["data_caricamento"] = now_iso()
    await get_db().documenti_cloud.insert_one(doc_dict)
    return serialize_doc(doc_dict)

@router.delete("/documenti/{documento_id}")
async def delete_documento(documento_id: str):
    result = await get_db().documenti_cloud.delete_one({"id": documento_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Documento non trovato")
    return {"message": "Documento eliminato"}


_CF_DOC_RE = re.compile(r'\b([A-Z]{6}\d{2}[A-Z]\d{2}[A-Z]\d{3}[A-Z])\b')
CATEGORIE_DOC = ["UNILAV", "CERTIFICAZIONE_UNICA", "CONTRATTO", "RIDUZIONE_ORARIO", "BONIFICO",
                 "CODICE_FISCALE", "CARTA_IDENTITA", "BUSTA_PAGA", "ALTRO"]


def classifica_documento(text: str, filename: str = "") -> str:
    """Riconosce il tipo di documento dal testo e, in fallback, dal nome del file
    (utile per le scansioni-immagine senza testo). Diciture standard italiane."""
    t = (text or "").lower()
    fn = (filename or "").lower()

    def H(s, *ks):
        return any(k in s for k in ks)
    # 1) Segnali forti dal TESTO
    if H(t, "unilav", "comunicazione obbligatoria", "modello unificato lav", "centro per l'impiego"):
        return "UNILAV"
    if H(t, "certificazione unica", "redditi di lavoro dipendente e assimilati"):
        return "CERTIFICAZIONE_UNICA"
    if H(t, "contratto individuale di lavoro", "contratto di lavoro", "patto di prova", "lettera di assunzione"):
        return "CONTRATTO"
    if H(t, "bonifico", "ordinante", "beneficiario", "disposizione di pagamento", "sepa credit"):
        return "BONIFICO"
    if H(t, "busta paga", "cedolino", "netto in busta", "retribuzione lorda"):
        return "BUSTA_PAGA"
    if H(t, "carta di identità", "carta d'identità", "documento di identità", "carta d identita"):
        return "CARTA_IDENTITA"
    if H(t, "riduzione orario", "riduzione dell'orario", "riduzione dell orario", "trasformazione part-time", "riduzione part time"):
        return "RIDUZIONE_ORARIO"
    # 2) Nome FILE (per scansioni senza testo)
    if H(fn, "riduzione"):
        return "RIDUZIONE_ORARIO"
    if H(fn, "unilav"):
        return "UNILAV"
    if H(fn, "certificazione_unica", "certificazione unica", "_cu_", "cud"):
        return "CERTIFICAZIONE_UNICA"
    if H(fn, "contratto"):
        return "CONTRATTO"
    if H(fn, "bonific"):
        return "BONIFICO"
    if H(fn, "carta_di_identit", "carta d'identit", "carta identit", "carta_identit"):
        return "CARTA_IDENTITA"
    if H(fn, "codice_fiscale", "codice fiscale", "tessera_sanitaria", "tessera sanitaria"):
        return "CODICE_FISCALE"
    if H(fn, "busta", "cedolino"):
        return "BUSTA_PAGA"
    # 3) Segnale debole dal testo
    if H(t, "tessera sanitaria", "servizio sanitario nazionale"):
        return "CODICE_FISCALE"
    return "ALTRO"


@router.post("/documenti/upload-massivo")
async def upload_documenti_massivo(files: List[UploadFile] = File(...)):
    """Carica più documenti insieme: per ognuno riconosce il tipo (UNILAV, C.U., contratto,
    bonifico, codice fiscale…), trova il dipendente dal codice fiscale (o dal nome) nel testo,
    e lo archivia nella sua cartella. Anti-duplicati per hash del file."""
    import io
    import pdfplumber
    db = get_db()
    dips = await db.dipendenti.find({"merged_into": {"$exists": False}},
                                    {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1, "codice_fiscale": 1}).to_list(1000)

    def norm(s):
        return re.sub(r"\s+", " ", str(s or "").strip()).lower()
    by_cf, by_nome, by_cogn = {}, {}, {}
    for d in dips:
        cf = (d.get("codice_fiscale") or "").upper().strip()
        if cf:
            by_cf[cf] = d
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for v in {norm(d.get("nome_completo")), f"{c} {n}".strip(), f"{n} {c}".strip()}:
            if v and len(v) > 6:
                by_nome[v] = d
        if len(c) >= 4:
            by_cogn.setdefault(c, []).append(d)

    import zipfile
    caricati, duplicati, non_assegnati, per_categoria = [], [], [], {}

    async def processa(filename, raw, contesto=""):
        if not raw:
            return
        h = hashlib.sha256(raw).hexdigest()
        if await db.documenti_cloud.find_one({"hash": h}):
            duplicati.append(filename)
            return
        text = ""
        if raw[:4] == b"%PDF":
            try:
                with pdfplumber.open(io.BytesIO(raw)) as pdf:
                    for p in pdf.pages[:6]:
                        text += (p.extract_text() or "") + "\n"
            except Exception:
                text = ""
        categoria = classifica_documento(text, f"{contesto} {filename}".strip())
        d = None
        for cf in _CF_DOC_RE.findall((text or "").upper()):
            if cf in by_cf:
                d = by_cf[cf]
                break
        if not d:
            tl = norm(text)
            for nome_n, dd in by_nome.items():
                if nome_n in tl:
                    d = dd
                    break
        if not d:
            fn_norm = norm(f"{contesto} {filename}".replace("_", " ").replace("-", " "))
            for nome_n, dd in by_nome.items():
                if nome_n in fn_norm:
                    d = dd
                    break
            if not d:
                for cogn, lst in by_cogn.items():
                    if cogn in fn_norm and len(lst) == 1:
                        d = lst[0]
                        break
        doc = {"id": generate_id(),
               "dipendente_id": (d or {}).get("id"),
               "dipendente_nome": (f"{d.get('cognome','')} {d.get('nome','')}".strip() if d else None),
               "titolo": filename, "filename": filename,
               "tipo": categoria, "categoria": categoria, "hash": h,
               "file_data": base64.b64encode(raw).decode(),
               "assegnato": bool(d), "origine": "upload_massivo", "data_caricamento": now_iso()}
        await db.documenti_cloud.insert_one(doc)
        per_categoria[categoria] = per_categoria.get(categoria, 0) + 1
        (caricati if d else non_assegnati).append({"file": filename, "categoria": categoria, "dipendente": doc["dipendente_nome"]})

    for f in files:
        raw = await f.read()
        fn = f.filename or ""
        if fn.lower().endswith(".zip"):
            try:
                with zipfile.ZipFile(io.BytesIO(raw)) as zf:
                    for nm in zf.namelist():
                        if nm.endswith("/"):
                            continue
                        await processa(nm.split("/")[-1], zf.read(nm), contesto=fn)
            except zipfile.BadZipFile:
                non_assegnati.append({"file": fn, "categoria": "ALTRO"})
        else:
            await processa(fn, raw, contesto="")
    return {"caricati": len(caricati), "duplicati": duplicati,
            "non_assegnati": non_assegnati, "per_categoria": per_categoria,
            "dettaglio": caricati[:300]}


@router.get("/documenti/{documento_id}/file")
async def download_documento(documento_id: str):
    from fastapi.responses import Response
    doc = await get_db().documenti_cloud.find_one({"id": documento_id}, {"_id": 0})
    if not doc or not doc.get("file_data"):
        raise HTTPException(status_code=404, detail="File non disponibile")
    data = base64.b64decode(doc["file_data"])
    fn = doc.get("filename") or "documento.pdf"
    media = "application/pdf" if fn.lower().endswith(".pdf") else "application/octet-stream"
    return Response(content=data, media_type=media, headers={"Content-Disposition": f'inline; filename="{fn}"'})

# ============ DASHBOARD STATS ============

@router.get("/dashboard/stats")
async def get_dashboard_stats():
    dipendenti = await get_db().dipendenti.find({}, {"_id": 0}).to_list(1000)
    attivi = [d for d in dipendenti if d.get("attivo", True) is not False and d.get("stato", "attivo") not in ("cessato", "disattivo", "inattivo")]
    
    ferie_pending = await get_db().ferie_cloud.count_documents({"stato": "in_attesa"})
    missioni_pending = await get_db().missioni_cloud.count_documents({"stato": "in_attesa"})
    
    # Presenze oggi
    today = datetime.now().strftime("%Y-%m-%d")
    presenze_oggi = await get_db().presenze_cloud.count_documents({"data": today, "stato": "presente"})

    # Solo alert HR: la collezione `alerts` è condivisa con l'ERP contabile, qui
    # mostriamo soltanto i moduli del personale (niente fatture/fornitori/banca…).
    alert_aperti = await get_db().alerts.count_documents(
        {"stato": "aperto", "modulo": {"$in": MODULI_HR}})

    # Buste in attesa di pagamento (motore unico): busta presente ma non ancora pagata
    buste_attesa = 0
    importo_attesa = 0.0
    async for p in get_db().paghe_mensili.find(
            {"stato_pagamento": {"$in": ["in_attesa_pagamento", "parziale"]}},
            {"_id": 0, "saldo": 1, "importo_busta": 1, "bonifico_importo": 1}):
        saldo = p.get("saldo")
        if saldo is None:
            saldo = float(p.get("importo_busta") or 0) - float(p.get("bonifico_importo") or 0)
        if saldo and saldo > 0.5:
            buste_attesa += 1
            importo_attesa += saldo

    return {
        "totale_dipendenti": len(dipendenti),
        "dipendenti_attivi": len(attivi),
        "ferie_in_attesa": ferie_pending,
        "missioni_in_attesa": missioni_pending,
        "presenze_oggi": presenze_oggi,
        "alert_aperti": alert_aperti,
        "buste_in_attesa": buste_attesa,
        "importo_in_attesa": round(importo_attesa, 2),
    }


# Moduli di competenza HR (gli altri appartengono all'ERP contabile OpenClaw).
MODULI_HR = ["dipendenti", "cedolini"]


@router.get("/alerts")
async def lista_alert(modulo: str = "", severita: str = "", stato: str = "aperto"):
    """Elenco degli alert HR (scadenze contratti/prova, contestazioni…).
    La collezione `alerts` è condivisa con la contabilità: qui filtriamo ai soli
    moduli del personale. `stato`: 'aperto' (default), 'risolto' (archivio),
    'tutti' (entrambi). Gli alert risolti NON vengono cancellati: restano in archivio."""
    q = {}
    if stato and stato != "tutti":
        q["stato"] = stato
    if modulo:
        if modulo not in MODULI_HR:
            return {"totale": 0, "alerts": []}
        q["modulo"] = modulo
    else:
        q["modulo"] = {"$in": MODULI_HR}
    if severita:
        q["severita"] = severita
    sort_field = "resolved_at" if stato == "risolto" else "created_at"
    alerts = await get_db().alerts.find(q, {"_id": 0}).sort(sort_field, -1).to_list(500)
    return {"totale": len(alerts), "alerts": alerts}


@router.post("/alerts/{alert_id}/risolvi")
async def risolvi_alert_id(alert_id: str):
    """Segna un alert come risolto (manuale)."""
    r = await get_db().alerts.update_one(
        {"id": alert_id, "stato": "aperto"},
        {"$set": {"stato": "risolto", "risolto": True,
                  "resolved_at": now_iso(), "resolved_by": "admin"}})
    if r.matched_count == 0:
        raise HTTPException(status_code=404, detail="Alert non trovato o già risolto")
    return {"ok": True, "stato": "risolto"}

# ============ SEED DATA ============

@router.post("/seed-data")
async def seed_data():
    """Crea dati di esempio se non esistono"""
    existing = await get_db().dipendenti_cloud.count_documents({})
    if existing > 0:
        return {"message": "Dati già presenti"}
    
    # Crea dipendenti di esempio
    dipendenti_sample = [
        {"nome": "Mario", "cognome": "Rossi", "ruolo": "Responsabile", "stato": "attivo", "contratto": "Indeterminato"},
        {"nome": "Lucia", "cognome": "Bianchi", "ruolo": "Cameriere", "stato": "attivo", "contratto": "Determinato"},
        {"nome": "Giuseppe", "cognome": "Verdi", "ruolo": "Barista", "stato": "attivo", "contratto": "Indeterminato"},
    ]
    
    for d in dipendenti_sample:
        d["id"] = generate_id()
        d["created_at"] = now_iso()
        await get_db().dipendenti_cloud.insert_one(d)
    
    # Crea turni di esempio
    turni_sample = [
        {"nome": "Mattina", "orario_inizio": "06:00", "orario_fine": "14:00", "colore": "#3b82f6"},
        {"nome": "Pomeriggio", "orario_inizio": "14:00", "orario_fine": "22:00", "colore": "#10b981"},
        {"nome": "Notte", "orario_inizio": "22:00", "orario_fine": "06:00", "colore": "#8b5cf6"},
    ]
    
    for t in turni_sample:
        t["id"] = generate_id()
        await get_db().turni_cloud.insert_one(t)
    
    return {"message": "Dati di esempio creati"}
