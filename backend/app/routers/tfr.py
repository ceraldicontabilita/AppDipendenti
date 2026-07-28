"""
Router TFR - Gestione Trattamento Fine Rapporto
Accantonamento, rivalutazione ISTAT, liquidazione TFR e gestione acconti
"""
from fastapi import APIRouter, HTTPException, Query, Body, UploadFile, File
from pydantic import BaseModel
from typing import Dict, Any, Optional, List
from datetime import datetime, timezone, timedelta, date
from uuid import uuid4
import logging
import os
from pathlib import Path

from backend.app.database import Database
from backend.app.utils.error_handler import handle_errors

router = APIRouter()
logger = logging.getLogger(__name__)

# Cartella upload buste paga
PAYSLIPS_FOLDER = "/app/uploads/paghe"

# ============================================
# COSTANTI TFR
# ============================================

# Divisore per calcolo quota annuale TFR (art. 2120 c.c.)
TFR_DIVISORE = 13.5

# Rivalutazione minima ISTAT
RIVALUTAZIONE_FISSA = 1.5  # 1.5% fisso

# Aliquota tassazione separata TFR (approssimata al 23% per semplicità)
ALIQUOTA_TFR = 23.0


# ============================================
# MODELLI
# ============================================

class AccantonamentoTFRInput(BaseModel):
    dipendente_id: str
    anno: int
    retribuzione_annua: float
    indice_istat: Optional[float] = 0.0  # percentuale variazione ISTAT


class LiquidazioneTFRInput(BaseModel):
    dipendente_id: str
    data_liquidazione: str  # YYYY-MM-DD
    motivo: str  # "dimissioni", "licenziamento", "pensionamento", "anticipo"
    importo_richiesto: Optional[float] = None  # se anticipo, importo parziale
    note: Optional[str] = ""


# ============================================
# ENDPOINT
# ============================================

@router.get("/situazione/{dipendente_id}")
@handle_errors
async def get_situazione_tfr(dipendente_id: str) -> Dict[str, Any]:
    """
    Restituisce la situazione TFR completa di un dipendente.
    Include storico accantonamenti e rivalutazioni.
    """
    db = Database.get_db()
    
    # Recupera dipendente
    dipendente = await db["dipendenti"].find_one(
        {"id": dipendente_id},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # TFR accantonato totale
    tfr_accantonato = float(dipendente.get("tfr_accantonato", 0))
    
    # Storico accantonamenti
    accantonamenti = await db["tfr_accantonamenti"].find(
        {"dipendente_id": dipendente_id},
        {"_id": 0}
    ).sort("anno", 1).to_list(100)
    
    # Storico liquidazioni/anticipi
    liquidazioni = await db["tfr_liquidazioni"].find(
        {"dipendente_id": dipendente_id},
        {"_id": 0}
    ).sort("data", -1).to_list(100)
    
    totale_liquidato = sum(l.get("importo_lordo", 0) for l in liquidazioni)
    
    return {
        "dipendente_id": dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "tfr_accantonato": round(tfr_accantonato, 2),
        "tfr_disponibile": round(tfr_accantonato - totale_liquidato, 2),
        "totale_liquidato": round(totale_liquidato, 2),
        "num_accantonamenti": len(accantonamenti),
        "accantonamenti": accantonamenti,
        "liquidazioni": liquidazioni
    }


@router.post("/accantonamento")
@handle_errors
async def registra_accantonamento_tfr(input_data: AccantonamentoTFRInput) -> Dict[str, Any]:
    """
    Registra l'accantonamento annuale TFR per un dipendente.
    
    Formula TFR (Art. 2120 Codice Civile):
    - Quota annuale = Retribuzione annua / 13.5
    - Rivalutazione = TFR accumulato precedente * (1.5% + 75% * indice ISTAT)
    - Imposta sostitutiva rivalutazione: 17% (non applicata qui, gestita in sede fiscale)
    
    Args:
        input_data: Dati per il calcolo (dipendente_id, anno, retribuzione_annua, indice_istat)
    
    Returns:
        Dettaglio dell'accantonamento registrato
    """
    db = Database.get_db()
    
    if input_data.retribuzione_annua <= 0:
        raise HTTPException(status_code=400, detail="La retribuzione annua deve essere positiva")
    
    if input_data.anno < 2020 or input_data.anno > 2030:
        raise HTTPException(status_code=400, detail="Anno non valido")
    
    # Recupera dipendente
    dipendente = await db["dipendenti"].find_one(
        {"id": input_data.dipendente_id},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Calcola quota annuale
    quota_annuale = input_data.retribuzione_annua / TFR_DIVISORE
    
    # Calcola rivalutazione sul TFR accumulato precedente
    tfr_precedente = float(dipendente.get("tfr_accantonato", 0))
    # Art. 2120 c.c.: 1.5% fisso + 75% dell'indice ISTAT
    tasso_rivalutazione = (RIVALUTAZIONE_FISSA + input_data.indice_istat * 0.75) / 100
    rivalutazione = tfr_precedente * tasso_rivalutazione
    
    # Totale accantonamento anno
    totale_accantonamento = quota_annuale + rivalutazione
    
    # Nuovo TFR totale
    nuovo_tfr_totale = tfr_precedente + totale_accantonamento
    
    # Registra accantonamento
    accantonamento = {
        "id": str(uuid4()),
        "dipendente_id": input_data.dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "anno": input_data.anno,
        "retribuzione_annua": input_data.retribuzione_annua,
        "quota_annuale": round(quota_annuale, 2),
        "tfr_precedente": round(tfr_precedente, 2),
        "indice_istat": input_data.indice_istat,
        "tasso_rivalutazione": round(tasso_rivalutazione * 100, 2),
        "rivalutazione": round(rivalutazione, 2),
        "totale_accantonamento": round(totale_accantonamento, 2),
        "nuovo_tfr_totale": round(nuovo_tfr_totale, 2),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db["tfr_accantonamenti"].insert_one(accantonamento.copy())
    
    # Aggiorna TFR dipendente
    await db["dipendenti"].update_one(
        {"id": input_data.dipendente_id},
        {"$set": {"tfr_accantonato": round(nuovo_tfr_totale, 2)}}
    )
    
    # Registra movimento contabile
    movimento = {
        "id": str(uuid4()),
        "data": f"{input_data.anno}-12-31",
        "descrizione": f"Accantonamento TFR {input_data.anno} - {dipendente.get('nome_completo', '')}",
        "tipo": "tfr_accantonamento",
        "importo": round(totale_accantonamento, 2),
        "dipendente_id": input_data.dipendente_id,
        "anno": input_data.anno,
        "dettaglio": {
            "quota_annuale": round(quota_annuale, 2),
            "rivalutazione": round(rivalutazione, 2)
        },
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db["movimenti_contabili"].insert_one(movimento.copy())
    
    return {
        "success": True,
        "accantonamento_id": accantonamento["id"],
        "messaggio": f"TFR {input_data.anno} accantonato per {dipendente.get('nome_completo', '')}",
        "dettaglio": {
            "quota_annuale": round(quota_annuale, 2),
            "rivalutazione": round(rivalutazione, 2),
            "totale_accantonato": round(totale_accantonamento, 2),
            "nuovo_tfr_totale": round(nuovo_tfr_totale, 2)
        }
    }


@router.post("/liquidazione")
@handle_errors
async def liquida_tfr(input_data: LiquidazioneTFRInput) -> Dict[str, Any]:
    """
    Liquida il TFR a un dipendente (totale o parziale).
    
    Calcola ritenute fiscali con tassazione separata (Art. 19 TUIR).
    L'aliquota media è calcolata come approssimazione semplificata al 23%.
    Per anticipi (max 70% del TFR maturato, Art. 2120 c.c. comma 6-8).
    
    Args:
        input_data: Dati liquidazione (dipendente_id, data, motivo, importo)
    
    Returns:
        Dettaglio della liquidazione con importo lordo, ritenute e netto
    """
    db = Database.get_db()
    
    # Recupera dipendente
    dipendente = await db["dipendenti"].find_one(
        {"id": input_data.dipendente_id},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    tfr_disponibile = float(dipendente.get("tfr_accantonato", 0))
    
    # Determina importo da liquidare
    # Per anticipo è obbligatorio indicare un importo richiesto positivo.
    if input_data.motivo == "anticipo":
        if input_data.importo_richiesto is None:
            raise HTTPException(
                status_code=400,
                detail="Per anticipo TFR è obbligatorio specificare un importo richiesto"
            )
        if input_data.importo_richiesto <= 0:
            raise HTTPException(
                status_code=400,
                detail="Per anticipo TFR l'importo richiesto deve essere maggiore di zero"
            )

        max_anticipo = tfr_disponibile * 0.70
        if input_data.importo_richiesto > max_anticipo:
            raise HTTPException(
                status_code=400,
                detail=f"Anticipo TFR max 70%: richiesto €{input_data.importo_richiesto:.2f}, "
                       f"massimo consentito €{max_anticipo:.2f} (Art. 2120 c.c.)"
            )
        importo_lordo = min(input_data.importo_richiesto, tfr_disponibile)
    elif input_data.importo_richiesto is not None:
        if input_data.importo_richiesto <= 0:
            raise HTTPException(
                status_code=400,
                detail="L'importo richiesto deve essere maggiore di zero"
            )
        importo_lordo = min(input_data.importo_richiesto, tfr_disponibile)
    else:
        importo_lordo = tfr_disponibile
    
    if importo_lordo <= 0:
        raise HTTPException(status_code=400, detail="Nessun TFR disponibile da liquidare")
    
    # Calcola ritenute (tassazione separata semplificata)
    ritenute = importo_lordo * ALIQUOTA_TFR / 100
    importo_netto = importo_lordo - ritenute
    
    # Registra liquidazione
    liquidazione = {
        "id": str(uuid4()),
        "dipendente_id": input_data.dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "data": input_data.data_liquidazione,
        "motivo": input_data.motivo,
        "tfr_precedente": round(tfr_disponibile, 2),
        "importo_lordo": round(importo_lordo, 2),
        "aliquota_ritenuta": ALIQUOTA_TFR,
        "ritenute": round(ritenute, 2),
        "importo_netto": round(importo_netto, 2),
        "tfr_residuo": round(tfr_disponibile - importo_lordo, 2),
        "note": input_data.note,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db["tfr_liquidazioni"].insert_one(liquidazione.copy())
    
    # Aggiorna TFR dipendente
    nuovo_tfr = tfr_disponibile - importo_lordo
    await db["dipendenti"].update_one(
        {"id": input_data.dipendente_id},
        {"$set": {"tfr_accantonato": round(nuovo_tfr, 2)}}
    )
    
    # Registra movimenti contabili
    # 1. Utilizzo fondo TFR
    movimento_fondo = {
        "id": str(uuid4()),
        "data": input_data.data_liquidazione,
        "descrizione": f"Liquidazione TFR - {dipendente.get('nome_completo', '')}",
        "tipo": "tfr_liquidazione",
        "importo": round(importo_lordo, 2),
        "dipendente_id": input_data.dipendente_id,
        "motivo": input_data.motivo,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db["movimenti_contabili"].insert_one(movimento_fondo.copy())
    
    # 2. Ritenute
    if ritenute > 0:
        movimento_ritenute = {
            "id": str(uuid4()),
            "data": input_data.data_liquidazione,
            "descrizione": f"Ritenute TFR - {dipendente.get('nome_completo', '')}",
            "tipo": "ritenuta_tfr",
            "importo": round(ritenute, 2),
            "dipendente_id": input_data.dipendente_id,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db["movimenti_contabili"].insert_one(movimento_ritenute.copy())
    
    return {
        "success": True,
        "liquidazione_id": liquidazione["id"],
        "messaggio": f"TFR liquidato per {dipendente.get('nome_completo', '')}",
        "dettaglio": {
            "importo_lordo": round(importo_lordo, 2),
            "ritenute": round(ritenute, 2),
            "importo_netto": round(importo_netto, 2),
            "tfr_residuo": round(nuovo_tfr, 2)
        }
    }


@router.get("/riepilogo-aziendale")
@handle_errors
async def get_riepilogo_tfr_aziendale(anno: int = Query(None)) -> Dict[str, Any]:
    """
    Riepilogo TFR per tutti i dipendenti attivi.
    Utile per il bilancio e pianificazione finanziaria.
    """
    db = Database.get_db()
    
    if not anno:
        anno = datetime.now().year
    
    # Dipendenti attivi
    dipendenti = await db["dipendenti"].find(
        {"status": {"$in": ["attivo", "active"]}},
        {"_id": 0, "id": 1, "nome_completo": 1, "tfr_accantonato": 1}
    ).to_list(1000)
    
    # Accantonamenti dell'anno
    accantonamenti_anno = await db["tfr_accantonamenti"].aggregate([
        {"$match": {"anno": anno}},
        {"$group": {
            "_id": None,
            "totale_quota": {"$sum": "$quota_annuale"},
            "totale_rivalutazione": {"$sum": "$rivalutazione"},
            "totale_accantonato": {"$sum": "$totale_accantonamento"},
            "num_dipendenti": {"$sum": 1}
        }}
    ]).to_list(1)
    
    # Liquidazioni dell'anno
    liquidazioni_anno = await db["tfr_liquidazioni"].aggregate([
        {"$match": {"data": {"$regex": f"^{anno}"}}},
        {"$group": {
            "_id": None,
            "totale_lordo": {"$sum": "$importo_lordo"},
            "totale_ritenute": {"$sum": "$ritenute"},
            "totale_netto": {"$sum": "$importo_netto"},
            "num_liquidazioni": {"$sum": 1}
        }}
    ]).to_list(1)
    
    # Totale fondo TFR
    totale_fondo = sum(float(d.get("tfr_accantonato", 0)) for d in dipendenti)
    
    # Dettaglio per dipendente
    dettaglio_dipendenti = [
        {
            "dipendente_id": d["id"],
            "nome": d.get("nome_completo", ""),
            "tfr_accantonato": round(float(d.get("tfr_accantonato", 0)), 2)
        }
        for d in dipendenti
        if float(d.get("tfr_accantonato", 0)) > 0
    ]
    
    return {
        "anno": anno,
        "totale_fondo_tfr": round(totale_fondo, 2),
        "num_dipendenti_attivi": len(dipendenti),
        "accantonamenti_anno": {
            "totale_quota": round(accantonamenti_anno[0]["totale_quota"], 2) if accantonamenti_anno else 0,
            "totale_rivalutazione": round(accantonamenti_anno[0]["totale_rivalutazione"], 2) if accantonamenti_anno else 0,
            "totale_accantonato": round(accantonamenti_anno[0]["totale_accantonato"], 2) if accantonamenti_anno else 0,
            "num_dipendenti": accantonamenti_anno[0]["num_dipendenti"] if accantonamenti_anno else 0
        },
        "liquidazioni_anno": {
            "totale_lordo": round(liquidazioni_anno[0]["totale_lordo"], 2) if liquidazioni_anno else 0,
            "totale_ritenute": round(liquidazioni_anno[0]["totale_ritenute"], 2) if liquidazioni_anno else 0,
            "totale_netto": round(liquidazioni_anno[0]["totale_netto"], 2) if liquidazioni_anno else 0,
            "num_liquidazioni": liquidazioni_anno[0]["num_liquidazioni"] if liquidazioni_anno else 0
        },
        "dettaglio_dipendenti": sorted(dettaglio_dipendenti, key=lambda x: x["tfr_accantonato"], reverse=True)
    }


@router.post("/calcola-batch/{anno}")
@handle_errors
async def calcola_tfr_batch(anno: int) -> Dict[str, Any]:
    """
    Calcola TFR per tutti i dipendenti attivi per l'anno specificato.
    Usa i dati dei cedolini per determinare la retribuzione annua.
    """
    db = Database.get_db()
    
    # Dipendenti attivi
    dipendenti = await db["dipendenti"].find(
        {"status": {"$in": ["attivo", "active"]}},
        {"_id": 0}
    ).to_list(1000)
    
    risultati = []
    
    for dip in dipendenti:
        dip_id = dip["id"]
        
        # Verifica se già calcolato per quest'anno
        esistente = await db["tfr_accantonamenti"].find_one({
            "dipendente_id": dip_id,
            "anno": anno
        })
        
        if esistente:
            risultati.append({
                "dipendente": dip.get("nome_completo", ""),
                "stato": "già_calcolato",
                "importo": esistente.get("totale_accantonamento", 0)
            })
            continue
        
        # Calcola retribuzione annua dai cedolini
        cedolini = await db["cedolini"].aggregate([
            {"$match": {"dipendente_id": dip_id, "anno": anno}},
            {"$group": {"_id": None, "totale_lordo": {"$sum": "$lordo"}}}
        ]).to_list(1)
        
        if not cedolini or cedolini[0]["totale_lordo"] == 0:
            # Se non ci sono cedolini, usa una stima dalla prima nota salari
            salari = await db["prima_nota_salari"].aggregate([
                {"$match": {
                    "$or": [
                        {"dipendente_id": dip_id},
                        {"dipendente": dip.get("nome_completo", "")}
                    ],
                    "anno": anno
                }},
                {"$group": {"_id": None, "totale": {"$sum": "$importo_lordo"}}}
            ]).to_list(1)
            
            retribuzione_annua = salari[0]["totale"] if salari else 0
        else:
            retribuzione_annua = cedolini[0]["totale_lordo"]
        
        if retribuzione_annua <= 0:
            risultati.append({
                "dipendente": dip.get("nome_completo", ""),
                "stato": "nessuna_retribuzione",
                "importo": 0
            })
            continue
        
        # Registra accantonamento
        input_data = AccantonamentoTFRInput(
            dipendente_id=dip_id,
            anno=anno,
            retribuzione_annua=retribuzione_annua,
            indice_istat=0  # Può essere aggiornato con indice reale
        )
        
        try:
            result = await registra_accantonamento_tfr(input_data)
            risultati.append({
                "dipendente": dip.get("nome_completo", ""),
                "stato": "calcolato",
                "importo": result["dettaglio"]["totale_accantonato"]
            })
        except Exception as e:
            risultati.append({
                "dipendente": dip.get("nome_completo", ""),
                "stato": "errore",
                "errore": str(e)
            })
    
    totale_accantonato = sum(r.get("importo", 0) for r in risultati if r["stato"] == "calcolato")
    
    return {
        "anno": anno,
        "risultati": risultati,
        "totale_nuovo_accantonato": round(totale_accantonato, 2),
        "num_calcolati": len([r for r in risultati if r["stato"] == "calcolato"]),
        "num_già_esistenti": len([r for r in risultati if r["stato"] == "già_calcolato"]),
        "num_senza_retribuzione": len([r for r in risultati if r["stato"] == "nessuna_retribuzione"])
    }


# ============================================
# GESTIONE ACCONTI (TFR, Ferie, 13ima, 14ima, Prestiti)
# ============================================

class AccontoInput(BaseModel):
    dipendente_id: str
    # Tipi: "stipendio" | "tfr" | "ferie" | "tredicesima" | "quattordicesima" | "prestito"
    tipo: str
    importo: float
    data: str  # YYYY-MM-DD
    note: Optional[str] = ""

    # === CAMPI ESTESI (gestione completa flusso acconti) ===
    # Distinzione tra acconto su lavoro futuro vs su pregresso (lavoro già svolto).
    # Default "su_futuro" perché è il caso standard (anticipo su busta del mese).
    natura_acconto: Optional[str] = "su_futuro"  # "su_futuro" | "su_pregresso"

    # Tipo di bonifico bancario. Ceraldi Group eroga acconti SOLO via banca,
    # mai in contanti. Distinguere standard vs istantaneo aiuta nella
    # riconciliazione (task 2) perché i bonifici standard appaiono in
    # estratto conto in 1-2 giorni lavorativi, gli istantanei lo stesso
    # giorno (anche festivi).
    tipo_bonifico: Optional[str] = "standard"  # "standard" | "istantaneo"

    # Mese/anno del cedolino su cui questo acconto verrà scalato.
    # Per default: stesso mese della data dell'acconto. L'utente può forzare
    # un mese diverso (es. "anticipo dato il 30/04 ma scalato su busta di maggio")
    scalato_su_anno_mese: Optional[str] = None  # formato "YYYY-MM"


class AccontoUpdateInput(BaseModel):
    """Modello per PUT /acconti/{id}: tutti i campi opzionali per update parziale."""
    importo: Optional[float] = None
    data: Optional[str] = None
    tipo: Optional[str] = None
    note: Optional[str] = None
    natura_acconto: Optional[str] = None
    tipo_bonifico: Optional[str] = None
    scalato_su_anno_mese: Optional[str] = None
    stato: Optional[str] = None


# Costanti per validazione (esposte a livello modulo per riuso in altri router)
TIPI_ACCONTO_VALIDI = {
    "stipendio", "tfr", "ferie", "tredicesima", "quattordicesima", "prestito",
}
NATURE_VALIDE = {"su_futuro", "su_pregresso"}
TIPI_BONIFICO_VALIDI = {"standard", "istantaneo"}
STATI_VALIDI = {
    "registrato",            # appena inserito
    "riconciliato_banca",    # collegato a movimento estratto conto
    "scalato_su_cedolino",   # confermato sul cedolino paga
    "annullato",             # rimosso dal flusso (non eliminato per audit)
}


@router.get("/acconti/{dipendente_id}")
@handle_errors
async def get_acconti_dipendente(dipendente_id: str) -> Dict[str, Any]:
    """
    Restituisce tutti gli acconti di un dipendente raggruppati per tipo.
    Include: TFR, Ferie, 13ima, 14ima, Prestiti.
    """
    db = Database.get_db()
    
    # Verifica dipendente
    dipendente = await db["dipendenti"].find_one(
        {"id": dipendente_id},
        {"_id": 0, "id": 1, "nome_completo": 1, "tfr_accantonato": 1}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Recupera tutti gli acconti
    acconti = await db["acconti_dipendenti"].find(
        {"dipendente_id": dipendente_id},
        {"_id": 0}
    ).sort("data", -1).to_list(500)
    
    # Raggruppa per tipo
    acconti_per_tipo = {
        "tfr": [],
        "ferie": [],
        "tredicesima": [],
        "quattordicesima": [],
        "prestito": []
    }
    
    totali = {
        "tfr": 0,
        "ferie": 0,
        "tredicesima": 0,
        "quattordicesima": 0,
        "prestito": 0
    }
    
    for acc in acconti:
        tipo = acc.get("tipo", "altro")
        if tipo in acconti_per_tipo:
            acconti_per_tipo[tipo].append(acc)
            totali[tipo] += acc.get("importo", 0)
    
    # Calcola saldi
    tfr_totale = float(dipendente.get("tfr_accantonato", 0))
    
    return {
        "dipendente_id": dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "tfr_accantonato": round(tfr_totale, 2),
        "tfr_acconti": round(totali["tfr"], 2),
        "tfr_saldo": round(tfr_totale - totali["tfr"], 2),
        "ferie_acconti": round(totali["ferie"], 2),
        "tredicesima_acconti": round(totali["tredicesima"], 2),
        "quattordicesima_acconti": round(totali["quattordicesima"], 2),
        "prestiti_totale": round(totali["prestito"], 2),
        "acconti": acconti_per_tipo,
        "totale_acconti": round(sum(totali.values()), 2)
    }


@router.post("/acconti")
@handle_errors
async def registra_acconto(input_data: AccontoInput) -> Dict[str, Any]:
    """
    Registra un acconto per un dipendente.

    Tipi supportati: stipendio, tfr, ferie, tredicesima, quattordicesima, prestito.

    Nuovi campi:
    - natura_acconto: "su_futuro" (anticipo su busta prossima) | "su_pregresso"
      (ripianamento su lavoro già svolto). Default "su_futuro".
    - tipo_bonifico: "standard" | "istantaneo". Default "standard".
      Tutti gli acconti Ceraldi sono via banca: distinguere standard da
      istantaneo aiuta nella riconciliazione con l'estratto conto (i
      bonifici istantanei arrivano anche in giornata festiva).
    - scalato_su_anno_mese: mese cedolino su cui andrà scalato (es. "2026-04").
      Se non fornito, derivato dalla data dell'acconto.

    Stato lifecycle:
        registrato → riconciliato_banca → scalato_su_cedolino
                  ↘ annullato

    L'acconto viene creato in stato "registrato". Le transizioni di stato
    avvengono via endpoint dedicati (riconcilia, scala-su-cedolino) che
    saranno aggiunti nei task successivi.
    """
    db = Database.get_db()

    # Verifica dipendente
    dipendente = await db["dipendenti"].find_one(
        {"id": input_data.dipendente_id},
        {"_id": 0}
    )

    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")

    # Validazioni di dominio
    if input_data.tipo not in TIPI_ACCONTO_VALIDI:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo non valido. Usa: {', '.join(sorted(TIPI_ACCONTO_VALIDI))}",
        )
    if input_data.importo <= 0:
        raise HTTPException(status_code=400, detail="L'importo deve essere positivo")

    natura = input_data.natura_acconto or "su_futuro"
    if natura not in NATURE_VALIDE:
        raise HTTPException(
            status_code=400,
            detail=f"Natura acconto non valida. Usa: {', '.join(sorted(NATURE_VALIDE))}",
        )

    metodo = input_data.tipo_bonifico or "standard"
    if metodo not in TIPI_BONIFICO_VALIDI:
        raise HTTPException(
            status_code=400,
            detail=f"Tipo bonifico non valido. Usa: {', '.join(sorted(TIPI_BONIFICO_VALIDI))}",
        )

    # Deriva scalato_su_anno_mese da data se non fornito
    scalato_su = input_data.scalato_su_anno_mese
    if not scalato_su:
        try:
            # input_data.data è in formato YYYY-MM-DD
            scalato_su = input_data.data[:7]  # estrae YYYY-MM
        except Exception:
            scalato_su = None

    # Estrae anno/mese numerici dalla data per query rapide su DB
    anno_int = mese_int = None
    try:
        if input_data.data and len(input_data.data) >= 7:
            anno_int = int(input_data.data[:4])
            mese_int = int(input_data.data[5:7])
    except Exception:
        pass

    now_iso = datetime.now(timezone.utc).isoformat()

    # Crea record acconto con schema esteso
    acconto = {
        "id": str(uuid4()),
        "dipendente_id": input_data.dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "tipo": input_data.tipo,
        "importo": round(input_data.importo, 2),
        "data": input_data.data,
        "anno": anno_int,
        "mese": mese_int,
        "note": input_data.note or "",

        # Nuovi campi
        "natura_acconto": natura,
        "tipo_bonifico": metodo,
        "scalato_su_anno_mese": scalato_su,

        # Stato lifecycle
        "stato": "registrato",
        "movimento_bancario_id": None,
        "riconciliato_il": None,
        "cedolino_id": None,
        "importo_scalato_effettivo": None,

        # Audit
        "source": "manuale",
        "created_at": now_iso,
        "updated_at": now_iso,
    }

    await db["acconti_dipendenti"].insert_one(acconto.copy())

    # Se è un acconto TFR, aggiorna anche il TFR del dipendente (logica preesistente)
    if input_data.tipo == "tfr":
        tfr_attuale = float(dipendente.get("tfr_accantonato", 0))
        nuovo_tfr = max(0, tfr_attuale - input_data.importo)
        await db["dipendenti"].update_one(
            {"id": input_data.dipendente_id},
            {"$set": {"tfr_accantonato": round(nuovo_tfr, 2)}}
        )

        # Registra movimento contabile
        movimento = {
            "id": str(uuid4()),
            "data": input_data.data,
            "descrizione": f"Acconto TFR - {dipendente.get('nome_completo', '')}",
            "tipo": "acconto_tfr",
            "importo": round(input_data.importo, 2),
            "dipendente_id": input_data.dipendente_id,
            "note": input_data.note or "",
            "created_at": now_iso,
        }
        await db["movimenti_contabili"].insert_one(movimento.copy())

    return {
        "success": True,
        "acconto_id": acconto["id"],
        "messaggio": f"Acconto {input_data.tipo} ({natura}) registrato per {dipendente.get('nome_completo', '')}",
        "importo": round(input_data.importo, 2),
        "natura": natura,
        "tipo_bonifico": metodo,
        "scalato_su": scalato_su,
        "stato": "registrato",
    }


@router.put("/acconti/{acconto_id}")
@handle_errors
async def modifica_acconto(acconto_id: str, input_data: dict) -> Dict[str, Any]:
    """Modifica un acconto esistente.

    Accetta tutti i campi del modello esteso. Se viene cambiata la data,
    ricalcola anche `anno`, `mese` numerici e `scalato_su_anno_mese`
    (a meno che quest'ultimo sia stato fornito esplicitamente).
    """
    db = Database.get_db()

    # Trova acconto
    acconto = await db["acconti_dipendenti"].find_one({"id": acconto_id})
    if not acconto:
        raise HTTPException(status_code=404, detail="Acconto non trovato")

    # Prepara aggiornamento
    update_fields: Dict[str, Any] = {}

    if "importo" in input_data and input_data["importo"] is not None:
        vecchio_importo = acconto.get("importo", 0)
        nuovo_importo = float(input_data["importo"])
        if nuovo_importo <= 0:
            raise HTTPException(status_code=400, detail="L'importo deve essere positivo")
        update_fields["importo"] = round(nuovo_importo, 2)

        # Se è un acconto TFR, aggiorna il saldo del dipendente
        if acconto.get("tipo") == "tfr":
            dipendente = await db["dipendenti"].find_one({"id": acconto["dipendente_id"]})
            if dipendente:
                tfr_attuale = float(dipendente.get("tfr_accantonato", 0))
                # Ripristina il vecchio importo e sottrai il nuovo
                nuovo_tfr = tfr_attuale + vecchio_importo - nuovo_importo
                await db["dipendenti"].update_one(
                    {"id": acconto["dipendente_id"]},
                    {"$set": {"tfr_accantonato": round(nuovo_tfr, 2)}}
                )

    if "data" in input_data and input_data["data"]:
        nuova_data = input_data["data"]
        update_fields["data"] = nuova_data
        # Ricalcola anno/mese numerici dal nuovo valore
        try:
            if len(nuova_data) >= 7:
                update_fields["anno"] = int(nuova_data[:4])
                update_fields["mese"] = int(nuova_data[5:7])
        except Exception:
            pass
        # Se l'utente non ha forzato scalato_su_anno_mese, derivalo dalla nuova data
        if "scalato_su_anno_mese" not in input_data:
            update_fields["scalato_su_anno_mese"] = nuova_data[:7]

    if "note" in input_data:
        update_fields["note"] = input_data["note"] or ""

    if "tipo" in input_data and input_data["tipo"]:
        if input_data["tipo"] not in TIPI_ACCONTO_VALIDI:
            raise HTTPException(
                status_code=400,
                detail=f"Tipo non valido. Usa: {', '.join(sorted(TIPI_ACCONTO_VALIDI))}",
            )
        update_fields["tipo"] = input_data["tipo"]

    if "natura_acconto" in input_data and input_data["natura_acconto"]:
        if input_data["natura_acconto"] not in NATURE_VALIDE:
            raise HTTPException(
                status_code=400,
                detail=f"Natura non valida. Usa: {', '.join(sorted(NATURE_VALIDE))}",
            )
        update_fields["natura_acconto"] = input_data["natura_acconto"]

    if "tipo_bonifico" in input_data and input_data["tipo_bonifico"]:
        if input_data["tipo_bonifico"] not in TIPI_BONIFICO_VALIDI:
            raise HTTPException(
                status_code=400,
                detail=f"Tipo bonifico non valido. Usa: {', '.join(sorted(TIPI_BONIFICO_VALIDI))}",
            )
        update_fields["tipo_bonifico"] = input_data["tipo_bonifico"]

    if "scalato_su_anno_mese" in input_data:
        # Accetta None per "rimuovi binding"
        update_fields["scalato_su_anno_mese"] = input_data["scalato_su_anno_mese"]

    if "stato" in input_data and input_data["stato"]:
        if input_data["stato"] not in STATI_VALIDI:
            raise HTTPException(
                status_code=400,
                detail=f"Stato non valido. Usa: {', '.join(sorted(STATI_VALIDI))}",
            )
        update_fields["stato"] = input_data["stato"]

    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db["acconti_dipendenti"].update_one(
            {"id": acconto_id},
            {"$set": update_fields}
        )

    return {
        "success": True,
        "messaggio": f"Acconto {acconto.get('tipo', '')} modificato",
        "acconto_id": acconto_id,
        "campi_aggiornati": sorted(update_fields.keys()),
    }


@router.delete("/acconti/{acconto_id}")
@handle_errors
async def elimina_acconto(acconto_id: str) -> Dict[str, Any]:
    """Elimina un acconto."""
    db = Database.get_db()
    
    # Trova acconto
    acconto = await db["acconti_dipendenti"].find_one({"id": acconto_id})
    if not acconto:
        raise HTTPException(status_code=404, detail="Acconto non trovato")
    
    # Se era un acconto TFR, ripristina il valore
    if acconto.get("tipo") == "tfr":
        dipendente = await db["dipendenti"].find_one({"id": acconto["dipendente_id"]})
        if dipendente:
            tfr_attuale = float(dipendente.get("tfr_accantonato", 0))
            nuovo_tfr = tfr_attuale + acconto.get("importo", 0)
            await db["dipendenti"].update_one(
                {"id": acconto["dipendente_id"]},
                {"$set": {"tfr_accantonato": round(nuovo_tfr, 2)}}
            )
    
    # Elimina acconto
    await db["acconti_dipendenti"].delete_one({"id": acconto_id})
    
    return {
        "success": True,
        "messaggio": f"Acconto {acconto.get('tipo', '')} eliminato"
    }


# ============================================
# RICONCILIAZIONE ACCONTO ↔ MOVIMENTO ESTRATTO CONTO
# ============================================

@router.get("/acconti/{acconto_id}/candidati-banca")
@handle_errors
async def candidati_banca_per_acconto(acconto_id: str) -> Dict[str, Any]:
    """Cerca movimenti dell'estratto conto compatibili con questo acconto.

    Logica di matching:
    - Solo movimenti uscita (importo < 0 o tipo='uscita')
    - Importo uguale a quello dell'acconto (tolleranza ±0.01€)
    - Range data dipende dal tipo_bonifico:
        * 'istantaneo' → stesso giorno della registrazione (±1gg margine)
        * 'standard' → entro 5 giorni dopo la registrazione (skip weekend non
          implementato perché alcuni istituti accreditano comunque il sabato)
    - Esclude movimenti già riconciliati con un altro acconto
    - La descrizione contiene il cognome o il nome del dipendente

    Restituisce candidati ordinati per "score" decrescente (best match prima).
    Score: data esatta=+50, importo esatto=+30, nome in descrizione=+20.
    """
    db = Database.get_db()

    acconto = await db["acconti_dipendenti"].find_one({"id": acconto_id}, {"_id": 0})
    if not acconto:
        raise HTTPException(status_code=404, detail="Acconto non trovato")

    if acconto.get("stato") == "riconciliato_banca":
        raise HTTPException(
            status_code=400,
            detail=f"Acconto già riconciliato. Movimento collegato: {acconto.get('movimento_bancario_id')}",
        )

    importo_target = abs(float(acconto.get("importo", 0)))
    if importo_target <= 0:
        raise HTTPException(status_code=400, detail="Acconto con importo non valido")

    data_acconto_str = acconto.get("data", "")
    try:
        data_acconto = datetime.strptime(data_acconto_str, "%Y-%m-%d")
    except (ValueError, TypeError):
        raise HTTPException(
            status_code=400,
            detail=f"Data acconto non valida: {data_acconto_str}",
        )

    tipo_bonifico = acconto.get("tipo_bonifico") or "standard"

    # Range temporale per la ricerca
    from datetime import timedelta
    if tipo_bonifico == "istantaneo":
        # Bonifico istantaneo: stesso giorno (±1gg margine per fusi orari/contabilità)
        data_min = data_acconto - timedelta(days=1)
        data_max = data_acconto + timedelta(days=1, hours=23, minutes=59)
    else:
        # Bonifico standard: range +0/+5gg dalla data registrazione
        # (alcuni istituti accreditano in giornata, altri D+1, raramente D+2/D+3)
        data_min = data_acconto - timedelta(days=1)
        data_max = data_acconto + timedelta(days=5, hours=23, minutes=59)

    # Recupera nome dipendente per matching descrizione
    dipendente_nome = (acconto.get("dipendente_nome") or "").strip()
    nome_parti = [p for p in dipendente_nome.split() if len(p) >= 3]

    # Query movimenti candidati
    # La collezione canonica è estratto_conto_movimenti.
    # data_contabile_obj è datetime per range query efficienti.
    query: Dict[str, Any] = {
        # Movimento di uscita: tipo='uscita' OR importo<0
        "$or": [
            {"tipo": "uscita"},
            {"importo": {"$lt": 0}},
        ],
        # Importo entro tolleranza ±0.01
        # NB: alcuni movimenti hanno importo negativo (uscite), altri positivo
        # con tipo='uscita' — facciamo confronto su valore assoluto via $expr
        "$and": [
            {"$expr": {
                "$lte": [
                    {"$abs": {"$subtract": [{"$abs": "$importo"}, importo_target]}},
                    0.01,
                ]
            }}
        ],
        # Range data
        "data_contabile_obj": {"$gte": data_min, "$lte": data_max},
        # Esclude movimenti già usati per altri acconti
        "$nor": [
            {"acconto_id": {"$exists": True, "$ne": None, "$ne": ""}},
        ],
    }

    movimenti = await db["estratto_conto_movimenti"].find(
        query, {"_id": 0}
    ).sort("data_contabile_obj", 1).limit(50).to_list(50)

    # Calcolo score per ranking
    candidati = []
    for m in movimenti:
        score = 0
        match_reasons = []

        # Data: esatta = +50, ±1gg = +30, oltre = +10
        try:
            data_mov = m.get("data_contabile_obj")
            if isinstance(data_mov, str):
                data_mov = datetime.fromisoformat(data_mov.replace("Z", ""))
            delta_giorni = abs((data_mov - data_acconto).days) if data_mov else 99
            if delta_giorni == 0:
                score += 50
                match_reasons.append("data esatta")
            elif delta_giorni <= 1:
                score += 30
                match_reasons.append(f"data ±{delta_giorni}gg")
            else:
                score += 10
                match_reasons.append(f"data +{delta_giorni}gg")
        except Exception:
            delta_giorni = None

        # Importo: già filtrato a tolleranza 0.01 → tutti hanno importo esatto
        score += 30
        match_reasons.append("importo esatto")

        # Nome dipendente in descrizione
        descrizione = (m.get("descrizione") or "").upper()
        if nome_parti:
            for parte in nome_parti:
                if parte.upper() in descrizione:
                    score += 20
                    match_reasons.append(f"nome '{parte}' in descrizione")
                    break

        candidati.append({
            "movimento_id": m.get("id"),
            "data": m.get("data") or (m.get("data_contabile_obj").strftime("%Y-%m-%d") if m.get("data_contabile_obj") else None),
            "descrizione": m.get("descrizione", ""),
            "importo": m.get("importo"),
            "categoria": m.get("categoria"),
            "fornitore": m.get("fornitore"),
            "score": score,
            "match_reasons": match_reasons,
            "delta_giorni": delta_giorni,
        })

    # Ordina per score desc
    candidati.sort(key=lambda c: c["score"], reverse=True)

    return {
        "success": True,
        "acconto": {
            "id": acconto.get("id"),
            "dipendente_nome": dipendente_nome,
            "importo": acconto.get("importo"),
            "data": data_acconto_str,
            "tipo_bonifico": tipo_bonifico,
        },
        "ricerca": {
            "data_min": data_min.strftime("%Y-%m-%d"),
            "data_max": data_max.strftime("%Y-%m-%d"),
            "tolleranza_importo": 0.01,
        },
        "totale_candidati": len(candidati),
        "candidati": candidati,
    }


class RiconciliaBancaInput(BaseModel):
    movimento_id: str


@router.post("/acconti/{acconto_id}/riconcilia-banca")
@handle_errors
async def riconcilia_acconto_banca(
    acconto_id: str, payload: RiconciliaBancaInput
) -> Dict[str, Any]:
    """Collega manualmente un acconto a un movimento dell'estratto conto.

    Effetti:
    - Acconto: stato='riconciliato_banca', movimento_bancario_id=<id>,
      riconciliato_il=<now>
    - Movimento: acconto_id=<id> (per evitare doppia riconciliazione)
    """
    db = Database.get_db()

    acconto = await db["acconti_dipendenti"].find_one({"id": acconto_id})
    if not acconto:
        raise HTTPException(status_code=404, detail="Acconto non trovato")

    if acconto.get("stato") == "riconciliato_banca":
        raise HTTPException(
            status_code=400,
            detail="Acconto già riconciliato. Annulla prima la riconciliazione esistente.",
        )

    movimento = await db["estratto_conto_movimenti"].find_one(
        {"id": payload.movimento_id}, {"_id": 0}
    )
    if not movimento:
        raise HTTPException(status_code=404, detail="Movimento estratto conto non trovato")

    if movimento.get("acconto_id"):
        raise HTTPException(
            status_code=409,
            detail=f"Movimento già collegato all'acconto {movimento.get('acconto_id')}",
        )

    now_iso = datetime.now(timezone.utc).isoformat()

    # Aggiorna acconto
    await db["acconti_dipendenti"].update_one(
        {"id": acconto_id},
        {"$set": {
            "movimento_bancario_id": payload.movimento_id,
            "riconciliato_il": now_iso,
            "stato": "riconciliato_banca",
            "updated_at": now_iso,
        }},
    )

    # Aggiorna movimento (link inverso per anti-doppia-riconciliazione)
    await db["estratto_conto_movimenti"].update_one(
        {"id": payload.movimento_id},
        {"$set": {
            "acconto_id": acconto_id,
            "categoria_acconto": acconto.get("tipo", "stipendio"),
            "dipendente_nome": acconto.get("dipendente_nome", ""),
            "updated_at": now_iso,
        }},
    )

    return {
        "success": True,
        "messaggio": f"Acconto riconciliato con movimento del {movimento.get('data', '?')}",
        "acconto_id": acconto_id,
        "movimento_id": payload.movimento_id,
        "stato": "riconciliato_banca",
    }


@router.post("/acconti/{acconto_id}/annulla-riconciliazione-banca")
@handle_errors
async def annulla_riconciliazione_banca(acconto_id: str) -> Dict[str, Any]:
    """Annulla la riconciliazione bancaria di un acconto.

    Riporta lo stato a 'registrato' e rimuove il link sul movimento.
    Utile in caso di errore di abbinamento.
    """
    db = Database.get_db()

    acconto = await db["acconti_dipendenti"].find_one({"id": acconto_id})
    if not acconto:
        raise HTTPException(status_code=404, detail="Acconto non trovato")

    if acconto.get("stato") != "riconciliato_banca":
        raise HTTPException(
            status_code=400,
            detail=f"Acconto non in stato riconciliato_banca (stato attuale: {acconto.get('stato')})",
        )

    movimento_id = acconto.get("movimento_bancario_id")
    now_iso = datetime.now(timezone.utc).isoformat()

    # Rimuovi link da acconto
    await db["acconti_dipendenti"].update_one(
        {"id": acconto_id},
        {
            "$set": {"stato": "registrato", "updated_at": now_iso},
            "$unset": {"movimento_bancario_id": "", "riconciliato_il": ""},
        },
    )

    # Rimuovi link da movimento
    if movimento_id:
        await db["estratto_conto_movimenti"].update_one(
            {"id": movimento_id},
            {
                "$unset": {"acconto_id": "", "categoria_acconto": ""},
                "$set": {"updated_at": now_iso},
            },
        )

    return {
        "success": True,
        "messaggio": "Riconciliazione bancaria annullata",
        "acconto_id": acconto_id,
        "movimento_id": movimento_id,
        "stato": "registrato",
    }


# ============================================
# SCALATURA ACCONTI SU CEDOLINO PAGA (Task 3)
# ============================================
# Quando arriva il cedolino mensile di un dipendente, il sistema confronta
# il valore "acconto_mese_precedente" (estratto dal PDF dal parser AI) con
# il totale degli acconti `scalato_su_anno_mese == cedolino_periodo`
# registrati per quel dipendente. Se quadra, marca tutti come scalati.
# Se non quadra, restituisce la discrepanza per scelta manuale.

def _estrai_acconto_da_cedolino(cedolino: Dict[str, Any]) -> Optional[float]:
    """Estrae il valore 'acconto già erogato' dal cedolino salvato.

    Cerca in vari campi possibili (dipende dal flusso di import):
    - cedolino.enhanced_parsing.importi_finali.acconto_mese_precedente (parser AI)
    - cedolino.acconto_mese_precedente (campo flat se promosso)
    - cedolino.importi_finali.acconto_mese_precedente (legacy)

    Returns None se non disponibile (cedolino non parsato col parser AI).
    """
    candidates = []

    enhanced = cedolino.get("enhanced_parsing") or {}
    if isinstance(enhanced, dict):
        importi = enhanced.get("importi_finali") or {}
        if isinstance(importi, dict):
            v = importi.get("acconto_mese_precedente")
            if v is not None:
                candidates.append(v)

    importi_flat = cedolino.get("importi_finali") or {}
    if isinstance(importi_flat, dict):
        v = importi_flat.get("acconto_mese_precedente")
        if v is not None:
            candidates.append(v)

    v = cedolino.get("acconto_mese_precedente")
    if v is not None:
        candidates.append(v)

    for c in candidates:
        try:
            f = float(c)
            if f > 0:
                return round(f, 2)
        except (ValueError, TypeError):
            continue

    return None


async def _trova_acconti_da_scalare(
    db, dipendente_id: str, codice_fiscale: Optional[str], anno: int, mese: int
) -> List[Dict[str, Any]]:
    """Trova acconti candidati alla scalatura per un cedolino.

    Criteri:
    - scalato_su_anno_mese == "{anno}-{mese:02d}"
    - stato in ('registrato', 'riconciliato_banca') — non già scalati né annullati
    - matching dipendente: prima per id, fallback per CF (per record legacy
      senza dipendente_id ma con codice_fiscale)
    """
    periodo = f"{anno}-{str(mese).zfill(2)}"
    stati_eligibili = ["registrato", "riconciliato_banca"]

    # Match primario per dipendente_id
    query = {
        "dipendente_id": dipendente_id,
        "scalato_su_anno_mese": periodo,
        "stato": {"$in": stati_eligibili},
    }
    items = await db["acconti_dipendenti"].find(query, {"_id": 0}).to_list(500)

    # Fallback su CF se non ho trovato nulla via id (record legacy)
    if not items and codice_fiscale:
        query_cf = {
            "codice_fiscale": codice_fiscale.upper().strip(),
            "scalato_su_anno_mese": periodo,
            "stato": {"$in": stati_eligibili},
        }
        items = await db["acconti_dipendenti"].find(query_cf, {"_id": 0}).to_list(500)

    return items


@router.get("/cedolini/{cedolino_id}/preview-scalatura-acconti")
@handle_errors
async def preview_scalatura_acconti(cedolino_id: str) -> Dict[str, Any]:
    """Anteprima della scalatura acconti per un cedolino (NON scrive sul DB).

    Estrae da cedolino.enhanced_parsing.importi_finali.acconto_mese_precedente
    il totale dichiarato dal cedolino. Confronta con la somma degli acconti
    registrati per quel dipendente con scalato_su_anno_mese == periodo.

    Risposta:
        {
          "cedolino": {id, dipendente_id, anno, mese, ...},
          "acconto_mese_precedente": float | null,
          "acconti_registrati": [...],
          "totale_acconti_registrati": float,
          "delta": float,                # acconto_cedolino - totale_registrati
          "stato_match": "quadra" | "discrepanza" | "nessun_dato_cedolino" |
                         "nessun_acconto",
          "messaggio": str,
        }
    """
    db = Database.get_db()

    cedolino = await db["cedolini"].find_one({"id": cedolino_id}, {"_id": 0})
    if not cedolino:
        raise HTTPException(status_code=404, detail="Cedolino non trovato")

    dipendente_id = cedolino.get("dipendente_id")
    codice_fiscale = cedolino.get("codice_fiscale") or cedolino.get("cf")
    anno = cedolino.get("anno")
    mese = cedolino.get("mese")

    if not anno or not mese:
        raise HTTPException(
            status_code=400,
            detail="Cedolino senza anno/mese - impossibile fare matching",
        )

    if not dipendente_id and not codice_fiscale:
        raise HTTPException(
            status_code=400,
            detail="Cedolino senza dipendente_id né codice_fiscale - impossibile fare matching",
        )

    acconto_cedolino = _estrai_acconto_da_cedolino(cedolino)
    acconti = await _trova_acconti_da_scalare(
        db, dipendente_id, codice_fiscale, anno, mese
    )
    totale_registrati = round(sum(float(a.get("importo", 0) or 0) for a in acconti), 2)

    # Determina stato_match
    stato_match: str
    messaggio: str
    delta: Optional[float] = None

    if acconto_cedolino is None and not acconti:
        stato_match = "nessun_dato"
        messaggio = (
            "Il cedolino non riporta acconti del mese precedente e nel sistema "
            "non risultano acconti registrati per questo periodo."
        )
    elif acconto_cedolino is None:
        stato_match = "nessun_dato_cedolino"
        messaggio = (
            f"Il cedolino non riporta il valore 'acconto mese precedente' "
            f"(probabilmente non parsato con AI), ma nel sistema risultano "
            f"{len(acconti)} acconti registrati per €{totale_registrati:.2f}. "
            f"Riprocessa il cedolino con parser AI o scala manualmente."
        )
    elif not acconti:
        stato_match = "nessun_acconto"
        messaggio = (
            f"Il cedolino dichiara €{acconto_cedolino:.2f} di acconti già erogati "
            f"ma nel sistema non risultano acconti registrati per il mese "
            f"{anno}-{str(mese).zfill(2)}. Verifica di aver registrato gli acconti."
        )
    else:
        delta = round(acconto_cedolino - totale_registrati, 2)
        if abs(delta) < 0.01:
            stato_match = "quadra"
            messaggio = (
                f"Match perfetto: cedolino dichiara €{acconto_cedolino:.2f}, "
                f"sistema ha {len(acconti)} acconti registrati per "
                f"€{totale_registrati:.2f}."
            )
        else:
            stato_match = "discrepanza"
            messaggio = (
                f"DISCREPANZA: cedolino dichiara €{acconto_cedolino:.2f}, "
                f"sistema ha {len(acconti)} acconti registrati per "
                f"€{totale_registrati:.2f} (delta: €{delta:+.2f}). "
                f"Verifica manualmente prima di applicare la scalatura."
            )

    return {
        "cedolino": {
            "id": cedolino.get("id"),
            "dipendente_id": dipendente_id,
            "dipendente_nome": cedolino.get("dipendente_nome") or cedolino.get("nome_dipendente"),
            "codice_fiscale": codice_fiscale,
            "anno": anno,
            "mese": mese,
            "periodo": f"{anno}-{str(mese).zfill(2)}",
        },
        "acconto_mese_precedente": acconto_cedolino,
        "acconti_registrati": [
            {
                "id": a.get("id"),
                "data": a.get("data"),
                "importo": a.get("importo"),
                "tipo": a.get("tipo"),
                "natura_acconto": a.get("natura_acconto") or "su_futuro",
                "tipo_bonifico": a.get("tipo_bonifico") or "standard",
                "stato": a.get("stato") or "registrato",
                "note": a.get("note") or "",
            }
            for a in acconti
        ],
        "totale_acconti_registrati": totale_registrati,
        "delta": delta,
        "stato_match": stato_match,
        "messaggio": messaggio,
    }


class ScalaturaInput(BaseModel):
    forza_anche_se_discrepanza: Optional[bool] = False


@router.post("/cedolini/{cedolino_id}/scala-acconti")
@handle_errors
async def scala_acconti_su_cedolino(
    cedolino_id: str, payload: ScalaturaInput = Body(default=ScalaturaInput())
) -> Dict[str, Any]:
    """Scala gli acconti sul cedolino: marca tutti gli acconti del periodo
    come 'scalato_su_cedolino', linkando il cedolino_id e l'importo scalato.

    Comportamento:
    - Se quadra (delta < 0.01): scala tutti, ognuno per il suo importo intero
    - Se discrepanza: errore 400 con dettagli, A MENO CHE
      forza_anche_se_discrepanza=True (l'utente ha verificato e accetta)
    - Se nessun acconto registrato: 400
    - Se nessun dato cedolino: 400 (chiede di riprocessare con AI)

    Effetti per ogni acconto:
    - stato='scalato_su_cedolino'
    - cedolino_id=<id>
    - importo_scalato_effettivo=<importo dell'acconto, integralmente>
    """
    db = Database.get_db()

    # Riusa la logica del preview per coerenza
    preview = await preview_scalatura_acconti(cedolino_id)
    stato_match = preview["stato_match"]

    if stato_match in ("nessun_acconto", "nessun_dato", "nessun_dato_cedolino"):
        raise HTTPException(
            status_code=400,
            detail=preview["messaggio"],
        )

    if stato_match == "discrepanza" and not payload.forza_anche_se_discrepanza:
        raise HTTPException(
            status_code=409,
            detail={
                "stato_match": "discrepanza",
                "messaggio": preview["messaggio"],
                "delta": preview["delta"],
                "totale_registrati": preview["totale_acconti_registrati"],
                "totale_cedolino": preview["acconto_mese_precedente"],
                "hint": "Imposta forza_anche_se_discrepanza=true per applicare comunque",
            },
        )

    # Procedi con la scalatura
    now_iso = datetime.now(timezone.utc).isoformat()
    acconti_ids = [a["id"] for a in preview["acconti_registrati"]]
    aggiornati = 0
    for acconto_data in preview["acconti_registrati"]:
        await db["acconti_dipendenti"].update_one(
            {"id": acconto_data["id"]},
            {"$set": {
                "stato": "scalato_su_cedolino",
                "cedolino_id": cedolino_id,
                "importo_scalato_effettivo": acconto_data["importo"],
                "scalato_il": now_iso,
                "updated_at": now_iso,
            }},
        )
        aggiornati += 1

    return {
        "success": True,
        "messaggio": (
            f"Scalati {aggiornati} acconti su cedolino "
            f"{preview['cedolino']['periodo']} di "
            f"{preview['cedolino']['dipendente_nome'] or preview['cedolino']['codice_fiscale']}"
            + (" (FORZATO con discrepanza)" if stato_match == "discrepanza" else "")
        ),
        "cedolino_id": cedolino_id,
        "acconti_scalati": acconti_ids,
        "totale_scalato": preview["totale_acconti_registrati"],
        "totale_cedolino": preview["acconto_mese_precedente"],
        "delta": preview["delta"],
        "stato_match": stato_match,
        "forzato": stato_match == "discrepanza",
    }


@router.post("/cedolini/{cedolino_id}/annulla-scalatura-acconti")
@handle_errors
async def annulla_scalatura_acconti(cedolino_id: str) -> Dict[str, Any]:
    """Annulla la scalatura: riporta gli acconti del cedolino allo stato
    precedente (riconciliato_banca se hanno movimento_bancario_id, altrimenti
    registrato).
    """
    db = Database.get_db()

    cedolino = await db["cedolini"].find_one({"id": cedolino_id}, {"_id": 0})
    if not cedolino:
        raise HTTPException(status_code=404, detail="Cedolino non trovato")

    # Trova acconti scalati su questo cedolino
    acconti_scalati = await db["acconti_dipendenti"].find(
        {"cedolino_id": cedolino_id, "stato": "scalato_su_cedolino"},
        {"_id": 0}
    ).to_list(500)

    if not acconti_scalati:
        raise HTTPException(
            status_code=404,
            detail="Nessun acconto risulta scalato su questo cedolino",
        )

    now_iso = datetime.now(timezone.utc).isoformat()
    for a in acconti_scalati:
        # Ripristina stato precedente in base alla presenza di movimento bancario
        nuovo_stato = "riconciliato_banca" if a.get("movimento_bancario_id") else "registrato"
        await db["acconti_dipendenti"].update_one(
            {"id": a["id"]},
            {
                "$set": {"stato": nuovo_stato, "updated_at": now_iso},
                "$unset": {
                    "cedolino_id": "",
                    "importo_scalato_effettivo": "",
                    "scalato_il": "",
                },
            },
        )

    return {
        "success": True,
        "messaggio": f"Annullata scalatura di {len(acconti_scalati)} acconti dal cedolino",
        "cedolino_id": cedolino_id,
        "acconti_ripristinati": [a["id"] for a in acconti_scalati],
    }


@router.get("/parse-payslips")
@handle_errors
async def parse_payslips_for_tfr() -> Dict[str, Any]:
    """
    Analizza i PDF delle buste paga per estrarre i dati TFR.
    Legge dalla cartella /app/uploads/paghe.
    """
    try:
        from backend.app.services.payslip_pdf_parser import parse_all_payslips
        
        if not os.path.exists(PAYSLIPS_FOLDER):
            return {
                "success": False,
                "error": f"Cartella {PAYSLIPS_FOLDER} non trovata",
                "data": []
            }
        
        # Conta PDF disponibili
        pdf_files = list(Path(PAYSLIPS_FOLDER).glob("Libro*.pdf"))
        
        if not pdf_files:
            return {
                "success": False,
                "error": "Nessun file 'Libro Unico' trovato nella cartella",
                "data": []
            }
        
        # Parse tutti i PDF
        data = parse_all_payslips(PAYSLIPS_FOLDER)
        
        return {
            "success": True,
            "num_pdf_analizzati": len(pdf_files),
            "num_dipendenti_trovati": len(data),
            "dipendenti": data
        }
        
    except Exception as e:
        logger.error(f"Errore parsing buste paga: {e}")
        return {
            "success": False,
            "error": str(e),
            "data": []
        }


@router.get("/storico-tfr/{dipendente_id}")
@handle_errors
async def get_storico_tfr(dipendente_id: str) -> Dict[str, Any]:
    """
    Restituisce lo storico completo del TFR di un dipendente.
    Include: accantonamenti, acconti, variazioni.
    """
    db = Database.get_db()
    
    # Verifica dipendente
    dipendente = await db["dipendenti"].find_one(
        {"id": dipendente_id},
        {"_id": 0}
    )
    
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")
    
    # Accantonamenti
    accantonamenti = await db["tfr_accantonamenti"].find(
        {"dipendente_id": dipendente_id},
        {"_id": 0}
    ).sort("anno", -1).to_list(100)
    
    # Liquidazioni
    liquidazioni = await db["tfr_liquidazioni"].find(
        {"dipendente_id": dipendente_id},
        {"_id": 0}
    ).sort("data", -1).to_list(100)
    
    # Acconti TFR
    acconti_tfr = await db["acconti_dipendenti"].find(
        {"dipendente_id": dipendente_id, "tipo": "tfr"},
        {"_id": 0}
    ).sort("data", -1).to_list(100)
    
    # Calcola totali
    totale_accantonato = sum(a.get("totale_accantonamento", 0) for a in accantonamenti)
    totale_liquidato = sum(l.get("importo_lordo", 0) for l in liquidazioni)
    totale_acconti = sum(acc.get("importo", 0) for acc in acconti_tfr)
    
    tfr_attuale = float(dipendente.get("tfr_accantonato", 0))
    
    return {
        "dipendente_id": dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "tfr_attuale": round(tfr_attuale, 2),
        "totale_accantonato": round(totale_accantonato, 2),
        "totale_liquidato": round(totale_liquidato, 2),
        "totale_acconti": round(totale_acconti, 2),
        "accantonamenti": accantonamenti,
        "liquidazioni": liquidazioni,
        "acconti": acconti_tfr
    }


# ============================================
# SIMULATORE TFR STORICO (periodo per periodo)
# ============================================
# Ricostruisce il TFR maturato PRIMA che l'app calcolasse tutto in automatico
# dai cedolini, e da lì in poi lo tiene aggiornato da solo. Modello:
# - Ogni dipendente ha una catena di periodi con una paga settimanale ciascuno.
# - L'ULTIMO periodo può essere APERTO (data_fine assente = "tuttora in corso"):
#   il suo maturato viene ricalcolato al volo fino a oggi ad ogni lettura, senza
#   bisogno di aggiungere un periodo ogni anno.
# - L'unica azione manuale prevista è registrare un aumento: si comunica la
#   nuova paga settimanale (e la data da cui vale, di default oggi); il sistema
#   chiude da solo il periodo aperto precedente (data_fine = giorno prima) e
#   apre il nuovo, sempre aperto. L'import da Excel (una tantum, per tutti i
#   dipendenti insieme) fa lo stesso: l'ultima riga di ciascun foglio diventa il
#   periodo aperto, ignorando la data di fine eventualmente scritta nel file.
# Formula (stessa del foglio storico del titolare, verificata al centesimo sulla
# riga 2023: 300€/sett → lordo 1.314,44, tassazione 27% 354,90, netto 959,54):
#   mesi  = giorni del periodo / 30            (il giorno finale non si conta,
#                                               come DATEDIF di Excel: un anno
#                                               intero = 364/30 = 12,13 mesi)
#   lordo = importo_settimanale × 52/12/12 × mesi   (mensilità/12 per ogni mese)
#   tassazione = lordo × aliquota del periodo   (23% di default, modificabile
#                                               periodo per periodo — il foglio
#                                               usava 27% su alcuni anni)
#   netto = lordo − tassazione                  (il "netto da ricevere")
# 13ª/14ª (liquidazione): stessa struttura con divisore 30,416 come nel foglio.
# È un sandbox separato da 'dipendenti.tfr_accantonato' (quello alimentato in
# automatico dai cedolini): la simulazione non lo tocca, per non mescolare un
# dato storico ancora da verificare con quello già vivo in produzione.

class PeriodoSimulazioneInput(BaseModel):
    importo_settimanale: float
    # Se omessa: giorno dopo la fine dell'ultimo periodo chiuso, oggi se il
    # precedente è aperto (è un aumento), o data assunzione per il primo periodo.
    data_inizio: Optional[str] = None
    # Se omessa: periodo APERTO (tuttora in corso). Se valorizzata: periodo storico chiuso.
    data_fine: Optional[str] = None
    # Percentuale di tassazione del periodo (default 23; il foglio storico usava 27 su alcuni anni)
    aliquota_tassazione: Optional[float] = None


class RateSimulazioneInput(BaseModel):
    numero_rate: int
    data_prima_rata: Optional[str] = None  # YYYY-MM-DD, opzionale


def _parse_data_tfr(s: str) -> datetime:
    return datetime.strptime(s[:10], "%Y-%m-%d")


def _oggi_tfr() -> datetime:
    return datetime.now(timezone.utc).replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=None)


def _mesi_excel(dal: datetime, al: datetime, divisore: float = 30.0) -> float:
    """Mesi del periodo come nel foglio storico: giorni (senza contare il giorno
    finale, come DATEDIF di Excel) diviso 30 — un anno intero 1/1→31/12 fa
    364/30 = 12,13 mesi. Per 13ª/14ª il foglio usava il divisore 30,416."""
    if al <= dal:
        return 0.0
    return (al - dal).days / divisore


def _calcola_periodo_tfr(data_inizio: datetime, data_fine: datetime, importo_settimanale: float,
                         aliquota_tassazione: Optional[float] = None) -> Dict[str, Any]:
    """Quota TFR del periodo, con la STESSA formula del foglio storico del titolare
    (verificata al centesimo sulla riga 2023: 300€/sett, 27% → 1.314,44 / 354,90 /
    959,54): lordo = importo_settimanale × 52/12/12 × mesi (mesi = giorni/30, giorno
    finale escluso), tassazione = lordo × aliquota del periodo (default 23%),
    netto = lordo − tassazione."""
    aliquota = ALIQUOTA_TFR if aliquota_tassazione is None else aliquota_tassazione
    mesi = _mesi_excel(data_inizio, data_fine)
    lordo = importo_settimanale * 52 / 12 / 12 * mesi
    tassazione = lordo * aliquota / 100
    netto = lordo - tassazione
    return {
        "giorni": max(0, (data_fine - data_inizio).days),
        "mesi": round(mesi, 2),
        "lordo": round(lordo, 2),
        "tassazione": round(tassazione, 2),
        "netto": round(netto, 2),
    }


def _fino_a_calcolo(dipendente: Dict[str, Any]) -> datetime:
    """Limite temporale per ogni maturato: se il dipendente è cessato, la data di
    cessazione (il rapporto non matura più nulla dopo); altrimenti oggi."""
    if dipendente.get("stato") == "cessato":
        data_cess = dipendente.get("data_dimissione") or dipendente.get("data_cessazione")
        if data_cess:
            try:
                return _parse_data_tfr(data_cess)
            except ValueError:
                pass
    return _oggi_tfr()


def _periodo_con_calcolo_live(p: Dict[str, Any], fino_a: Optional[datetime] = None) -> Dict[str, Any]:
    """Se il periodo è aperto (senza data_fine), ricalcola il maturato al volo fino a
    'fino_a' (oggi, o la cessazione se il dipendente ha lasciato), senza scrivere nulla
    sul DB: il valore resta sempre aggiornato finché il rapporto è in corso."""
    if p.get("data_fine"):
        return {**p, "aperto": False}
    limite = fino_a or _oggi_tfr()
    calc = _calcola_periodo_tfr(_parse_data_tfr(p["data_inizio"]), limite,
                                p["importo_settimanale"], p.get("aliquota_tassazione"))
    return {**p, **calc, "data_fine": None, "aperto": True}


def _periodo_competenza_13(fino_a: datetime):
    """Tredicesima: competenza anno solare (gennaio-dicembre), corrisposta a dicembre."""
    anno = fino_a.year
    return datetime(anno, 1, 1), datetime(anno, 12, 31)


def _periodo_competenza_14(fino_a: datetime):
    """Quattordicesima: competenza 1° luglio - 30 giugno, corrisposta a luglio."""
    if fino_a.month >= 7:
        return datetime(fino_a.year, 7, 1), datetime(fino_a.year + 1, 6, 30)
    return datetime(fino_a.year - 1, 7, 1), datetime(fino_a.year, 6, 30)


def _quota_mensilita_aggiuntiva(periodi_grezzi: List[Dict[str, Any]], data_assunzione: Optional[datetime],
                                inizio_competenza: datetime, fine_competenza: datetime,
                                fino_a: datetime) -> Dict[str, Any]:
    """Rateo di tredicesima/quattordicesima maturato nel ciclo di competenza indicato,
    con la stessa struttura del foglio storico (righe "13°"/"14°"): per ogni pezzo di
    ciclo coperto da un periodo di paga, lordo = importo_settimanale × 52/12/12 × mesi
    (mesi = giorni/30,416 come nel foglio), tassazione con l'aliquota di quel periodo,
    netto = lordo − tassazione."""
    inizio_eff = max(inizio_competenza, data_assunzione) if data_assunzione else inizio_competenza
    fine_eff = min(fine_competenza, fino_a)
    if fine_eff < inizio_eff:
        return {"lordo": 0.0, "tassazione": 0.0, "netto": 0.0, "mesi": 0.0,
                "dal": inizio_eff.strftime("%Y-%m-%d"), "al": fine_eff.strftime("%Y-%m-%d")}
    tot_lordo, tot_tass, mesi_tot = 0.0, 0.0, 0.0
    for p in periodi_grezzi:
        p_inizio = _parse_data_tfr(p["data_inizio"])
        p_fine = fino_a if not p.get("data_fine") else _parse_data_tfr(p["data_fine"])
        oi, of = max(p_inizio, inizio_eff), min(p_fine, fine_eff)
        if of <= oi:
            continue
        mesi = _mesi_excel(oi, of, 30.416)
        lordo = p["importo_settimanale"] * 52 / 12 / 12 * mesi
        aliquota = p.get("aliquota_tassazione")
        aliquota = ALIQUOTA_TFR if aliquota is None else aliquota
        tot_lordo += lordo
        tot_tass += lordo * aliquota / 100
        mesi_tot += mesi
    return {"lordo": round(tot_lordo, 2), "tassazione": round(tot_tass, 2),
            "netto": round(tot_lordo - tot_tass, 2), "mesi": round(mesi_tot, 2),
            "dal": inizio_eff.strftime("%Y-%m-%d"), "al": fine_eff.strftime("%Y-%m-%d")}


@router.get("/simulazione/{dipendente_id}")
@handle_errors
async def get_simulazione_tfr(dipendente_id: str) -> Dict[str, Any]:
    """Elenca i periodi della simulazione storica TFR di un dipendente (l'eventuale
    periodo aperto è ricalcolato al volo fino a oggi), con i totali e la paga
    settimanale attualmente in corso."""
    db = Database.get_db()
    dipendente = await db["dipendenti"].find_one({"id": dipendente_id}, {"_id": 0})
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")

    fino_a = _fino_a_calcolo(dipendente)
    grezzi = await db["tfr_simulazione_periodi"].find(
        {"dipendente_id": dipendente_id}, {"_id": 0}
    ).sort("data_inizio", 1).to_list(500)
    periodi = [_periodo_con_calcolo_live(p, fino_a) for p in grezzi]

    periodo_aperto = periodi[-1] if periodi and periodi[-1]["aperto"] else None
    if not periodi:
        prossimo_inizio = (dipendente.get("data_assunzione") or "")[:10] or None
    elif periodo_aperto:
        prossimo_inizio = None  # nessun "prossimo": si registra un aumento, non un nuovo periodo in coda
    else:
        ultimo_fine = _parse_data_tfr(periodi[-1]["data_fine"])
        prossimo_inizio = (ultimo_fine + timedelta(days=1)).strftime("%Y-%m-%d")

    return {
        "dipendente_id": dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "periodi": periodi,
        "totale_lordo": round(sum(p["lordo"] for p in periodi), 2),
        "totale_tassazione": round(sum(p["tassazione"] for p in periodi), 2),
        "totale_netto": round(sum(p["netto"] for p in periodi), 2),
        "prossimo_data_inizio": prossimo_inizio,
        "paga_attuale": periodo_aperto["importo_settimanale"] if periodo_aperto else None,
        "paga_attuale_dal": periodo_aperto["data_inizio"] if periodo_aperto else None,
        "cessato": dipendente.get("stato") == "cessato",
        "data_cessazione": (dipendente.get("data_dimissione") or dipendente.get("data_cessazione") or None),
    }


@router.post("/simulazione/{dipendente_id}/periodi")
@handle_errors
async def aggiungi_periodo_simulazione(dipendente_id: str, input_data: PeriodoSimulazioneInput) -> Dict[str, Any]:
    """Aggiunge un periodo alla simulazione. Se non indichi la data di fine, il
    periodo resta APERTO (in corso): è il caso normale per registrare un aumento,
    che chiude da solo il periodo aperto precedente (se c'era) il giorno prima
    della nuova data. Se non indichi la data di inizio: riparte dal giorno dopo
    la fine dell'ultimo periodo chiuso, da oggi se il precedente era aperto
    (è un aumento), o dalla data di assunzione per il primissimo periodo."""
    db = Database.get_db()
    dipendente = await db["dipendenti"].find_one({"id": dipendente_id}, {"_id": 0})
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")

    if input_data.importo_settimanale <= 0:
        raise HTTPException(status_code=400, detail="L'importo settimanale deve essere positivo")

    periodi_esistenti = await db["tfr_simulazione_periodi"].find(
        {"dipendente_id": dipendente_id}, {"_id": 0}
    ).sort("data_inizio", 1).to_list(500)

    ultimo = periodi_esistenti[-1] if periodi_esistenti else None
    ultimo_aperto = bool(ultimo) and not ultimo.get("data_fine")

    if input_data.data_inizio:
        data_inizio_str = input_data.data_inizio[:10]
    elif ultimo_aperto:
        data_inizio_str = _oggi_tfr().strftime("%Y-%m-%d")
    elif ultimo:
        ultimo_fine = _parse_data_tfr(ultimo["data_fine"])
        data_inizio_str = (ultimo_fine + timedelta(days=1)).strftime("%Y-%m-%d")
    elif dipendente.get("data_assunzione"):
        data_inizio_str = dipendente["data_assunzione"][:10]
    else:
        raise HTTPException(
            status_code=400,
            detail="Indica la data di inizio del primo periodo (il dipendente non ha "
                   "una data di assunzione in anagrafica)")

    try:
        data_inizio = _parse_data_tfr(data_inizio_str)
        data_fine = _parse_data_tfr(input_data.data_fine) if input_data.data_fine else None
    except ValueError:
        raise HTTPException(status_code=400, detail="Date non valide, usa il formato YYYY-MM-DD")

    if data_fine and data_fine <= data_inizio:
        raise HTTPException(status_code=400, detail="La data di fine deve essere successiva alla data di inizio")

    if ultimo_aperto:
        ultimo_inizio = _parse_data_tfr(ultimo["data_inizio"])
        if data_inizio <= ultimo_inizio:
            raise HTTPException(
                status_code=400,
                detail=f"La data del nuovo periodo deve essere successiva all'inizio di quello in corso "
                       f"({ultimo['data_inizio']})")
        # Chiude da solo il periodo aperto precedente: finisce il giorno prima del nuovo.
        data_fine_chiusura = data_inizio - timedelta(days=1)
        calc_chiusura = _calcola_periodo_tfr(ultimo_inizio, data_fine_chiusura,
                                             ultimo["importo_settimanale"], ultimo.get("aliquota_tassazione"))
        await db["tfr_simulazione_periodi"].update_one(
            {"id": ultimo["id"]},
            {"$set": {"data_fine": data_fine_chiusura.strftime("%Y-%m-%d"),
                      "chiuso_automaticamente": True, **calc_chiusura}})
    elif ultimo:
        ultimo_fine = _parse_data_tfr(ultimo["data_fine"])
        if data_inizio <= ultimo_fine:
            raise HTTPException(
                status_code=400,
                detail=f"Il periodo si sovrappone: l'ultimo periodo finisce il "
                       f"{ultimo['data_fine']}, questo dovrebbe iniziare dopo")

    aliquota = ALIQUOTA_TFR if input_data.aliquota_tassazione is None else input_data.aliquota_tassazione
    if not (0 <= aliquota < 100):
        raise HTTPException(status_code=400, detail="L'aliquota di tassazione deve essere tra 0 e 100")

    periodo = {
        "id": str(uuid4()),
        "dipendente_id": dipendente_id,
        "data_inizio": data_inizio_str,
        "data_fine": data_fine.strftime("%Y-%m-%d") if data_fine else None,
        "importo_settimanale": round(input_data.importo_settimanale, 2),
        "aliquota_tassazione": round(aliquota, 2),
        "chiuso_automaticamente": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    if data_fine:
        periodo.update(_calcola_periodo_tfr(data_inizio, data_fine, input_data.importo_settimanale, aliquota))
    await db["tfr_simulazione_periodi"].insert_one(periodo.copy())

    return {"success": True, "periodo": _periodo_con_calcolo_live(periodo)}


class ModificaPeriodoInput(BaseModel):
    importo_settimanale: Optional[float] = None
    data_inizio: Optional[str] = None
    data_fine: Optional[str] = None  # ha effetto solo se il periodo era già chiuso
    aliquota_tassazione: Optional[float] = None  # percentuale, es. 23 o 27


@router.put("/simulazione/{dipendente_id}/periodi/{periodo_id}")
@handle_errors
async def modifica_periodo_simulazione(dipendente_id: str, periodo_id: str,
                                       input_data: ModificaPeriodoInput) -> Dict[str, Any]:
    """Corregge un periodo esistente — QUALSIASI, non solo l'ultimo — utile se hai
    sbagliato l'importo settimanale o una data. Ricalcola solo quel periodo, senza
    toccare gli altri. Non cambia se il periodo è aperto o chiuso: se era aperto
    resta aperto (torna a maturare al volo), se era chiuso puoi correggerne anche la
    data di fine. Controlla che non si sovrapponga ai periodi immediatamente
    precedente/successivo."""
    db = Database.get_db()
    tutti = await db["tfr_simulazione_periodi"].find(
        {"dipendente_id": dipendente_id}, {"_id": 0}).sort("data_inizio", 1).to_list(500)
    idx = next((i for i, p in enumerate(tutti) if p["id"] == periodo_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="Periodo non trovato")
    periodo = tutti[idx]
    era_aperto = not periodo.get("data_fine")

    nuovo_importo = (input_data.importo_settimanale if input_data.importo_settimanale is not None
                     else periodo["importo_settimanale"])
    if nuovo_importo <= 0:
        raise HTTPException(status_code=400, detail="L'importo settimanale deve essere positivo")

    nuova_inizio_str = input_data.data_inizio[:10] if input_data.data_inizio else periodo["data_inizio"]
    nuova_fine_str = periodo.get("data_fine")
    if not era_aperto and input_data.data_fine:
        nuova_fine_str = input_data.data_fine[:10]

    try:
        nuova_inizio = _parse_data_tfr(nuova_inizio_str)
        nuova_fine = _parse_data_tfr(nuova_fine_str) if nuova_fine_str else None
    except ValueError:
        raise HTTPException(status_code=400, detail="Date non valide, usa il formato YYYY-MM-DD")

    if nuova_fine and nuova_fine <= nuova_inizio:
        raise HTTPException(status_code=400, detail="La data di fine deve essere successiva alla data di inizio")

    if idx > 0:
        prec_fine = tutti[idx - 1].get("data_fine")
        if prec_fine and nuova_inizio <= _parse_data_tfr(prec_fine):
            raise HTTPException(status_code=400,
                                detail=f"Si sovrapporrebbe al periodo precedente (finisce il {prec_fine})")
    if idx < len(tutti) - 1:
        succ_inizio_str = tutti[idx + 1]["data_inizio"]
        if nuova_fine and nuova_fine >= _parse_data_tfr(succ_inizio_str):
            raise HTTPException(status_code=400,
                                detail=f"Si sovrapporrebbe al periodo successivo (inizia il {succ_inizio_str})")

    nuova_aliquota = (input_data.aliquota_tassazione if input_data.aliquota_tassazione is not None
                      else periodo.get("aliquota_tassazione"))
    nuova_aliquota = ALIQUOTA_TFR if nuova_aliquota is None else nuova_aliquota
    if not (0 <= nuova_aliquota < 100):
        raise HTTPException(status_code=400, detail="L'aliquota di tassazione deve essere tra 0 e 100")

    aggiornamento: Dict[str, Any] = {"data_inizio": nuova_inizio_str,
                                     "importo_settimanale": round(nuovo_importo, 2),
                                     "aliquota_tassazione": round(nuova_aliquota, 2)}
    update_op: Dict[str, Any] = {"$set": aggiornamento}
    if nuova_fine:
        aggiornamento["data_fine"] = nuova_fine_str
        aggiornamento.update(_calcola_periodo_tfr(nuova_inizio, nuova_fine, nuovo_importo, nuova_aliquota))
    else:
        update_op["$unset"] = {"giorni": "", "mesi": "", "lordo": "", "tassazione": "", "netto": ""}

    await db["tfr_simulazione_periodi"].update_one({"id": periodo_id}, update_op)

    dipendente = await db["dipendenti"].find_one({"id": dipendente_id}, {"_id": 0}) or {}
    aggiornato = await db["tfr_simulazione_periodi"].find_one({"id": periodo_id}, {"_id": 0})
    return {"success": True, "periodo": _periodo_con_calcolo_live(aggiornato, _fino_a_calcolo(dipendente))}


@router.delete("/simulazione/{dipendente_id}/periodi/{periodo_id}")
@handle_errors
async def elimina_periodo_simulazione(dipendente_id: str, periodo_id: str) -> Dict[str, Any]:
    """Elimina un periodo dalla simulazione. Consentito solo per l'ULTIMO periodo
    (in ordine di data), per non lasciare buchi nella catena di date. Se
    l'eliminazione scopre un periodo che era stato chiuso automaticamente da
    questo (un aumento annullato), lo riapre."""
    db = Database.get_db()
    periodi = await db["tfr_simulazione_periodi"].find(
        {"dipendente_id": dipendente_id}, {"_id": 0}
    ).sort("data_inizio", 1).to_list(500)
    if not periodi:
        raise HTTPException(status_code=404, detail="Nessun periodo trovato per questo dipendente")
    if periodi[-1]["id"] != periodo_id:
        raise HTTPException(
            status_code=400,
            detail="Puoi eliminare solo l'ultimo periodo inserito (in ordine di data), per non lasciare buchi")
    await db["tfr_simulazione_periodi"].delete_one({"id": periodo_id})

    if len(periodi) >= 2:
        precedente = periodi[-2]
        if precedente.get("chiuso_automaticamente"):
            await db["tfr_simulazione_periodi"].update_one(
                {"id": precedente["id"]},
                {"$set": {"data_fine": None, "chiuso_automaticamente": False},
                 "$unset": {"giorni": "", "mesi": "", "lordo": "", "tassazione": "", "netto": ""}})

    return {"success": True, "messaggio": "Periodo eliminato"}


@router.post("/simulazione/{dipendente_id}/rate")
@handle_errors
async def dividi_in_rate_simulazione(dipendente_id: str, input_data: RateSimulazioneInput) -> Dict[str, Any]:
    """Divide il netto TFR totale della simulazione in N rate (l'ultima assorbe
    l'arrotondamento). Solo calcolo: non registra nulla, così titolare e dipendente
    possono valutare la proposta prima di deciderla."""
    if input_data.numero_rate < 1:
        raise HTTPException(status_code=400, detail="Il numero di rate deve essere almeno 1")

    db = Database.get_db()
    dipendente = await db["dipendenti"].find_one({"id": dipendente_id}, {"_id": 0}) or {}
    grezzi = await db["tfr_simulazione_periodi"].find(
        {"dipendente_id": dipendente_id}, {"_id": 0}).to_list(500)
    if not grezzi:
        raise HTTPException(status_code=400, detail="Nessun periodo inserito nella simulazione")
    periodi = [_periodo_con_calcolo_live(p, _fino_a_calcolo(dipendente)) for p in grezzi]

    totale_netto = round(sum(p["netto"] for p in periodi), 2)
    if totale_netto <= 0:
        raise HTTPException(status_code=400, detail="Il totale netto della simulazione è zero")

    n = input_data.numero_rate
    rata_base = round(totale_netto / n, 2)
    data_rata = _parse_data_tfr(input_data.data_prima_rata) if input_data.data_prima_rata else None

    rate = []
    for i in range(1, n + 1):
        importo = round(totale_netto - rata_base * (n - 1), 2) if i == n else rata_base
        voce = {"numero": i, "importo": importo}
        if data_rata:
            mese_idx = data_rata.month - 1 + (i - 1)
            anno_target = data_rata.year + mese_idx // 12
            mese_target = mese_idx % 12 + 1
            giorno = min(data_rata.day, 28)  # evita overflow sui mesi corti
            voce["data"] = f"{anno_target:04d}-{mese_target:02d}-{giorno:02d}"
        rate.append(voce)

    return {"totale_netto": totale_netto, "numero_rate": n, "rate": rate}


@router.get("/simulazione/{dipendente_id}/liquidazione")
@handle_errors
async def liquidazione_finale_simulazione(dipendente_id: str) -> Dict[str, Any]:
    """Stima di liquidazione finale con la stessa scaletta di periodi del simulatore
    TFR: rateo di tredicesima e quattordicesima maturati (competenza gennaio-dicembre
    per la 13ª, 1° luglio-30 giugno per la 14ª) e controvalore delle ferie residue.
    Se il dipendente è cessato, tutto si ferma alla data di cessazione anziché a oggi
    — stessa logica del periodo aperto che si chiude su un aumento.
    Tredicesima e quattordicesima usano la stessa formula del foglio storico (righe
    "13°"/"14°": mensilità/12 × mesi con divisore 30,416, tassazione con l'aliquota
    del periodo) e restituiscono lordo, tassazione e netto. Le ferie residue riusano
    il dato già tracciato dall'app dai cedolini (dipendenti.ferie_residue), non un
    calcolo parallelo."""
    db = Database.get_db()
    dipendente = await db["dipendenti"].find_one({"id": dipendente_id}, {"_id": 0})
    if not dipendente:
        raise HTTPException(status_code=404, detail="Dipendente non trovato")

    cessato = dipendente.get("stato") == "cessato"
    data_cessazione = (dipendente.get("data_dimissione") or dipendente.get("data_cessazione") or "")[:10]
    fino_a = _fino_a_calcolo(dipendente)

    periodi = await db["tfr_simulazione_periodi"].find(
        {"dipendente_id": dipendente_id}, {"_id": 0}).sort("data_inizio", 1).to_list(500)
    if not periodi:
        raise HTTPException(
            status_code=400,
            detail="Nessun periodo nella simulazione: inserisci prima la paga settimanale in corso")

    data_assunzione = None
    if dipendente.get("data_assunzione"):
        try:
            data_assunzione = _parse_data_tfr(dipendente["data_assunzione"])
        except ValueError:
            data_assunzione = None

    inizio_13, fine_13 = _periodo_competenza_13(fino_a)
    inizio_14, fine_14 = _periodo_competenza_14(fino_a)
    tredicesima = _quota_mensilita_aggiuntiva(periodi, data_assunzione, inizio_13, fine_13, fino_a)
    quattordicesima = _quota_mensilita_aggiuntiva(periodi, data_assunzione, inizio_14, fine_14, fino_a)

    # Ferie: riusa il residuo già tracciato dall'app (da cedolino) — un solo sistema, non un doppione.
    ferie_residue_giorni = dipendente.get("ferie_residue")
    paga_settimanale_attuale = periodi[-1]["importo_settimanale"]
    paga_giornaliera = round(paga_settimanale_attuale / 6, 2)  # CCNL: settimana lavorativa di 6 giorni
    ferie = None
    if ferie_residue_giorni is not None:
        giorni = round(float(ferie_residue_giorni), 2)
        ferie = {
            "giorni_residui": giorni,
            "paga_giornaliera": paga_giornaliera,
            "controvalore": round(giorni * paga_giornaliera, 2),
            "fonte": "residuo tracciato dall'app (da cedolino)",
        }

    return {
        "dipendente_id": dipendente_id,
        "dipendente_nome": dipendente.get("nome_completo", ""),
        "cessato": cessato,
        "data_cessazione": data_cessazione or None,
        "calcolato_fino_a": fino_a.strftime("%Y-%m-%d"),
        "tredicesima": tredicesima,
        "quattordicesima": quattordicesima,
        "ferie": ferie,
    }


@router.post("/simulazione/importa-da-excel")
@handle_errors
async def importa_simulazione_da_excel(
    file: UploadFile = File(...),
    sostituisci: bool = Query(False, description="Sovrascrive i periodi già salvati per i dipendenti trovati nel file"),
) -> Dict[str, Any]:
    """Importa i periodi della simulazione storica TFR dal file Excel 'calcolo ferie e TFR'
    (un foglio per dipendente: colonna B=inizio periodo, C=fine periodo, G=paga settimanale
    anno per anno). Ogni periodo viene RICALCOLATO con la formula di legge (art. 2120 c.c.),
    non con l'approssimazione "una mensilità" del foglio originale.

    Abbinamento foglio → dipendente: SOLO sul nome della scheda (il contenuto della cella A2
    di alcune schede è un'etichetta rimasta da un copia-incolla e non è affidabile — es. le
    schede 'VINCENZO' e 'VALERIO' hanno entrambe A2='CAPEZZUTO'). Si cerca il nome della
    scheda nel foglio 'DATI' (per nome o cognome; se il nome compare come nome E cognome nella
    stessa riga si preferisce quella riga, utile per distinguere due persone con lo stesso
    cognome) per risalire al codice fiscale, poi si cerca quel CF in anagrafica. Se il foglio
    non è nel DATI o il CF non è in anagrafica, si prova un abbinamento diretto per nome/cognome.
    Se l'abbinamento è ambiguo o assente, il foglio viene segnalato e saltato (nessuna scelta
    a caso). Se un dipendente ha già periodi salvati, restano intatti a meno di sostituisci=true."""
    import io
    import re as re_mod
    import openpyxl
    raw = await file.read()
    if raw[:2] != b"PK":
        raise HTTPException(400, "Il file deve essere un .xlsx")
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel non valido: {e}")

    def norm(s):
        return re_mod.sub(r"[^a-z0-9]", "", str(s or "").strip().lower())

    db = Database.get_db()

    # --- Indice dal foglio DATI: nome/cognome ripuliti -> codice fiscale ---
    dati_idx = []
    if "DATI" in wb.sheetnames:
        ws_dati = wb["DATI"]
        for row in ws_dati.iter_rows(min_row=2, max_row=ws_dati.max_row, values_only=True):
            if not row or len(row) < 9:
                continue
            nome, cognome, cf = row[0], row[1], row[8]
            if not (nome or cognome):
                continue
            dati_idx.append({
                "nome_norm": norm(nome), "cognome_norm": norm(cognome),
                "self_ref": norm(nome) == norm(cognome) and norm(nome) != "",
                "cf": (str(cf).strip().upper() if cf else ""),
            })

    def cerca_cf_in_dati(token):
        esatti = [d for d in dati_idx if d["nome_norm"] == token or d["cognome_norm"] == token]
        pool = esatti
        if not pool and len(token) >= 3:
            pool = [d for d in dati_idx if token in d["nome_norm"] or token in d["cognome_norm"]]
        if not pool:
            return None, "nessuna riga DATI corrispondente"
        forti = [d for d in pool if d["self_ref"]]
        scelta = forti if forti else pool
        cf_set = {d["cf"] for d in scelta if d["cf"]}
        if len(cf_set) == 1:
            return next(iter(cf_set)), "ok"
        if len(cf_set) > 1:
            return None, "più righe DATI diverse corrispondono a questo nome (ambiguo)"
        return None, "riga DATI trovata ma senza codice fiscale"

    # --- Anagrafica app: indici per CF e per nome/cognome (fallback) ---
    dips = await db.dipendenti.find(
        {"merged_into": {"$exists": False}},
        {"_id": 0, "id": 1, "nome": 1, "cognome": 1, "nome_completo": 1, "codice_fiscale": 1}
    ).to_list(1000)
    by_cf, by_nome = {}, {}
    bare_counts: Dict[str, int] = {}
    for d in dips:
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for bare in {n, c}:
            if bare:
                bare_counts[bare] = bare_counts.get(bare, 0) + 1
    for d in dips:
        cf = (d.get("codice_fiscale") or "").strip().upper()
        if cf:
            by_cf[cf] = d
        n, c = norm(d.get("nome")), norm(d.get("cognome"))
        for v in {norm(d.get("nome_completo")), c + n, n + c}:
            if v:
                by_nome[v] = d
        # Fallback su singolo nome o cognome (utile per file senza scheda DATI):
        # solo se quel nome/cognome, da solo, identifica un unico dipendente attivo.
        for bare in {n, c}:
            if bare and bare_counts.get(bare) == 1:
                by_nome.setdefault(bare, d)

    def risolvi_dipendente(token):
        cf, msg = cerca_cf_in_dati(token)
        if cf and cf in by_cf:
            return by_cf[cf], f"DATI → CF {cf}"
        dip = by_nome.get(token)
        if dip:
            return dip, "nome (fallback anagrafica)" if not cf else f"nome (fallback: {msg})"
        return None, msg if not cf else "codice fiscale del DATI non presente in anagrafica"

    SHEET_SKIP = {norm(x) for x in ("DATI", "Foglio1", "13", "14", "12")}
    risultati = []
    dipendenti_processati = {}  # dipendente_id -> foglio (per rilevare conflitti nello stesso import)

    for ws in wb.worksheets:
        titolo = ws.title
        tok = norm(titolo)
        if tok in SHEET_SKIP:
            continue

        dip, motivo = risolvi_dipendente(tok)
        if not dip:
            risultati.append({"foglio": titolo, "abbinato": False, "motivo": motivo})
            continue

        nome_dip = dip.get("nome_completo") or f"{dip.get('cognome', '')} {dip.get('nome', '')}".strip()

        if dip["id"] in dipendenti_processati:
            risultati.append({"foglio": titolo, "abbinato": True, "dipendente": nome_dip,
                              "periodi_importati": 0,
                              "nota": f"stesso dipendente già risolto dal foglio "
                                      f"'{dipendenti_processati[dip['id']]}' in questo import: saltato"})
            continue

        # Scaletta anno per anno: righe 2-20, colonne B(inizio) C(fine) G(paga settimanale);
        # H(lordo) e I(tassazione) servono solo a ricavare l'aliquota usata dal foglio su
        # quella riga (23% o 27% a seconda dell'anno) — se mancano, default 23%.
        righe = []
        for r in range(2, 21):
            b, c, g = ws[f"B{r}"].value, ws[f"C{r}"].value, ws[f"G{r}"].value
            if not isinstance(b, (datetime, date)) or not isinstance(c, (datetime, date)):
                continue
            try:
                g = float(g)
            except (TypeError, ValueError):
                continue
            if g <= 0:
                continue
            b = b if isinstance(b, datetime) else datetime(b.year, b.month, b.day)
            c = c if isinstance(c, datetime) else datetime(c.year, c.month, c.day)
            aliquota = ALIQUOTA_TFR
            try:
                h, i_val = float(ws[f"H{r}"].value), float(ws[f"I{r}"].value)
                if h > 0 and 0 <= i_val / h < 1:
                    aliquota = round(i_val / h * 100, 2)
            except (TypeError, ValueError):
                pass
            righe.append((b, c, g, aliquota))

        if not righe:
            risultati.append({"foglio": titolo, "abbinato": True, "dipendente": nome_dip,
                              "motivo_abbinamento": motivo, "periodi_importati": 0,
                              "nota": "nessuna paga settimanale valorizzata nel foglio"})
            continue

        righe.sort(key=lambda x: x[0])

        esistenti = await db.tfr_simulazione_periodi.count_documents({"dipendente_id": dip["id"]})
        if esistenti and not sostituisci:
            risultati.append({"foglio": titolo, "abbinato": True, "dipendente": nome_dip,
                              "motivo_abbinamento": motivo, "periodi_importati": 0,
                              "nota": f"il dipendente ha già {esistenti} periodi salvati: "
                                      f"ripeti con sostituisci=true per sovrascriverli"})
            continue

        if esistenti:
            await db.tfr_simulazione_periodi.delete_many({"dipendente_id": dip["id"]})

        inseriti = 0
        for idx, (b, c, g, aliquota) in enumerate(righe):
            e_ultimo = idx == len(righe) - 1
            periodo = {
                "id": str(uuid4()), "dipendente_id": dip["id"],
                "data_inizio": b.strftime("%Y-%m-%d"),
                # L'ultima riga (la più recente) diventa il periodo APERTO: la data di fine
                # scritta nel file è solo "quando è stato esportato", non una vera chiusura —
                # da qui in avanti il sistema la tiene aggiornata da sola fino ad oggi.
                "data_fine": None if e_ultimo else c.strftime("%Y-%m-%d"),
                "importo_settimanale": round(g, 2),
                "aliquota_tassazione": aliquota,
                "chiuso_automaticamente": False,
                "fonte": "excel_calcolo_ferie_tfr", "created_at": datetime.now(timezone.utc).isoformat(),
            }
            if not e_ultimo:
                periodo.update(_calcola_periodo_tfr(b, c, g, aliquota))
            await db.tfr_simulazione_periodi.insert_one(periodo)
            inseriti += 1

        dipendenti_processati[dip["id"]] = titolo
        risultati.append({"foglio": titolo, "abbinato": True, "dipendente": nome_dip,
                          "motivo_abbinamento": motivo, "periodi_importati": inseriti})

    return {
        "fogli_processati": len(risultati),
        "fogli_abbinati": len([r for r in risultati if r.get("abbinato")]),
        "fogli_non_abbinati": len([r for r in risultati if not r.get("abbinato")]),
        "risultati": risultati,
    }


# ============================================
# CALCOLO NETTO DA LORDO (stipendio mensile, IRPEF 2026)
# ============================================
# Strumento SEPARATO dal simulatore TFR sopra (che usa la formula del foglio
# storico con aliquota fissa per periodo): qui si parte da un importo settimanale
# LORDO per stimare quanto arriva netto in busta con le regole fiscali italiane
# 2026 (INPS 9,19%, scaglioni IRPEF progressivi, detrazione lavoro dipendente).
# Non tocca né legge nulla del simulatore TFR.
#
# Non include: addizionale regionale/comunale (variano per comune/regione) né
# eventuali bonus (es. trattamento integrativo/cuneo fiscale per i redditi più
# bassi) — per il netto esatto in busta paga, verifica con il commercialista.

# Scaglioni IRPEF 2026 (soglia cumulativa, aliquota marginale in quella fascia)
_SCAGLIONI_IRPEF_2026 = [(28000.0, 0.23), (50000.0, 0.33), (float("inf"), 0.43)]
_ALIQUOTA_INPS_DIPENDENTE = 0.0919  # quota INPS a carico del lavoratore


def _irpef_lorda(imponibile: float) -> float:
    """IRPEF a scaglioni PROGRESSIVI (marginale): ogni fascia di reddito paga
    la propria aliquota, non tutto il reddito all'aliquota più alta raggiunta."""
    irpef, residuo, soglia_precedente = 0.0, imponibile, 0.0
    for soglia, aliquota in _SCAGLIONI_IRPEF_2026:
        fascia = min(residuo, soglia - soglia_precedente)
        if fascia <= 0:
            break
        irpef += fascia * aliquota
        residuo -= fascia
        soglia_precedente = soglia
    return irpef


def _detrazione_lavoro_dipendente_2026(reddito: float) -> float:
    """Detrazione lavoro dipendente 2026: piena (1.955€) fino a 15.000€ di
    reddito, poi decrescente in due tratti fino ad azzerarsi a 50.000€."""
    if reddito <= 15000:
        return 1955.0
    if reddito <= 28000:
        return 1910 + 1190 * (28000 - reddito) / 13000
    if reddito <= 50000:
        return 1910 * (50000 - reddito) / 22000
    return 0.0


class CalcoloNettoInput(BaseModel):
    importo_settimanale_lordo: float
    settimane_lavorate: float
    mesi_lavorati: float  # es. 12 per un anno intero, o i mesi effettivi del periodo


@router.post("/calcolo-netto-da-lordo")
@handle_errors
async def calcolo_netto_da_lordo(input_data: CalcoloNettoInput) -> Dict[str, Any]:
    """Da un importo settimanale LORDO: calcola la media mensile lorda del periodo
    (settimane_lavorate × importo_settimanale_lordo / mesi_lavorati), la annualizza
    (×12) per determinare correttamente gli scaglioni IRPEF e la detrazione lavoro
    dipendente, poi applica la stessa aliquota media alla mensilità per ottenere il
    netto mensile. Formula INPS 9,19% + IRPEF 2026 (23/33/43%) + detrazione lavoro
    dipendente 2026 — non include addizionali regionali/comunali né bonus."""
    if input_data.importo_settimanale_lordo <= 0:
        raise HTTPException(status_code=400, detail="L'importo settimanale lordo deve essere positivo")
    if input_data.settimane_lavorate <= 0:
        raise HTTPException(status_code=400, detail="Le settimane lavorate devono essere positive")
    if input_data.mesi_lavorati <= 0:
        raise HTTPException(status_code=400, detail="I mesi lavorati devono essere positivi")

    lordo_periodo = input_data.importo_settimanale_lordo * input_data.settimane_lavorate
    lordo_mensile_medio = lordo_periodo / input_data.mesi_lavorati
    ral_equivalente = lordo_mensile_medio * 12

    contributi_inps = ral_equivalente * _ALIQUOTA_INPS_DIPENDENTE
    imponibile_fiscale = ral_equivalente - contributi_inps
    irpef_lorda = _irpef_lorda(imponibile_fiscale)
    detrazione = _detrazione_lavoro_dipendente_2026(imponibile_fiscale)
    irpef_netta = max(0.0, irpef_lorda - detrazione)

    netto_annuo_equivalente = ral_equivalente - contributi_inps - irpef_netta
    aliquota_media_effettiva = (contributi_inps + irpef_netta) / ral_equivalente if ral_equivalente else 0.0
    netto_mensile = lordo_mensile_medio * (1 - aliquota_media_effettiva)

    return {
        "lordo_periodo": round(lordo_periodo, 2),
        "lordo_mensile_medio": round(lordo_mensile_medio, 2),
        "ral_equivalente_annua": round(ral_equivalente, 2),
        "contributi_inps": round(contributi_inps, 2),
        "imponibile_fiscale": round(imponibile_fiscale, 2),
        "irpef_lorda": round(irpef_lorda, 2),
        "detrazione_lavoro_dipendente": round(detrazione, 2),
        "irpef_netta": round(irpef_netta, 2),
        "aliquota_media_effettiva": round(aliquota_media_effettiva * 100, 2),
        "netto_mensile": round(netto_mensile, 2),
        "netto_periodo": round(netto_mensile * input_data.mesi_lavorati, 2),
    }

